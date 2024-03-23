# import the libraries

import pandas as pd
import yfinance as yf
import openpyxl
import os


def fun (   ticker = 'goog',
            filePath = r"D:\Python Tools\ChartMaker\SourceDocuments\OutPut_Excel\def ",
            sheetName = 'Yahoo Fundamentals'

):

    # Set Varibles
    filePathTicker = filePath + ticker + ".xlsx"
    filePathTemp = filePath + " Temp " + ".xlsx"
    book  = openpyxl.load_workbook(filePathTicker)

    # Download Fundamentals from Yahoo Finance
    try:
        funda = yf.Ticker(ticker).info
        df= pd.DataFrame([funda])
        df = df.set_index('symbol')
        print(df)

    except Exception:
        print(f' Error on Fundamentals : No data!')
        df = pd.DataFrame(columns=['symbol','address1','city','state','zip','country','phone','website','industry','industryDisp','sector',	
                                'longBusinessSummary','fullTimeEmployees','companyOfficers','auditRisk','boardRisk','compensationRisk',
                                'shareHolderRightsRisk','overallRisk','governanceEpochDate','compensationAsOfEpochDate','maxAge','priceHint',
                                'previousClose','open','dayLow','dayHigh','regularMarketPreviousClose','regularMarketOpen','regularMarketDayLow',
                                'regularMarketDayHigh','payoutRatio','beta','trailingPE','forwardPE','volume','regularMarketVolume','averageVolume',
                                'averageVolume10days','averageDailyVolume10Day','bid','ask','bidSize','askSize','marketCap','fiftyTwoWeekLow','fiftyTwoWeekHigh',
                                'priceToSalesTrailing12Months','fiftyDayAverage','twoHundredDayAverage','trailingAnnualDividendRate','trailingAnnualDividendYield',
                                'currency','enterpriseValue','profitMargins','floatShares','sharesOutstanding','sharesShort','sharesShortPriorMonth',
                                'sharesShortPreviousMonthDate','dateShortInterest','sharesPercentSharesOut','heldPercentInsiders','heldPercentInstitutions',
                                'shortRatio','shortPercentOfFloat','impliedSharesOutstanding','bookValue','priceToBook','lastFiscalYearEnd','nextFiscalYearEnd',
                                'mostRecentQuarter','earningsQuarterlyGrowth','netIncomeToCommon','trailingEps','forwardEps','pegRatio','enterpriseToRevenue',
                                'enterpriseToEbitda','52WeekChange','SandP52WeekChange','exchange','quoteType','underlyingSymbol','shortName','longName',
                                'firstTradeDateEpochUtc','timeZoneFullName','timeZoneShortName','uuid','messageBoardId','gmtOffSetMilliseconds','currentPrice',
                                'targetHighPrice','targetLowPrice','targetMeanPrice','targetMedianPrice','recommendationMean','recommendationKey','numberOfAnalystOpinions',
                                'totalCash','totalCashPerShare','ebitda','totalDebt','quickRatio','currentRatio','totalRevenue','debtToEquity','revenuePerShare',
                                'returnOnAssets','returnOnEquity','grossProfits','freeCashflow','operatingCashflow','earningsGrowth','revenueGrowth','grossMargins',
                                'ebitdaMargins','operatingMargins','financialCurrency','trailingPegRatio','lastSplitFactor','lastSplitDate','dividendRate',
                                'dividendYield','exDividendDate','fiveYearAvgDividendYield','lastDividendValue','lastDividendDate'
                                    ]) # need to change this
        

    df.to_excel( filePathTemp , sheetName)
    book2 = openpyxl.load_workbook(filePathTemp).active
    book2._parent = book
    book._add_sheet(book2)
    book.save(filePathTicker)
    os.remove(filePathTemp)




#_____________________________________________________________________________________________________________________________________________________





def fun2 (   ticker = 'goog',
            filePath = r"D:\Python Tools\ChartMaker\SourceDocuments\OutPut_Excel\def ",
            sheetName = 'Yahoo Fundamentals'

):

    # Set Varibles
    filePathTicker = filePath + ticker + ".xlsx"
    filePathTemp = filePath + " Temp " + ".xlsx"
    book  = openpyxl.load_workbook(filePathTicker)

    # Download Fundamentals from Yahoo Finance
    try:
        funda = yf.Ticker(ticker).get_info()
        df= pd.DataFrame([funda])
        # Swap Rows and Columns                 Help Link :- https://note.nkmk.me/en/python-pandas-t-transpose/
        df= df.transpose()
        # Rename column in dataframe :-         Help Link :- https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.rename.html
        df= df.rename(columns={0:"Data"})
        # Add an empty column to a dataframe    Help Link :- https://stackoverflow.com/questions/16327055/how-to-add-an-empty-column-to-a-dataframe
        df["Premarket High"]    = ""
        df["High"]              = ""
        df["Low"]               = ""
        df["Open"]              = ""
        df["Close"]             = ""
        df["Previos_Close"]     = ""
        df["Volume"]            = ""
        df["H_Vol"]             = ""
        df["L_Vol"]             = ""

        print(df)


    except Exception:
        print(f' Error on Fundamentals : No data!')
        index = [
        'address1','address2','city','state','zip','country','phone','website','industry'
        ,'industryKey','industryDisp','sector','sectorKey','sectorDisp','longBusinessSummary','companyOfficers','compensationAsOfEpochDate','maxAge','priceHint'
        ,'previousClose','open','dayLow','dayHigh','regularMarketPreviousClose','regularMarketOpen','regularMarketDayLow','regularMarketDayHigh','forwardPE','volume'
        ,'regularMarketVolume','averageVolume','averageVolume10days','averageDailyVolume10Day','bid','ask','bidSize','askSize','marketCap','fiftyTwoWeekLow','fiftyTwoWeekHigh'
        ,'priceToSalesTrailing12Months','fiftyDayAverage','twoHundredDayAverage','currency','enterpriseValue','profitMargins','floatShares','sharesOutstanding','sharesShort'
        ,'sharesShortPriorMonth','sharesShortPreviousMonthDate','dateShortInterest','sharesPercentSharesOut','heldPercentInsiders','heldPercentInstitutions','shortRatio'
        ,'shortPercentOfFloat','impliedSharesOutstanding','bookValue','priceToBook','lastFiscalYearEnd','nextFiscalYearEnd','mostRecentQuarter','netIncomeToCommon'
        ,'trailingEps','forwardEps','lastSplitFactor','lastSplitDate','enterpriseToRevenue','enterpriseToEbitda','52WeekChange','SandP52WeekChange','exchange'
        ,'quoteType','symbol','underlyingSymbol','shortName','longName','firstTradeDateEpochUtc','timeZoneFullName','timeZoneShortName','uuid','messageBoardId'
        ,'gmtOffSetMilliseconds','currentPrice','targetHighPrice','targetLowPrice','targetMeanPrice','targetMedianPrice','recommendationKey','numberOfAnalystOpinions'
        ,'totalCash','totalCashPerShare','ebitda','totalDebt','quickRatio','currentRatio','totalRevenue','debtToEquity','revenuePerShare','returnOnAssets'
        ,'returnOnEquity','freeCashflow','operatingCashflow','revenueGrowth','grossMargins','ebitdaMargins','operatingMargins','financialCurrency','trailingPegRatio'
        ]

        df = pd.DataFrame(columns=['INDEX','Data','PremarketHigh','High','Low','Open','Close','Previos_Close','Volume','H_Vol','L_Vol' ])
        df.INDEX = index

    df.to_excel( filePathTemp , sheetName)
    book2 = openpyxl.load_workbook(filePathTemp).active
    book2._parent = book
    book._add_sheet(book2)
    book.save(filePathTicker)
    os.remove(filePathTemp)




