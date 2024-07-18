

import pandas as pd
import openpyxl
import time
import os

def fun(    ticker = 'goog',
            filePath = r"D:\Python Tools\ChartMaker\SourceDocuments\OutPut_Excel\def ",
            sheetName = 'Samary'
        ):
    
    filePathTicker = filePath + ticker + ".xlsx"
    df = pd.DataFrame(columns=['Symbol', 'Premarket_High','High','Low','Open','Close', 'Previos_Close','Volume',
                               'H_Vol','L_Vol', 'Relative_Volume','Average_Volume','Shortable_Shares',
                               'gap','PH_pst','PH_O','Max_Gain','intra_Range','Market_Cap','Float',
                               'Sector','Industry','Country','Watch_List','Zamzam_Effect','My_Interaction',
                               'File','Daily','Min_5','Min_1',
                               'Sec0920_0930','Sec0930_0940' ,'Sec0940_0950','Sec0950_1000','Sec1000_1010',
                               'Sec1010_1020','Sec1020_1030','Sec1030_1040','Sec1040_1050','Sec1050_1100','Sec1100_1110' 
                                
                               ]) # need to change this
    df.to_excel( filePathTicker , sheetName)

#______________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________                    

def funYahooImport (filePath_fun, bookName = "Yahoo 1m") :

    # Set Variable    
    intraDayList = ['Yahoo Hours','Yahoo 30m','Yahoo 15m','Yahoo 5m','Yahoo 2m','Yahoo 1m']
    swingDayList = ['Yahoo Dayes']
    
    # Import  data
    #------------------------------------------------  
    book = openpyxl.load_workbook(filePath_fun)
    sheet = book[bookName]
    df= pd.DataFrame(sheet.values)


    # format  data
    #------------------------------------------------  
    #Convert row to column header for Pandas DataFrame : Link https://stackoverflow.com/questions/26147180/convert-row-to-column-header-for-pandas-dataframe
    df.columns = df.iloc[0]
    df = df.drop(0)

    #change object to datetime64[ns]
    if (bookName in intraDayList):
        df.Datetime = pd.to_datetime(df.Datetime.astype('datetime64[ns]'))
        df = df.set_index('Datetime')
    elif (bookName in swingDayList):
        df.Date = pd.to_datetime(df.Date.astype('datetime64[ns]'))
        df = df.set_index('Date')

    #change object to float
    # Link https://stackoverflow.com/questions/36814100/pandas-to-numeric-for-multiple-columns
    cols = df.columns
    df[cols] = df[cols].apply(pd.to_numeric, errors='coerce')


    return df

#______________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________                    

def funIBKRimport (filePath_fun, bookName = 'IBKR 1m') :

    # Set Variable    
    intraDayList = ['IBKR 1s','IBKR 5s','IBKR 10s','IBKR 15s','IBKR 30s',
                    'IBKR 1m','IBKR 2m','IBKR 3m','IBKR 5m','IBKR 10m','IBKR 15m','IBKR 20m','IBKR 30m',
                    'IBKR 1H','IBKR 2H','IBKR 3H','IBKR 4H','IBKR 8H' ]
    swingDayList = ['IBKR 1Day','IBKR 1week','IBKR 1Month']
    
    # Import  data
    #------------------------------------------------  
    book = openpyxl.load_workbook(filePath_fun)
    sheet = book[bookName]
    df= pd.DataFrame(sheet.values)


    # format  data
    #------------------------------------------------  
    #Convert row to column header for Pandas DataFrame : Link https://stackoverflow.com/questions/26147180/convert-row-to-column-header-for-pandas-dataframe
    df.columns = df.iloc[0]
    df = df.drop(0)

    #change object to datetime64[ns]
    if (bookName in intraDayList):
        df.Datetime = df.Datetime.str.replace("US/Eastern", "")
        df.Datetime = pd.to_datetime(df.Datetime.astype('datetime64[ns]'))
        df = df.set_index('Datetime')
    elif (bookName in swingDayList):
        df.Datetime = pd.to_datetime(df.Datetime.astype('datetime64[ns]'))
        df = df.set_index('Datetime')

    #change object to float
    # Link https://stackoverflow.com/questions/36814100/pandas-to-numeric-for-multiple-columns
    cols = df.columns
    df[cols] = df[cols].apply(pd.to_numeric, errors='coerce')


    return df

#______________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________                    

def funSamaryImport (filePath_fun, bookName = 'Samary') :
    # Import  data
    #------------------------------------------------  
    book = openpyxl.load_workbook(filePath_fun)
    sheet = book[bookName]
    df= pd.DataFrame(sheet.values)
    # format  data
    #------------------------------------------------  
    #Convert row to column header for Pandas DataFrame : Link https://stackoverflow.com/questions/26147180/convert-row-to-column-header-for-pandas-dataframe
    df.columns = df.iloc[0]
    df = df.drop(0)
    df.reset_index(inplace=True)
    df = df.drop(columns=['index', None])
    return df

#______________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________                    

def funSamaryUpdate(     ticker = 'TGL'
                        ,fExcelPath = r"C:\Users\lenovo\Desktop\TGL\Session [20240323_051432]\TGL\OK TGL.xlsx"
                        ,sheetName = 'Samary'
                    ):    
    print("Update the Samary Sheet for Ticker:- ",ticker)
    fExcelPath = fExcelPath + ticker + ".xlsx"

    dfS = funSamaryImport (filePath_fun = fExcelPath ,bookName = sheetName)          # "Samary"
    dfD = funYahooImport  (filePath_fun = fExcelPath ,bookName = "Yahoo Dayes" )    # "Yahoo Dayes"
    dfT = funIBKRimport   (filePath_fun = fExcelPath ,bookName = 'IBKR 30m')        # "IBKR 30m"  
            
    EDay = str( dfD.index[-1])
    ETime=EDay[0:11] + "09:00:00"  
    STime=EDay[0:11] + "04:00:00"
    df = dfT.loc[STime:ETime] #XXXXXXXX

    #__________________________________________________________________________________________________                    
    Premarket_High  = df['High'].max()
    Premarket__Low  = df['Low' ].min()
    H_Vol           = df['Volume' ].sum()
    L_Vol           = 0
    
    Volume          = dfT['CumVol'][-1]

    High            = dfD['High' ][-1]
    Low             = dfD['Low'  ][-1] 
    Open            = dfD['Open' ][-1] 
    Close           = dfD['Close'][-1] 
    Previos_Close   = dfD['Close'][-2] 
    
    #__________________________________________________________________________________________________ 
    # Add one row in an existing Pandas DataFrame   Help Link :- https://www.geeksforgeeks.org/how-to-add-one-row-in-an-existing-pandas-dataframe/
    dfN =   {
                 "Symbol"           : ticker
                ,"Premarket_High"   : Premarket_High
                ,"High"             : High
                ,"Low"              : Low
                ,"Open"             : Open
                ,"Close"            : Close
                ,"Previos_Close"    : Previos_Close
                ,"Volume"           : Volume
                ,"H_Vol"            : H_Vol
                ,"L_Vol"            : L_Vol
                # ,"" : 
            }
    dfS = dfS.append(dfN, ignore_index = True) 

    

    #__________________________________________________________________________________________________ 
    # Save DataFrame and replace the Old Sheet      Help Link :-    https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.ExcelWriter.html
    with pd.ExcelWriter(
                        fExcelPath,
                        mode="a",
                        engine="openpyxl",
                        if_sheet_exists="replace",
                        ) as writer:
                            dfS.to_excel(writer, sheet_name=sheetName)  
    
    #__________________________________________________________________________________________________ 
    # print("Premarket_High is :- ",Premarket_High)
    # print("High is :-           ",High)
    # print("Low is :-            ",Low)
    # print("Open is :-           ",Open)
    # print("Close is :-          ",Close)
    # print("Previos_Close is :-  ",Previos_Close)
    # print("Volume is :-         ",Volume)
    # print("H_Vol is :-          ",H_Vol)
    # print("L_Vol is :-          ",L_Vol)
    # # print(" is :-",)
    # print(dfS)

# funSamaryUpdate()
#______________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________                    


                            




#______________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________________                    

def funSamaryUpdate2(     ticker = 'TGL'
                        ,fExcelPath = r"C:\Users\lenovo\Desktop\TGL\Session [20240323_051432]\TGL\OK TGL.xlsx"
                        ,sheetName = 'Samary'
                    ):    
    print("Update the Samary Sheet for Ticker:- ",ticker)
    fExcelPath = fExcelPath + ticker + ".xlsx"

    dfS = funSamaryImport (filePath_fun = fExcelPath ,bookName = sheetName)         # "Samary"
    dfF = funYahooImport  (filePath_fun = fExcelPath ,bookName = "Yahoo Fundamentals" )    # "Yahoo Fundamentals"
    dfD = funYahooImport  (filePath_fun = fExcelPath ,bookName = "Yahoo Dayes" )    # "Yahoo Dayes"
    dfT = funIBKRimport   (filePath_fun = fExcelPath ,bookName = 'IBKR 30m')        # "IBKR 30m"
    dfM = funIBKRimport   (filePath_fun = fExcelPath ,bookName = "IBKR 1m")         # "IBKR 1m"  
            
    EDay = str( dfD.index[-1])
    ETime=EDay[0:11] + "09:00:00"  
    STime=EDay[0:11] + "04:00:00"
    df = dfT.loc[STime:ETime] #XXXXXXXX

    #__________________________________________________________________________________________________                    
    Premarket_High  = df['High'].max()
    Premarket__Low  = df['Low' ].min()
    H_Vol           = df['Volume' ].sum()
    L_Vol           = 0
    
    Volume          = dfT['CumVol'][-1]

    High            = dfD['High' ][-1]
    Low             = dfD['Low'  ][-1] 
    Open            = dfD['Open' ][-1] 
    Close           = dfD['Close'][-1] 
    Previos_Close   = dfD['Close'][-2] 
    
    #__________________________________________________________________________________________________ 
    # Add one row in an existing Pandas DataFrame   Help Link :- https://www.geeksforgeeks.org/how-to-add-one-row-in-an-existing-pandas-dataframe/
    # dfN =   {
    #              "Symbol"           : ticker
    #             ,"Premarket_High"   : Premarket_High
    #             ,"High"             : High
    #             ,"Low"              : Low
    #             ,"Open"             : Open
    #             ,"Close"            : Close
    #             ,"Previos_Close"    : Previos_Close
    #             ,"Volume"           : Volume
    #             ,"H_Vol"            : H_Vol
    #             ,"L_Vol"            : L_Vol
    #             # ,"" : 
    #         }
    

    #__________________________________________________________________________________________________ 
    # stylRead = (StyRow.Value[StyRow['Attribute'] == AttributeList[AttributeIndex]].values[0])
    Z_Date                  = EDay
    Z_Ticker                = (dfF.Value[dfF['Data'] == "symbol"].values[0])
    Z_Company_Name          = (dfF.Value[dfF['Data'] == "shortName"].values[0])
    Z_Activity_Sector       = (dfF.Value[dfF['Data'] == "sector"].values[0])
    Z_Issuer_Country        = (dfF.Value[dfF['Data'] == "country"].values[0])
    Z_Market_Cap_Million    = (dfF.Value[dfF['Data'] == "marketCap"].values[0])
    Z_Float_Million         = (dfF.Value[dfF['Data'] == "floatShares"].values[0])
    Z_Short_Interest        = (dfF.Value[dfF['Data'] == "shortRatio"].values[0])        # Need to consult Rabi3
    Z_ATR                   = "0"                                                       # Need to consult Rabi3
    
    Z_Recent_Reverse_Split  = int (dfF.Value[dfF['Data'] == "lastSplitDate"].values[0])
    Z_Recent_Reverse_Split  = time.strftime("%Y/%m/%d ", time.localtime(Z_Recent_Reverse_Split))


    #__________________________________________________________________________________________________ 
    # Add one row in an existing Pandas DataFrame   Help Link :- https://www.geeksforgeeks.org/how-to-add-one-row-in-an-existing-pandas-dataframe/
    dfN =   {
                 "Date"                             :   Z_Date  
                ,"Ticker"                           :   Z_Ticker
                ,"Company Name"                     :   Z_Company_Name
                ,"Activity - Sector"                :   Z_Activity_Sector
                ,"Issuer Country"                   :   Z_Issuer_Country
                ,"Market Cap (Million)"             :	Z_Market_Cap_Million
                ,"Float (Million)"                  :	Z_Float_Million
                ,"Short Interest %"                 :	Z_Short_Interest
                ,"ATR"                              :	Z_ATR
                ,"Recent Reverse Split"             :	Z_Recent_Reverse_Split
                ,"Active Shelf Registration"        :	Z_Active_Shelf_Registration
                ,"Catalyst"                         :	Z_Catalyst
                ,"Pre Runner"                       :	Z_Pre_Runner
                ,"Previous Day High"                :	Z_Previous_Day_High
                ,"Previous Day Close"               :	Z_Previous_Day_Close
                ,"4:00am price"                     :	Z_price_0400am
                ,"Low price before first move"      :	Z_Low_price_before_first_move
                ,"Low time before first move"       :	Z_Low_time_before_first_move
                ,"First Dip %"                      :	Z_First_Dip
                ,"PreMarket High Level"             :	Z_PreMarket_High_Level
                ,"PreMarket High Time"              :	Z_PreMarket_High_Time
                ,"Market Open Level"                :	Z_Market_Open_Level
                ,"Gap %"                            :	Z_Gap
                ,"IntraDay High Level"              :	Z_IntraDay_High_Level
                ,"IntraDay High Time"               :	Z_IntraDay_High_Time
                ,"Low price after IntraDay High"    :	Z_Low_price_after_IntraDay_High
                ,"Low time after IntraDay High"     :	Z_Low_time_after_IntraDay_High
                ,"Sell off %"                       :	Z_Sell_off
                ,"Day Close Level"                  :	Z_Day_Close_Level
                ,"After Hours High Level"           :	Z_After_Hours_High_Level
                ,"After Hours High Time"            :	Z_After_Hours_High_Time
                ,"Max % Gain (From Previous Day)"   :	Z_Max_Gain__From_Previous_Day
                ,"Max % Gain (From Market Open)"    :	Z_Max_Gain__From_Market_Open
                ,"Halts to the up side"             :	Z_Halts_to_the_up_side
                ,"Halts to the down side"           :	Z_Halts_to_the_down_side
                ,"Volume (7am)"                     :	Z_Volume_0700am
                ,"7:00 am float rotation"           :	Z_float_rotation_0700_am
                ,"Volume (9:30am)"                  :	Z_Volume_0930am
                ,"9:30 am float rotation"           :	Z_float_rotation_0930_am 
                ,"Volume (Full Day)"                :	Z_Volume_Full_Day
                ,"full day float rotation"          :	Z_float_rotation_full_day 
                ,"7am move"                         :	Z_move_0700_am 
                ,"PreMarket move (Non 7am)"         :	Z_PreMarket_move_Non_0700_am
                ,"9:30am move"                      :	Z_move_0930_am 
                ,"IntraDay move (After 10am)"       :	Z_IntraDay_move_After_1000_am
                ,"After Hours move"                 :	Z_After_Hours_move
                ,"5min breakouts"                   :	Z_breakouts_5min 
                ,"1min breakouts"                   :	Z_breakouts_1min 
                ,"General Comment"                  :	Z_General_Comment
                ,"Chart Screenshot"                 :	Z_Chart_Screenshot

            }
    

    dfS = dfS.append(dfN, ignore_index = True) 

    

    #__________________________________________________________________________________________________ 
    # Save DataFrame and replace the Old Sheet      Help Link :-    https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.ExcelWriter.html
    with pd.ExcelWriter(
                        fExcelPath,
                        mode="a",
                        engine="openpyxl",
                        if_sheet_exists="replace",
                        ) as writer:
                            dfS.to_excel(writer, sheet_name=sheetName)  
    
    #__________________________________________________________________________________________________ 
    # print("Premarket_High is :- ",Premarket_High)
    # print("High is :-           ",High)
    # print("Low is :-            ",Low)
    # print("Open is :-           ",Open)
    # print("Close is :-          ",Close)
    # print("Previos_Close is :-  ",Previos_Close)
    # print("Volume is :-         ",Volume)
    # print("H_Vol is :-          ",H_Vol)
    # print("L_Vol is :-          ",L_Vol)
    # # print(" is :-",)
    # print(dfS)






















