



import GetChart


import pandas as pd
import numpy as np
import math
import ImportData
import openpyxl
import time
import os

from datetime import datetime,date, timedelta













#--------------------------------------------------------------------------------------------------------------------------

def fun6(    filePath           = r"C:\Users\lenovo\Desktop\TGL\BCG\OK BCG.xlsx"
            ,timeStampPath      = r"D:\Python Tools\ChartMaker\SourceDocuments\InPut_Excel\Time_Stamp.xlsx"
            ,timeStampSheet     = "Ver1"
            ,DaySheetName       = "IBKR 1Day" 
            ,scope              =  "5m" # ["5s","1m","5m","30m","1h","1d"]  
            ,scopeBook          = {
                                     "5s"       :"5s"
                                    ,'5 secs'   :"5s"
                                    ,"1m"       :"1m"
                                    ,'1 min'    :"1m"
                                    ,"5m"       :"5m"
                                    ,'5 mins'   :"5m"
                                    ,"30m"      :"30m"
                                    ,'30 mins'  :"30m"
                                    ,"1h"       :"1h"
                                    ,'1 hour'   :"1h"
                                    ,"1d"       :"1d" 
                                    ,'1 day'    :"1d" 
                                } 
      
        ):
            #----------------------------------------------------------------------------------------------------------------------------------------------     
            # Date sheet:- Import data and change index to string     Help Link :-    https://stackoverflow.com/questions/44741587/pandas-timestamp-series-to-string
            #   1-Import data
            dfD = ImportData.fun2 (filePath_fun = filePath , bookName = DaySheetName) # "IBKR 1Day"
            
            if(len(dfD.index)==0):
                if(DaySheetName == "IBKR 1Day"):
                    DaySheetName = "Yahoo Dayes"
                    dfD = ImportData.fun (filePath_fun = filePath , bookName = DaySheetName )   # "Yahoo Dayes"
                elif(DaySheetName == "Yahoo Dayes"):
                    DaySheetName = "IBKR 1Day"
                    dfD = ImportData.fun2 (filePath_fun = filePath , bookName = DaySheetName )   # "Yahoo Dayes"
                else:
                    print("Wrong Day Sheet Name")
                # dfD = ImportData.fun2 (filePath_fun = filePath , bookName = DaySheetName )   # "Yahoo Dayes"
            #   2-Change index Type to string
            dfD.index = dfD.index.astype(str)
            #----------------------------------------------------------------------------------------------------------------------------------------------     



            #----------------------------------------------------------------------------------------------------------------------------------------------     
            # Time Stamp sheet:- Import  data and Set column header 
            #   1-Import  data
            book = openpyxl.load_workbook(timeStampPath)
            sheet = book[timeStampSheet]
            dfS= pd.DataFrame(sheet.values)
            #   2-Set column header :- Convert row to column header for Pandas DataFrame : Link https://stackoverflow.com/questions/26147180/convert-row-to-column-header-for-pandas-dataframe
            dfS.columns = dfS.iloc[0]
            dfS = dfS.drop(0)
            #----------------------------------------------------------------------------------------------------------------------------------------------     


            
            #----------------------------------------------------------------------------------------------------------------------------------------------     
            # OutPut Dataframe "taro":-     #   01- Create dataframe :-
            #-------------------------------#   02- Rename column in dataframe :- 
            #-------------------------------#   03- Clean None Rows :-
            #-------------------------------#   04- Change object to float :-
            #-------------------------------#   05- Add date to index colum :-
            #-------------------------------#   06- Change type of index colum to Time stamp :-
            #-------------------------------#   07- Return OutPut Dataframe "taro" :-
            #   01- Create dataframe :- from "dfS" another dataframe    Help Link :- https://www.statology.org/pandas-create-dataframe-from-existing-dataframe/
            taro = dfS[[    
                                         ("Index1_"         + scopeBook[scope])
                                        ,("Index2_"         + scopeBook[scope])
                                        ,("Name_"           + scopeBook[scope])
                                        ,("StartDelta_"     + scopeBook[scope])
                                        ,("EndDelta_"       + scopeBook[scope])
                                        ,("Fibonacci_"      + scopeBook[scope])
                                        ,("Scope_"          + scopeBook[scope])
                                        ,("HourConstant_"   + scopeBook[scope])
                                        ,("MinConstant_"    + scopeBook[scope])
                                        ,("SecConstant_"    + scopeBook[scope])
                                        ,("leftSide_"       + scopeBook[scope])
                                        ,("Blank_"          + scopeBook[scope])
                                        ]].copy()                        
            #   02- Rename column in dataframe :-      Help Link :- https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.rename.html
            taro = taro.rename(columns={                                                                  
                                         "Index1_"         + scopeBook[scope]    :"Index1"
                                        ,"Index2_"         + scopeBook[scope]    :"Index2"
                                        ,"Name_"           + scopeBook[scope]    :"Name"
                                        ,"StartDelta_"     + scopeBook[scope]    :"StartDelta"
                                        ,"EndDelta_"       + scopeBook[scope]    :"EndDelta"
                                        ,"Fibonacci_"      + scopeBook[scope]    :"Fibonacci"
                                        ,"Scope_"          + scopeBook[scope]    :"Scope"
                                        ,"HourConstant_"   + scopeBook[scope]    :"HourConstant"
                                        ,"MinConstant_"    + scopeBook[scope]    :"MinConstant"
                                        ,"SecConstant_"    + scopeBook[scope]    :"SecConstant"
                                        ,"leftSide_"       + scopeBook[scope]    :"leftSide"
                                        ,"Blank_"          + scopeBook[scope]    :"Date"
                                        })
            #   03- Clean None Rows :- Help Link :- https://www.digitalocean.com/community/tutorials/pandas-dropna-drop-null-na-values-from-dataframe    ,   https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.dropna.html
            taro = taro.dropna(how='all')
            taro.reset_index(inplace=True, drop=True)
            #   04- Change object to float :-   Help Link https://stackoverflow.com/questions/36814100/pandas-to-numeric-for-multiple-columns
            cols = taro.columns.drop(["Index1","Index2","Name","Fibonacci","Scope","Date"])   # set colums you dont want to change         
            taro[cols] = taro[cols].apply(pd.to_numeric, errors='coerce')            
            #   05- Add date to index colum :-            
            taro["Date"] = dfD.index[taro.EndDelta]
            taro.Index2 = taro.Date + " " + taro.Index2                        
            for x in range (len(taro)): 
                # IF taro.StartDelta out of bounds for dfD.index set the firist Day is the start of dfD.index    
                if(len(dfD.index) >= abs(taro.loc[x].StartDelta)): 
                    # print(">>>>_______________________________<<<<")
                    # print(dfD.index[taro.loc[x].StartDelta])    
                    taro.Date[x] = dfD.index[taro.loc[x].StartDelta]      
                else:                       
                    taro.Date[x] =    dfD.index[0]                                    
            taro.Index1 = taro.Date + " " + taro.Index1
            #   06- Change type of index colum to Time stamp :-            
            print(">>>>$$$$$$$$$$$$$$<<<<")
            print(dfD)
            print(taro)
            print(">>>>$$$$$$$$$$$$$$<<<<")
            taro.Index1 = pd.to_datetime(taro.Index1.astype('datetime64[ns]'))
            taro.Index2 = pd.to_datetime(taro.Index2.astype('datetime64[ns]'))            
            #   07- Return OutPut Dataframe "taro" :-
            return taro
            #----------------------------------------------------------------------------------------------------------------------------------------------     


#--------------------------------------------------------------------------------------------------------------------------








            # OutPut Dataframe "taro":-     #   01- Create dataframe :-
            #-------------------------------#   02- Rename column in dataframe :- *****
            #-------------------------------#   03- Clean None Rows :-
            #-------------------------------#   04- Change object to float :-
            #-------------------------------#   05- Add date to index colum :- *****
            #-------------------------------#   06- Change type of index colum to Time stamp :- *****
            #-------------------------------#   07- Return OutPut Dataframe "taro" :- *****



# 1 Get Style of Chart
# 2 Get Day Sheet Data Frame
# 3 Get Time Stamp Data Frame




# #Set Style:
# # link : C:\Users\lenovo\anaconda3\Lib\site-packages\mplfinance\_styledata\mike.py
# EhabStaylo = dict(  style_name    = 'Ehab_Staylo',
#                     base_mpl_style= 'dark_background', 
#                     marketcolors  = {'candle'  : {'up':'#14CE1C', 'down':'#CE1414'},
#                                       'edge'    : {'up':'#14CE1C', 'down':'#CE1414'},
#                                       'wick'    : {'up':'#ffffff', 'down':'#ffffff'},
#                                       'ohlc'    : {'up':'#ffffff', 'down':'#ffffff'},
#                                       'volume'  : {'up':'#14CE1C', 'down':'#CE1414'},
#                                       'vcdopcod': False, # Volume Color Depends On Price Change On Day
#                                       'alpha'   : 2.0,
#                                     },
#                     mavcolors     = ['#ec009c','#78ff8f','#fcf120'],
#                     y_on_right    = True,
#                     gridcolor     = None,
#                     gridstyle     = None,
#                     facecolor     = 'Black',
#                     figcolor      = 'gray',
#                     rc            = [ ('axes.edgecolor'  , 'white'   ),
#                                       ('axes.linewidth'  ,  1.5      ),
#                                       ('axes.labelsize'  , 'large'   ),
#                                       ('axes.labelweight', 'semibold'),
#                                       ('axes.grid'       , True      ),
#                                       ('axes.grid.axis'  , 'both'    ),
#                                       ('axes.grid.which' , 'major'   ),
#                                       ('grid.alpha'      ,  0.9      ),
#                                       ('grid.color'      , '#EBEE24' ),
#                                       ('grid.linestyle'  , ':'      ),
#                                       ('grid.linewidth'  ,  1.0      ),
#                                       ('figure.facecolor', '#0a0a0a' ),
#                                       ('patch.linewidth' ,  1.0      ),
#                                       ('lines.linewidth' ,  1.0      ),
#                                       ('font.weight'     , 'medium'  ),
#                                       ('font.size'       ,  8.0     ),
#                                       ('figure.titlesize', 'x-large' ),
#                                       ('figure.titleweight','semibold'),
#                                     ],
#                     base_mpf_style= 'mike'
#                   )
















#--------------------------------------------------------------------------------------------------------------------------
def fGetDayDataFrame(    
                        filePath           = r"C:\Users\lenovo\Desktop\TGL\BCG\OK BCG.xlsx"            
                        ,DaySheetName       = "IBKR 1Day"                   
                    ):
            #----------------------------------------------------------------------------------------------------------------------------------------------     
            # Date sheet:- Import data and change index to string     Help Link :-    https://stackoverflow.com/questions/44741587/pandas-timestamp-series-to-string
            #   1-Import data
            dfD = ImportData.fun2 (filePath_fun = filePath , bookName = DaySheetName) # "IBKR 1Day"
            
            if(len(dfD.index)==0):
                if(DaySheetName == "IBKR 1Day"):
                    DaySheetName = "Yahoo Dayes"
                    dfD = ImportData.fun (filePath_fun = filePath , bookName = DaySheetName )   # "Yahoo Dayes"
                elif(DaySheetName == "Yahoo Dayes"):
                    DaySheetName = "IBKR 1Day"
                    dfD = ImportData.fun2 (filePath_fun = filePath , bookName = DaySheetName )   # "Yahoo Dayes"
                else:
                    print("Wrong Day Sheet Name")
                # dfD = ImportData.fun2 (filePath_fun = filePath , bookName = DaySheetName )   # "Yahoo Dayes"
            #   2-Change index Type to string
            dfD.index = dfD.index.astype(str)
            #----------------------------------------------------------------------------------------------------------------------------------------------     
            return dfD
#--------------------------------------------------------------------------------------------------------------------------

#--------------------------------------------------------------------------------------------------------------------------
def fGetTimeStampDataFrame(    
            timeStampPath      = r"D:\Python Tools\ChartMaker\SourceDocuments\InPut_Excel\Watch_List_Zamzam .xlsx"
            ,timeStampSheet     = "Sheet2"                  
        ):

            #----------------------------------------------------------------------------------------------------------------------------------------------     
            # Time Stamp sheet:- Import  data and Set column header 
            #   1-Import  data
            book = openpyxl.load_workbook(timeStampPath)
            sheet = book[timeStampSheet]
            dfS= pd.DataFrame(sheet.values)
            #   2-Set column header :- Convert row to column header for Pandas DataFrame : Link https://stackoverflow.com/questions/26147180/convert-row-to-column-header-for-pandas-dataframe
            dfS.columns = dfS.iloc[0]
            dfS = dfS.drop(0)
            #----------------------------------------------------------------------------------------------------------------------------------------------     
            return dfS
#--------------------------------------------------------------------------------------------------------------------------

#--------------------------------------------------------------------------------------------------------------------------
def fun7(       # for Time Stamp
             dfD
            ,dfS            
            ,scope              =  "5m" # ["5s","1m","5m","30m","1h","1d"]  
            ,scopeBook          = {
                                     "5s"       :"5s"
                                    ,'5 secs'   :"5s"
                                    ,"1m"       :"1m"
                                    ,'1 min'    :"1m"
                                    ,"5m"       :"5m"
                                    ,'5 mins'   :"5m"
                                    ,"30m"      :"30m"
                                    ,'30 mins'  :"30m"
                                    ,"1h"       :"1h"
                                    ,'1 hour'   :"1h"
                                    ,"1d"       :"1d" 
                                    ,'1 day'    :"1d" 
                                } 
      
        ):
                        
            #----------------------------------------------------------------------------------------------------------------------------------------------     
            # OutPut Dataframe "taro":-     #   01- Create dataframe :-
            #-------------------------------#   02- Rename column in dataframe :- 
            #-------------------------------#   03- Clean None Rows :-
            #-------------------------------#   04- Change object to float :-
            #-------------------------------#   05- Add date to index colum :-
            #-------------------------------#   06- Change type of index colum to Time stamp :-
            #-------------------------------#   07- Return OutPut Dataframe "taro" :-
            #   01- Create dataframe :- from "dfS" another dataframe    Help Link :- https://www.statology.org/pandas-create-dataframe-from-existing-dataframe/
            taro = dfS[[    
                                         ("Index1_"         + scopeBook[scope])
                                        ,("Index2_"         + scopeBook[scope])
                                        ,("Name_"           + scopeBook[scope])
                                        ,("StartDelta_"     + scopeBook[scope])
                                        ,("EndDelta_"       + scopeBook[scope])
                                        ,("Fibonacci_"      + scopeBook[scope])
                                        ,("Scope_"          + scopeBook[scope])
                                        ,("HourConstant_"   + scopeBook[scope])
                                        ,("MinConstant_"    + scopeBook[scope])
                                        ,("SecConstant_"    + scopeBook[scope])
                                        ,("leftSide_"       + scopeBook[scope])
                                        ,("Blank_"          + scopeBook[scope])
                                        ]].copy()                        
            #   02- Rename column in dataframe :-      Help Link :- https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.rename.html
            taro = taro.rename(columns={                                                                  
                                         "Index1_"         + scopeBook[scope]    :"Index1"
                                        ,"Index2_"         + scopeBook[scope]    :"Index2"
                                        ,"Name_"           + scopeBook[scope]    :"Name"
                                        ,"StartDelta_"     + scopeBook[scope]    :"StartDelta"
                                        ,"EndDelta_"       + scopeBook[scope]    :"EndDelta"
                                        ,"Fibonacci_"      + scopeBook[scope]    :"Fibonacci"
                                        ,"Scope_"          + scopeBook[scope]    :"Scope"
                                        ,"HourConstant_"   + scopeBook[scope]    :"HourConstant"
                                        ,"MinConstant_"    + scopeBook[scope]    :"MinConstant"
                                        ,"SecConstant_"    + scopeBook[scope]    :"SecConstant"
                                        ,"leftSide_"       + scopeBook[scope]    :"leftSide"
                                        ,"Blank_"          + scopeBook[scope]    :"Date"
                                        })
            #   03- Clean None Rows :- Help Link :- https://www.digitalocean.com/community/tutorials/pandas-dropna-drop-null-na-values-from-dataframe    ,   https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.dropna.html
            taro = taro.dropna(how='all')
            taro.reset_index(inplace=True, drop=True)
            #   04- Change object to float :-   Help Link https://stackoverflow.com/questions/36814100/pandas-to-numeric-for-multiple-columns
            cols = taro.columns.drop(["Index1","Index2","Name","Fibonacci","Scope","Date"])   # set colums you dont want to change         
            taro[cols] = taro[cols].apply(pd.to_numeric, errors='coerce')            
            #   05- Add date to index colum :-            
            taro["Date"] = dfD.index[taro.EndDelta]
            taro.Index2 = taro.Date + " " + taro.Index2                        
            for x in range (len(taro)): 
                # IF taro.StartDelta out of bounds for dfD.index set the firist Day is the start of dfD.index    
                if(len(dfD.index) >= abs(taro.loc[x].StartDelta)): 
                    # print(">>>>_______________________________<<<<")
                    # print(dfD.index[taro.loc[x].StartDelta])    
                    taro.Date[x] = dfD.index[taro.loc[x].StartDelta]      
                else:                       
                    taro.Date[x] =    dfD.index[0]                                    
            taro.Index1 = taro.Date + " " + taro.Index1
            #   06- Change type of index colum to Time stamp :-            
            print(">>>>$$$$$$$$$$$$$$<<<<")
            print(dfD)
            print(taro)
            print(">>>>$$$$$$$$$$$$$$<<<<")
            taro.Index1 = pd.to_datetime(taro.Index1.astype('datetime64[ns]'))
            taro.Index2 = pd.to_datetime(taro.Index2.astype('datetime64[ns]'))            
            #   07- Return OutPut Dataframe "taro" :-
            return taro
            #----------------------------------------------------------------------------------------------------------------------------------------------     
#--------------------------------------------------------------------------------------------------------------------------

#--------------------------------------------------------------------------------------------------------------------------
def fGetStyle(                 
             dfS
            ,AttributeIndex=0
            ,AttributeList =[
                 "style_name"
                ,"base_mpl_style"
                ,"marketcolors"
                ,"candle_Up"
                ,"candle_Down"
                ,"edge_Up"
                ,"edge_Down"
                ,"wick_Up"
                ,"wick_Down"
                ,"ohlc_Up"
                ,"ohlc_Down"
                ,"volume_Up"
                ,"volume_Down"
                ,"vcdopcod"
                ,"alpha"
                ,"mavcolors"
                ,"y_on_right"
                ,"gridcolor"
                ,"gridstyle"
                ,"facecolor"
                ,"figcolor"
                ,"rc"
                ,"axes.edgecolor"
                ,"axes.linewidth"
                ,"axes.labelsize"
                ,"axes.labelweight"
                ,"axes.grid"
                ,"axes.grid.axis"
                ,"axes.grid.which"
                ,"grid.alpha"
                ,"grid.color"
                ,"grid.linestyle"
                ,"grid.linewidth"
                ,"figure.facecolor"
                ,"patch.linewidth"
                ,"lines.linewidth"
                ,"font.weight"
                ,"font.size"
                ,"figure.titlesize"
                ,"figure.titleweight"
                ,"base_mpf_style"]
            ,AttributeDic = {
                                 "style_name"           :	"Ehab_Staylo"
                                ,"base_mpl_style"       :	"dark_background"
                                ,"marketcolors"         :	"marketcolors"
                                ,"candle_Up"            :	"#14CE1C"
                                ,"candle_Down"          :	"#CE1414"
                                ,"edge_Up"              :	"#14CE1C"
                                ,"edge_Down"            :	"#CE1414"
                                ,"wick_Up"              :	"#ffffff"
                                ,"wick_Down"            :	"#ffffff"
                                ,"ohlc_Up"              :	"#ffffff"
                                ,"ohlc_Down"            :	"#ffffff"
                                ,"volume_Up"            :	"#14CE1C"
                                ,"volume_Down"          :	"#CE1414"
                                ,"vcdopcod"             :	 False	
                                ,"alpha"                :	 2.00	
                                ,"mavcolors"            :	 ['#ec009c','#78ff8f','#fcf120']	
                                ,"y_on_right"           :	 True	
                                ,"gridcolor"            :	 None	
                                ,"gridstyle"            :	 None	
                                ,"facecolor"            :	"Black"
                                ,"figcolor"             :	"gray"
                                ,"rc"                   :	"rc"
                                ,"axes.edgecolor"       :	"white"
                                ,"axes.linewidth"       :	 1.50	
                                ,"axes.labelsize"       :	"large"
                                ,"axes.labelweight"     :	"semibold"
                                ,"axes.grid"            :	 True	
                                ,"axes.grid.axis"       :	"both"
                                ,"axes.grid.which"      :	"major"
                                ,"grid.alpha"           :	 0.90	
                                ,"grid.color"           :	"#EBEE24"
                                ,"grid.linestyle"       :	":"	
                                ,"grid.linewidth"       :	 1.00	
                                ,"figure.facecolor"     :	"#0a0a0a"
                                ,"patch.linewidth"      :	 1.00	
                                ,"lines.linewidth"      :	 1.00	
                                ,"font.weight"          :	"medium"
                                ,"font.size"            :	 8.00	
                                ,"figure.titlesize"     :	"x-large"
                                ,"figure.titleweight"   :	"semibold"
                                ,"base_mpf_style"       :	"mike"
                                }


        ):
                        
            

            #--------------------------------------------------------------------------------------------------------------------------
            #   01- Create dataframe :- from "dfS" another dataframe    Help Link :- https://www.statology.org/pandas-create-dataframe-from-existing-dataframe/
            StyRow = dfS[[    
                             ("Attribute")
                            ,("Value")
                        ]].copy()                        
            
            #   02- Clean None Rows :- Help Link :- https://www.digitalocean.com/community/tutorials/pandas-dropna-drop-null-na-values-from-dataframe    ,   https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.dropna.html
            StyRow = StyRow.dropna(how='all')
            StyRow.reset_index(inplace=True, drop=True)


            #--------------------------------------------------------------------------------------------------------------------------
            #   03- Read Data from StyRow
            for AttributeIndex in range(len(AttributeList)):

                stylIndex = AttributeList[AttributeIndex]
                stylRead = (StyRow.Value[StyRow['Attribute'] == AttributeList[AttributeIndex]].values[0])
                AttributeDic[stylIndex] = stylRead

                # print(stylIndex , "=" , stylRead  ,"--->", type(stylRead) )
                # print(">>>>|||||||||||||||<<<<")


            #--------------------------------------------------------------------------------------------------------------------------
            #   03- Cerat Staylo Dict
                    #Set Style:
                    # link : C:\Users\lenovo\anaconda3\Lib\site-packages\mplfinance\_styledata\mike.py
            Staylo =     dict(  style_name    = AttributeDic["style_name"],
                                base_mpl_style= AttributeDic["base_mpl_style"], 
                                marketcolors  = {'candle'   : {'up':AttributeDic["candle_Up"]   ,'down':AttributeDic["candle_Down"] },
                                                  'edge'    : {'up':AttributeDic["edge_Up"]     ,'down':AttributeDic["edge_Down"]   },
                                                  'wick'    : {'up':AttributeDic["wick_Up"]     ,'down':AttributeDic["wick_Down"]   },
                                                  'ohlc'    : {'up':AttributeDic["ohlc_Up"]     ,'down':AttributeDic["ohlc_Down"]   },
                                                  'volume'  : {'up':AttributeDic["volume_Up"]   ,'down':AttributeDic["volume_Down"] },
                                                  'vcdopcod':       AttributeDic["vcdopcod"], # Volume Color Depends On Price Change On Day
                                                  'alpha'   :       AttributeDic["alpha"],
                                                },
                                mavcolors     = AttributeDic["mavcolors"].split(","),  #    Help Link:- https://www.w3schools.com/python/ref_string_split.asp
                                y_on_right    = AttributeDic["y_on_right"],
                                gridcolor     = None if AttributeDic["gridcolor"] == 'None' else AttributeDic["gridcolor"], # None AttributeDic["gridcolor"],     Help Link:-   https://stackoverflow.com/questions/26481774/pythonic-way-to-convert-the-string-none-to-a-proper-none
                                gridstyle     = None if AttributeDic["gridstyle"] == 'None' else AttributeDic["gridstyle"], # None AttributeDic["gridstyle"],     Help Link:-   https://stackoverflow.com/questions/26481774/pythonic-way-to-convert-the-string-none-to-a-proper-none    
                                facecolor     = AttributeDic["facecolor"],
                                figcolor      = AttributeDic["figcolor"],
                                rc            = [ ('axes.edgecolor'     , AttributeDic["axes.edgecolor"]        ),
                                                  ('axes.linewidth'     , AttributeDic["axes.linewidth"]        ),
                                                  ('axes.labelsize'     , AttributeDic["axes.labelsize"]        ),
                                                  ('axes.labelweight'   , AttributeDic["axes.labelweight"]      ),
                                                  ('axes.grid'          , AttributeDic["axes.grid"]             ),
                                                  ('axes.grid.axis'     , AttributeDic["axes.grid.axis"]        ),
                                                  ('axes.grid.which'    , AttributeDic["axes.grid.which"]       ),
                                                  ('grid.alpha'         , AttributeDic["grid.alpha"]            ),
                                                  ('grid.color'         , AttributeDic["grid.color"]            ),
                                                  ('grid.linestyle'     , AttributeDic["grid.linestyle"]        ),
                                                  ('grid.linewidth'     , AttributeDic["grid.linewidth"]        ),
                                                  ('figure.facecolor'   , AttributeDic["figure.facecolor"]      ),
                                                  ('patch.linewidth'    , AttributeDic["patch.linewidth"]       ),
                                                  ('lines.linewidth'    , AttributeDic["lines.linewidth"]       ),
                                                  ('font.weight'        , AttributeDic["font.weight"]           ),
                                                  ('font.size'          , AttributeDic["font.size"]             ),
                                                  ('figure.titlesize'   , AttributeDic["figure.titlesize"]      ),
                                                  ('figure.titleweight' , AttributeDic["figure.titleweight"]    ),
                                                ],
                                base_mpf_style= AttributeDic["base_mpf_style"]
                              )

            
            #--------------------------------------------------------------------------------------------------------------------------            
            #   07- Return OutPut  "Staylo" :-

            if(False):
                print(">>>>**************<<<<")
                print(StyRow.loc[StyRow['Attribute'] == "base_mpl_style"])
                print(">>>>**************<<<<")

                print(">>>>$$$$$$$$$$$$$$<<<<")
                print(dfS)
                print(StyRow)
                print(">>>>$$$$$$$$$$$$$$<<<<")

                print(">>>>*******1*******<<<<")
                stylYest = (StyRow.Value[StyRow['Attribute'] == "gridcolor"].values[0])
                print(stylYest, type(stylYest), type(None))
                print(">>>>*******2*******<<<<")
                print(AttributeDic)
                print(">>>>*******3*******<<<<")
                print(Staylo)
                print(">>>>**************<<<<")

            return Staylo
            #----------------------------------------------------------------------------------------------------------------------------------------------     





def fGetStyle2(                 
             dfS
            ,ShowOutPut = False
            ,AttributeIndex=0
            ,AttributeList =[
                 "style_name"
                ,"base_mpl_style"
                ,"marketcolors"
                ,"candle_Up"
                ,"candle_Down"
                ,"edge_Up"
                ,"edge_Down"
                ,"wick_Up"
                ,"wick_Down"
                ,"ohlc_Up"
                ,"ohlc_Down"
                ,"volume_Up"
                ,"volume_Down"
                ,"vcdopcod"
                ,"alpha"
                ,"mavcolors"
                ,"y_on_right"
                ,"gridcolor"
                ,"gridstyle"
                ,"facecolor"
                ,"figcolor"
                ,"rc"
                ,"axes.edgecolor"
                ,"axes.linewidth"
                ,"axes.labelsize"
                ,"axes.labelweight"
                ,"axes.grid"
                ,"axes.grid.axis"
                ,"axes.grid.which"
                ,"grid.alpha"
                ,"grid.color"
                ,"grid.linestyle"
                ,"grid.linewidth"
                ,"figure.facecolor"
                ,"patch.linewidth"
                ,"lines.linewidth"
                ,"font.weight"
                ,"font.size"
                ,"figure.titlesize"
                ,"figure.titleweight"
                ,"base_mpf_style"]
            ,AttributeDic = {
                                 "style_name"           :	"Ehab_Staylo"
                                ,"base_mpl_style"       :	"dark_background"
                                ,"marketcolors"         :	"marketcolors"
                                ,"candle_Up"            :	"#14CE1C"
                                ,"candle_Down"          :	"#CE1414"
                                ,"edge_Up"              :	"#14CE1C"
                                ,"edge_Down"            :	"#CE1414"
                                ,"wick_Up"              :	"#ffffff"
                                ,"wick_Down"            :	"#ffffff"
                                ,"ohlc_Up"              :	"#ffffff"
                                ,"ohlc_Down"            :	"#ffffff"
                                ,"volume_Up"            :	"#14CE1C"
                                ,"volume_Down"          :	"#CE1414"
                                ,"vcdopcod"             :	 False	
                                ,"alpha"                :	 2.00	
                                ,"mavcolors"            :	 ['#ec009c','#78ff8f','#fcf120']	
                                ,"y_on_right"           :	 True	
                                ,"gridcolor"            :	 None	
                                ,"gridstyle"            :	 None	
                                ,"facecolor"            :	"Black"
                                ,"figcolor"             :	"gray"
                                ,"rc"                   :	"rc"
                                ,"axes.edgecolor"       :	"white"
                                ,"axes.linewidth"       :	 1.50	
                                ,"axes.labelsize"       :	"large"
                                ,"axes.labelweight"     :	"semibold"
                                ,"axes.grid"            :	 True	
                                ,"axes.grid.axis"       :	"both"
                                ,"axes.grid.which"      :	"major"
                                ,"grid.alpha"           :	 0.90	
                                ,"grid.color"           :	"#EBEE24"
                                ,"grid.linestyle"       :	":"	
                                ,"grid.linewidth"       :	 1.00	
                                ,"figure.facecolor"     :	"#0a0a0a"
                                ,"patch.linewidth"      :	 1.00	
                                ,"lines.linewidth"      :	 1.00	
                                ,"font.weight"          :	"medium"
                                ,"font.size"            :	 8.00	
                                ,"figure.titlesize"     :	"x-large"
                                ,"figure.titleweight"   :	"semibold"
                                ,"base_mpf_style"       :	"mike"
                                }


        ):
                        
            

            #--------------------------------------------------------------------------------------------------------------------------
            #   01- Create dataframe :- from "dfS" another dataframe    Help Link :- https://www.statology.org/pandas-create-dataframe-from-existing-dataframe/
            StyRow = dfS[[    
                             ("Attribute")
                            ,("Value")
                        ]].copy()                        
            
            #   02- Clean None Rows :- Help Link :- https://www.digitalocean.com/community/tutorials/pandas-dropna-drop-null-na-values-from-dataframe    ,   https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.dropna.html
            StyRow = StyRow.dropna(how='all')
            StyRow.reset_index(inplace=True, drop=True)


            #--------------------------------------------------------------------------------------------------------------------------
            #   03- Read Data from StyRow
            timeScope = (StyRow.Value[StyRow['Attribute'] == 'timeScope'].values[0]).split(",")
            if(ShowOutPut): print(timeScope)

            for AttributeIndex in range(len(AttributeList)):
                stylIndex = AttributeList[AttributeIndex]
                stylRead = (StyRow.Value[StyRow['Attribute'] == AttributeList[AttributeIndex]].values[0])
                AttributeDic[stylIndex] = stylRead
                if(ShowOutPut):
                    print(stylIndex , "=" , stylRead  ,"--->", type(stylRead) )
                    print(">>>>|||||||||||||||<<<<")


            #--------------------------------------------------------------------------------------------------------------------------
            #   03- Cerat Staylo Dict
                    #Set Style:
                    # link : C:\Users\lenovo\anaconda3\Lib\site-packages\mplfinance\_styledata\mike.py
            Staylo =     dict(  style_name    = AttributeDic["style_name"],
                                base_mpl_style= AttributeDic["base_mpl_style"], 
                                marketcolors  = {'candle'   : {'up':AttributeDic["candle_Up"]   ,'down':AttributeDic["candle_Down"] },
                                                  'edge'    : {'up':AttributeDic["edge_Up"]     ,'down':AttributeDic["edge_Down"]   },
                                                  'wick'    : {'up':AttributeDic["wick_Up"]     ,'down':AttributeDic["wick_Down"]   },
                                                  'ohlc'    : {'up':AttributeDic["ohlc_Up"]     ,'down':AttributeDic["ohlc_Down"]   },
                                                  'volume'  : {'up':AttributeDic["volume_Up"]   ,'down':AttributeDic["volume_Down"] },
                                                  'vcdopcod':       AttributeDic["vcdopcod"], # Volume Color Depends On Price Change On Day
                                                  'alpha'   :       AttributeDic["alpha"],
                                                },
                                mavcolors     = AttributeDic["mavcolors"].split(","),  #    Help Link:- https://www.w3schools.com/python/ref_string_split.asp
                                y_on_right    = AttributeDic["y_on_right"],
                                gridcolor     = None if AttributeDic["gridcolor"] == 'None' else AttributeDic["gridcolor"], # None AttributeDic["gridcolor"],     Help Link:-   https://stackoverflow.com/questions/26481774/pythonic-way-to-convert-the-string-none-to-a-proper-none
                                gridstyle     = None if AttributeDic["gridstyle"] == 'None' else AttributeDic["gridstyle"], # None AttributeDic["gridstyle"],     Help Link:-   https://stackoverflow.com/questions/26481774/pythonic-way-to-convert-the-string-none-to-a-proper-none    
                                facecolor     = AttributeDic["facecolor"],
                                figcolor      = AttributeDic["figcolor"],
                                rc            = [ ('axes.edgecolor'     , AttributeDic["axes.edgecolor"]        ),
                                                  ('axes.linewidth'     , AttributeDic["axes.linewidth"]        ),
                                                  ('axes.labelsize'     , AttributeDic["axes.labelsize"]        ),
                                                  ('axes.labelweight'   , AttributeDic["axes.labelweight"]      ),
                                                  ('axes.grid'          , AttributeDic["axes.grid"]             ),
                                                  ('axes.grid.axis'     , AttributeDic["axes.grid.axis"]        ),
                                                  ('axes.grid.which'    , AttributeDic["axes.grid.which"]       ),
                                                  ('grid.alpha'         , AttributeDic["grid.alpha"]            ),
                                                  ('grid.color'         , AttributeDic["grid.color"]            ),
                                                  ('grid.linestyle'     , AttributeDic["grid.linestyle"]        ),
                                                  ('grid.linewidth'     , AttributeDic["grid.linewidth"]        ),
                                                  ('figure.facecolor'   , AttributeDic["figure.facecolor"]      ),
                                                  ('patch.linewidth'    , AttributeDic["patch.linewidth"]       ),
                                                  ('lines.linewidth'    , AttributeDic["lines.linewidth"]       ),
                                                  ('font.weight'        , AttributeDic["font.weight"]           ),
                                                  ('font.size'          , AttributeDic["font.size"]             ),
                                                  ('figure.titlesize'   , AttributeDic["figure.titlesize"]      ),
                                                  ('figure.titleweight' , AttributeDic["figure.titleweight"]    ),
                                                ],
                                base_mpf_style= AttributeDic["base_mpf_style"]
                              )

            
            #--------------------------------------------------------------------------------------------------------------------------            
            #   07- Return OutPut  "Staylo" :-

            if(ShowOutPut):
                print(">>>>**************<<<<")
                print(StyRow.loc[StyRow['Attribute'] == "base_mpl_style"])
                print(">>>>**************<<<<")

                print(">>>>$$$$$$$$$$$$$$<<<<")
                print(dfS)
                print(StyRow)
                print(">>>>$$$$$$$$$$$$$$<<<<")

                print(">>>>*******1*******<<<<")
                stylYest = (StyRow.Value[StyRow['Attribute'] == "gridcolor"].values[0])
                print(stylYest, type(stylYest), type(None))
                print(">>>>*******2*******<<<<")
                print(AttributeDic)
                print(">>>>*******3*******<<<<")
                print(Staylo)
                print(">>>>**************<<<<")

            return Staylo,timeScope
            #----------------------------------------------------------------------------------------------------------------------------------------------     







#--------------------------------------------------------------------------------------------------------------------------

###_____________________________________________________________________________________________________________________________________________________________________________________
def fun8(   # For Get Chart 
            # Set Varibles
            #------------------------------------------------
                 tickerName = 'AAPL'                    
                ,filePathExcel = r"D:\Python Tools\ChartMaker\SourceDocuments\OutPut_Excel\Session [20230921_221745]\AAPL\OK AAPL.xlsx"
                ,filePathChart = r'D:\Python Tools\ChartMaker\SourceDocuments\OutPut_jpg'            #C:\Users\lenovo\Desktop\Python Project\Ehab\Results\Chart test.jpg
                
                ,filePathWL = r"D:\Python Tools\ChartMaker\SourceDocuments\InPut_Excel\Watch_List_Zamzam .xlsx"
                ,timeStampSheet = "Sheet2"
                
                ,flagYahoo = False
                ,flagIBKR = True
                ,imageType= '.png'
        ):
            ###___________________________________________________________________________________________________________________________
            if (flagYahoo):
                # start the Function:-
                rxChartPath = CreateFile.fun(directory = "Yahoo", parent_dir = filePathChart)

                # Loop 1:-
                daySheet = "Yahoo Dayes"
                indexScope = 0
                bookScope = ["Yahoo 1m","Yahoo 5m","Yahoo 30m","Yahoo Hours","Yahoo Dayes"]
                timeScope = ["1m","5m","30m","1h","1d"] 

                for indexScope in range(len(bookScope)):
                    # Import  data
                    dft = ImportData.fun (filePath_fun = filePathExcel , bookName = bookScope[indexScope])
                    indexTS =TS.fun2(filePath=filePathExcel, bookScope=daySheet, scope=timeScope[indexScope])

                    # Loop 2:-
                    index = 0
                    for index in range(len(indexTS.index1)):
                        # Sampling data
                        df = dft[(indexTS.index1[index]):(indexTS.index2[index])]

                        #Make Indicator
                        EMA = addIndicator.fun(dataFrame = df, scope = timeScope[indexScope] )

                        #Make Yellow Color Candel 
                        mco = YellowCandel.fun(dataFrame = df) 

                        #Make a fibonacci lines
                        fabalines = FibonacciLines.fun (filePath_fun = filePathExcel)

                        #Saving plot to a file -> Link: https://github.com/matplotlib/mplfinance/blob/master/examples/savefig.ipynb
                        setName = '\\' + tickerName + '_' + timeScope[indexScope] + '_' + indexTS.index4[index] + '_to_' + indexTS.index5[index]
                        save = dict(fname= (rxChartPath + setName + imageType), dpi= 300, pad_inches= 0)

                        #Make a chart
                        mpf.plot(data                   = df,
                                title                  = (tickerName + ' ' + timeScope[indexScope]), 
                                type                   = 'candle', 
                                #mav                    = (20,50),
                                volume                 = True,
                                show_nontrading        = False,
                                tight_layout           = True,
                                figratio               = (1,1),
                                figscale               = 3,
                                scale_padding          = 0.30,
                                figsize               = (12,6),
                                #ylim                   = (((df.Low.min())*0.95) ,((df.High.max())*1.05)), # set min and max of Chart
                                ylim                   = (((df.Low.min())-0.03) ,((df.High.max())+0.03)), # set min and max of Chart
                                xrotation              = 0,
                                yscale                 = "linear", # y-axis scale: "linear", "log", "symlog", or "logit"
                                volume_yscale          = "linear", # Volume y-axis scale: "linear", "log", "symlog", or "logit"
                                style                  = EhabStaylo,
                                marketcolor_overrides  = mco,
                                mco_faceonly           = False,
                                addplot                = EMA,
                                hlines                 = fabalines,
                                savefig                = save # link: https://github.com/matplotlib/mplfinance/blob/master/examples/savefig.ipynb
                                #marketcolor_overrides = mco  # link https://github.com/matplotlib/mplfinance/blob/master/examples/marketcolor_overrides.ipynb
                                )
            ###___________________________________________________________________________________________________________________________                                                
                        
            if (flagIBKR):
                # start the Function:-

            #------------------------------------------------
                # Create File:        
                        # datetime object containing current date and time --- Help Link :- https://www.programiz.com/python-programming/datetime/current-datetime#google_vignette
                now = datetime.now()
                dt_string = now.strftime("[%Y%m%d_%H%M%S]") # YYmmdd_HHMMSS
                sessionSymbol =  "zzIBKR" + ' ' + str(dt_string)
                rxChartPath = CreateFile.fun(directory = sessionSymbol, parent_dir = filePathChart)

            #------------------------------------------------
                # Loop 1:-
                Fobi_Var        = FibonacciLines.FiboClass()
                imgCount        = 0
                daySheet        = "IBKR 1Day"
                indexScope      =  '5 secs'         
                timeScope       = ['5 secs','1 min','5 mins','30 mins','1 hour','1 day']      #   ['5 secs','1 min','5 mins','30 mins','1 hour','1 day']
                bookScope       = {
                                    '1 secs' :{'SizeIndex': 0, 'DurationIndex':1, 'DurationLenth': 1, 'Rolling':60, 'ChartDayDuration':  -1, 'indexTS': "5s", 'SheetName':'IBKR 1s'     },
                                    '5 secs' :{'SizeIndex': 1, 'DurationIndex':1, 'DurationLenth': 2, 'Rolling':60, 'ChartDayDuration':  -1, 'indexTS': "5s", 'SheetName':'IBKR 5s'     },
                                    '10 secs':{'SizeIndex': 2, 'DurationIndex':1, 'DurationLenth': 2, 'Rolling':30, 'ChartDayDuration':  -1, 'indexTS': "5s", 'SheetName':'IBKR 10s'    },
                                    '15 secs':{'SizeIndex': 3, 'DurationIndex':1, 'DurationLenth': 2, 'Rolling':20, 'ChartDayDuration':  -1, 'indexTS': "5s", 'SheetName':'IBKR 15s'    },
                                    '30 secs':{'SizeIndex': 4, 'DurationIndex':1, 'DurationLenth': 2, 'Rolling':10, 'ChartDayDuration':  -1, 'indexTS': "5s", 'SheetName':'IBKR 30s'    },
                                    '1 min'  :{'SizeIndex': 5, 'DurationIndex':1, 'DurationLenth': 2, 'Rolling': 5, 'ChartDayDuration':  -1, 'indexTS': "1m", 'SheetName':"IBKR 1m"     },
                                    '2 mins' :{'SizeIndex': 6, 'DurationIndex':1, 'DurationLenth': 2, 'Rolling': 5, 'ChartDayDuration':  -1, 'indexTS': "1m", 'SheetName':'IBKR 2m'     },
                                    '3 mins' :{'SizeIndex': 7, 'DurationIndex':1, 'DurationLenth': 2, 'Rolling':10, 'ChartDayDuration':  -1, 'indexTS': "1m", 'SheetName':'IBKR 3m'     },
                                    '5 mins' :{'SizeIndex': 8, 'DurationIndex':1, 'DurationLenth': 3, 'Rolling': 6, 'ChartDayDuration':  -1, 'indexTS': "5m", 'SheetName':'IBKR 5m'     },
                                    '10 mins':{'SizeIndex': 9, 'DurationIndex':1, 'DurationLenth': 9, 'Rolling': 6, 'ChartDayDuration':  -6, 'indexTS':"30m", 'SheetName':'IBKR 10m'    },
                                    '15 mins':{'SizeIndex':10, 'DurationIndex':1, 'DurationLenth': 9, 'Rolling':12, 'ChartDayDuration':  -6, 'indexTS':"30m", 'SheetName':'IBKR 15m'    },
                                    '20 mins':{'SizeIndex':11, 'DurationIndex':1, 'DurationLenth':12, 'Rolling':12, 'ChartDayDuration':  -6, 'indexTS':"30m", 'SheetName':'IBKR 20m'    },
                                    '30 mins':{'SizeIndex':12, 'DurationIndex':1, 'DurationLenth':12, 'Rolling':11, 'ChartDayDuration':  -6, 'indexTS':"30m", 'SheetName':'IBKR 30m'    },
                                    '1 hour' :{'SizeIndex':13, 'DurationIndex':1, 'DurationLenth':34, 'Rolling':16, 'ChartDayDuration': -21, 'indexTS': "1h", 'SheetName':'IBKR 1H'     },
                                    '2 hours':{'SizeIndex':14, 'DurationIndex':1, 'DurationLenth':34, 'Rolling': 8, 'ChartDayDuration': -21, 'indexTS': "1h", 'SheetName':'IBKR 2H'     },
                                    '3 hours':{'SizeIndex':15, 'DurationIndex':1, 'DurationLenth':34, 'Rolling':30, 'ChartDayDuration': -21, 'indexTS': "1h", 'SheetName':'IBKR 3H'     },
                                    '4 hours':{'SizeIndex':16, 'DurationIndex':1, 'DurationLenth':34, 'Rolling':20, 'ChartDayDuration': -21, 'indexTS': "1h", 'SheetName':'IBKR 4H'     },
                                    '8 hours':{'SizeIndex':17, 'DurationIndex':1, 'DurationLenth':34, 'Rolling':42, 'ChartDayDuration': -21, 'indexTS': "1h", 'SheetName':'IBKR 8H'     },
                                    '1 day'  :{'SizeIndex':18, 'DurationIndex':4, 'DurationLenth': 5, 'Rolling':21, 'ChartDayDuration':-220, 'indexTS': "1d", 'SheetName':'IBKR 1Day'   },
                                    '1W'     :{'SizeIndex':19, 'DurationIndex':4, 'DurationLenth': 5, 'Rolling':12, 'ChartDayDuration':-220, 'indexTS': "1d", 'SheetName':'IBKR 1week'  },
                                    '1M'     :{'SizeIndex':20, 'DurationIndex':4, 'DurationLenth': 5, 'Rolling':12, 'ChartDayDuration':-220, 'indexTS': "1d", 'SheetName':'IBKR 1Month' }
                                    }

                dfTimeStamp     = TS.fGetTimeStampDataFrame(    
                                                            timeStampPath      = filePathWL
                                                            ,timeStampSheet     = timeStampSheet                 
                                                            )
                ChartStyle      = TS.fGetStyle(                 
                                                dfTimeStamp
                                                ,AttributeIndex=0
                                                ,AttributeList =[
                                                                "style_name"
                                                                ,"base_mpl_style"
                                                                ,"marketcolors"
                                                                ,"candle_Up"
                                                                ,"candle_Down"
                                                                ,"edge_Up"
                                                                ,"edge_Down"
                                                                ,"wick_Up"
                                                                ,"wick_Down"
                                                                ,"ohlc_Up"
                                                                ,"ohlc_Down"
                                                                ,"volume_Up"
                                                                ,"volume_Down"
                                                                ,"vcdopcod"
                                                                ,"alpha"
                                                                ,"mavcolors"
                                                                ,"y_on_right"
                                                                ,"gridcolor"
                                                                ,"gridstyle"
                                                                ,"facecolor"
                                                                ,"figcolor"
                                                                ,"rc"
                                                                ,"axes.edgecolor"
                                                                ,"axes.linewidth"
                                                                ,"axes.labelsize"
                                                                ,"axes.labelweight"
                                                                ,"axes.grid"
                                                                ,"axes.grid.axis"
                                                                ,"axes.grid.which"
                                                                ,"grid.alpha"
                                                                ,"grid.color"
                                                                ,"grid.linestyle"
                                                                ,"grid.linewidth"
                                                                ,"figure.facecolor"
                                                                ,"patch.linewidth"
                                                                ,"lines.linewidth"
                                                                ,"font.weight"
                                                                ,"font.size"
                                                                ,"figure.titlesize"
                                                                ,"figure.titleweight"
                                                                ,"base_mpf_style"]
                                                ,AttributeDic = {
                                                                "style_name"           :	"Ehab_Staylo"
                                                                ,"base_mpl_style"       :	"dark_background"
                                                                ,"marketcolors"         :	"marketcolors"
                                                                ,"candle_Up"            :	"#14CE1C"
                                                                ,"candle_Down"          :	"#CE1414"
                                                                ,"edge_Up"              :	"#14CE1C"
                                                                ,"edge_Down"            :	"#CE1414"
                                                                ,"wick_Up"              :	"#ffffff"
                                                                ,"wick_Down"            :	"#ffffff"
                                                                ,"ohlc_Up"              :	"#ffffff"
                                                                ,"ohlc_Down"            :	"#ffffff"
                                                                ,"volume_Up"            :	"#14CE1C"
                                                                ,"volume_Down"          :	"#CE1414"
                                                                ,"vcdopcod"             :	 False	
                                                                ,"alpha"                :	 2.00	
                                                                ,"mavcolors"            :	 ['#ec009c','#78ff8f','#fcf120']	
                                                                ,"y_on_right"           :	 True	
                                                                ,"gridcolor"            :	 None	
                                                                ,"gridstyle"            :	 None	
                                                                ,"facecolor"            :	"Black"
                                                                ,"figcolor"             :	"gray"
                                                                ,"rc"                   :	"rc"
                                                                ,"axes.edgecolor"       :	"white"
                                                                ,"axes.linewidth"       :	 1.50	
                                                                ,"axes.labelsize"       :	"large"
                                                                ,"axes.labelweight"     :	"semibold"
                                                                ,"axes.grid"            :	 True	
                                                                ,"axes.grid.axis"       :	"both"
                                                                ,"axes.grid.which"      :	"major"
                                                                ,"grid.alpha"           :	 0.90	
                                                                ,"grid.color"           :	"#EBEE24"
                                                                ,"grid.linestyle"       :	":"	
                                                                ,"grid.linewidth"       :	 1.00	
                                                                ,"figure.facecolor"     :	"#0a0a0a"
                                                                ,"patch.linewidth"      :	 1.00	
                                                                ,"lines.linewidth"      :	 1.00	
                                                                ,"font.weight"          :	"medium"
                                                                ,"font.size"            :	 8.00	
                                                                ,"figure.titlesize"     :	"x-large"
                                                                ,"figure.titleweight"   :	"semibold"
                                                                ,"base_mpf_style"       :	"mike"
                                                                }
                                                )
                dfDay           = TS.fGetDayDataFrame(    
                                                        filePath           = filePathExcel      #r"C:\Users\lenovo\Desktop\TGL\BCG\OK BCG.xlsx"            
                                                        ,DaySheetName       = daySheet                   
                                                    )



                for indexScope in timeScope:
                #------------------------------------------------
                    # Import  data
                    dft = ImportData.fun2 (filePath_fun = filePathExcel , bookName = bookScope[indexScope]['SheetName'])    #   'YAHOO'      'IBKR 30m'  IBKR 1H
                    if( len(dft.index) == 0 ):
                        print ( "This Sheet : ",bookScope[indexScope]['SheetName']," is Empty")
                    if( len(dft.index) > 0 ):
                        # Replace NaN Values with Zeros in Pandas DataFrame     Help Link :- https://www.geeksforgeeks.org/replace-nan-values-with-zeros-in-pandas-dataframe/
                        dft = dft.fillna(0)
                        TimeStamp_df =TS.fun7(   dfD                = dfDay
                                                ,dfS                = dfTimeStamp 
                                                ,scope              = indexScope # ["5s","1m","5m","30m","1h","1d"]  
                                                ,scopeBook          = {
                                                                        "5s":"5s"
                                                                        ,'5 secs':"5s"
                                                                        ,"1m":"1m"
                                                                        ,'1 min':"1m"
                                                                        ,"5m":"5m"
                                                                        ,'5 mins':"5m"
                                                                        ,"30m":"30m"
                                                                        ,'30 mins':"30m"
                                                                        ,"1h":"1h"
                                                                        ,'1 hour':"1h"
                                                                        ,"1d":"1d" 
                                                                        ,'1 day':"1d" 
                                                                    }   
                                            )
                        

                    #------------------------------------------------
                        # Loop 2:-
                        
                        index = 0
                        for index in range(len(TimeStamp_df.Index1)):
                            
                            # Sampling data

                            df = dft[(TimeStamp_df.Index1[index]):(TimeStamp_df.Index2[index])]
                            if( len(df.index) == 0 ):   # Skip Empty Segment.
                                print (  "Ticker: ",tickerName,"  This Time Segment : " ,TimeStamp_df.Index1[index] ,":"  ,TimeStamp_df.Index2[index]   ," is Empty")
                            if( len(df.index) > 0 ):  
                                print ( "###############################################")
                                print ( "#######","Ticker: ",tickerName,"#######")
                                print (" in Time Segment : " ,TimeStamp_df.Index1[index] ,":"  ,TimeStamp_df.Index2[index]   ," is under Process")
                                print ( "###############################################")                         
                                        
                                dfH = df.High.max()
                                dfL = df.Low.min()
                                dfD = (dfH - dfL) * 0.05 

                                # df = dft['2023-09-20 09:30:00':'2023-09-20 10:00:00']
                                # df = dft
                                # print (df)

                            #------------------------------------------------
                                #Make Chart Style
                                
                                #Make Yellow Color Candel 
                                mco = YellowCandel.fun(dataFrame = df) 
                                
                                #Make Indicator
                                EMA = addIndicator.fun4(dataFrame = df, scope = indexScope )    # need one for IBKR

                                #Make a fibonacci lines
                                Fobi_Var = FibonacciLines.fun2 (     filePath_fun   = filePathExcel
                                                                    ,Fobi_Input     = Fobi_Var
                                                                    ,DaySheetName   = daySheet 
                                                                    ,fibonacciIndex = TimeStamp_df.Fibonacci[index]
                                                                    ,fibonacciBook  ={
                                                                                        "Last5Year"	    :{	'SheetName':	'IBKR 1Month'	,	"PeriodStart":	"04:00:00"	,	"PeriodEnd":	"20:00:00"	,	"TimeStart":	"04:00:00"	,	"TimeEnd":	"20:00:00"	,	"StartDelta":	-61	,	"EndDelta":	-1			},
                                                                                        "Last1Year"	    :{	'SheetName':	'IBKR 1Month'	,	"PeriodStart":	"04:00:00"	,	"PeriodEnd":	"20:00:00"	,	"TimeStart":	"04:00:00"	,	"TimeEnd":	"20:00:00"	,	"StartDelta":	-13	,	"EndDelta":	-1			},
                                                                                        "Last1Month"	:{	'SheetName':	'IBKR 1Day'	    ,	"PeriodStart":	"04:00:00"	,	"PeriodEnd":	"20:00:00"	,	"TimeStart":	"04:00:00"	,	"TimeEnd":	"20:00:00"	,	"StartDelta":	-31	,	"EndDelta":	-2			},
                                                                                        "Last1Week"	    :{	'SheetName':	'IBKR 1Day'	    ,	"PeriodStart":	"04:00:00"	,	"PeriodEnd":	"20:00:00"	,	"TimeStart":	"04:00:00"	,	"TimeEnd":	"20:00:00"	,	"StartDelta":	-6	,	"EndDelta":	-2			},
                                                                                        "Yesterday"	    :{	'SheetName':	'IBKR 4H'	    ,	"PeriodStart":	"04:00:00"	,	"PeriodEnd":	"20:00:00"	,	"TimeStart":	"04:00:00"	,	"TimeEnd":	"20:00:00"	,	"StartDelta":	-2	,	"EndDelta":	-1			},
                                                                                        "PreMarketI"	:{	'SheetName':	'IBKR 30m'	    ,	"PeriodStart":	"04:00:00"	,	"PeriodEnd":	"06:00:00"	,	"TimeStart":	"04:00:00"	,	"TimeEnd":	"20:00:00"	,	"StartDelta":	-2	,	"EndDelta":	-1			},
                                                                                        "PreMarketII"	:{	'SheetName':	'IBKR 30m'	    ,	"PeriodStart":	"06:00:00"	,	"PeriodEnd":	"08:00:00"	,	"TimeStart":	"04:00:00"	,	"TimeEnd":	"06:00:00"	,	"StartDelta":	-1	,	"EndDelta":	-1			},
                                                                                        "MarketOpen"	:{	'SheetName':	'IBKR 30m'	    ,	"PeriodStart":	"09:00:00"	,	"PeriodEnd":	"10:00:00"	,	"TimeStart":	"04:00:00"	,	"TimeEnd":	"09:00:00"	,	"StartDelta":	-1	,	"EndDelta":	-1			},
                                                                                        "Midday"	    :{	'SheetName':	'IBKR 30m'	    ,	"PeriodStart":	"10:00:00"	,	"PeriodEnd":	"14:00:00"	,	"TimeStart":	"04:00:00"	,	"TimeEnd":	"10:00:00"	,	"StartDelta":	-1	,	"EndDelta":	-1			},
                                                                                        "PowerHour"	    :{	'SheetName':	'IBKR 30m'	    ,	"PeriodStart":	"14:00:00"	,	"PeriodEnd":	"16:00:00"	,	"TimeStart":	"04:00:00"	,	"TimeEnd":	"14:00:00"	,	"StartDelta":	-1	,	"EndDelta":	-1			},
                                                                                        "AfterMarketI"	:{	'SheetName':	'IBKR 30m'	    ,	"PeriodStart":	"16:00:00"	,	"PeriodEnd":	"18:00:00"	,	"TimeStart":	"04:00:00"	,	"TimeEnd":	"16:00:00"	,	"StartDelta":	-1	,	"EndDelta":	-1			},
                                                                                        "AfterMarketII"	:{	'SheetName':	'IBKR 30m'	    ,	"PeriodStart":	"18:00:00"	,	"PeriodEnd":	"20:00:00"	,	"TimeStart":	"04:00:00"	,	"TimeEnd":	"18:00:00"	,	"StartDelta":	-1	,	"EndDelta":	-1			},
                                                                                        "Final"	        :{	'SheetName':	'IBKR 30m'	    ,	"PeriodStart":	"20:00:00"	,	"PeriodEnd":	"00:00:00"	,	"TimeStart":	"04:00:00"	,	"TimeEnd":	"20:00:00"	,	"StartDelta":	-1	,	"EndDelta":	-1			},
                                                                                        "Random1"	    :{	'SheetName':	'IBKR 8H'	    ,	"PeriodStart":	"XX:XX:XX"	,	"PeriodEnd":	"XX:XX:XX"	,	"TimeStart":	"XX:XX:XX"	,	"TimeEnd":	"XX:XX:XX"	,	"StartDelta":	-1	,	"EndDelta":	-1			},
                                                                                        "Random2"	    :{	'SheetName':	'IBKR 8H'	    ,	"PeriodStart":	"XX:XX:XX"	,	"PeriodEnd":	"XX:XX:XX"	,	"TimeStart":	"XX:XX:XX"	,	"TimeEnd":	"XX:XX:XX"	,	"StartDelta":	-1	,	"EndDelta":	-1			},
                                                                                        "Random3"	    :{	'SheetName':	'IBKR 8H'	    ,	"PeriodStart":	"XX:XX:XX"	,	"PeriodEnd":	"XX:XX:XX"	,	"TimeStart":	"XX:XX:XX"	,	"TimeEnd":	"XX:XX:XX"	,	"StartDelta":	-1	,	"EndDelta":	-1			},
                                                                                        "Random4"	    :{	'SheetName':	'IBKR 8H'	    ,	"PeriodStart":	"XX:XX:XX"	,	"PeriodEnd":	"XX:XX:XX"	,	"TimeStart":	"XX:XX:XX"	,	"TimeEnd":	"XX:XX:XX"	,	"StartDelta":	-1	,	"EndDelta":	-1			},
                                                                                        "Random5"	    :{	'SheetName':	'IBKR 8H'	    ,	"PeriodStart":	"XX:XX:XX"	,	"PeriodEnd":	"XX:XX:XX"	,	"TimeStart":	"XX:XX:XX"	,	"TimeEnd":	"XX:XX:XX"	,	"StartDelta":	-1	,	"EndDelta":	-1			}
                                                                                    }
                                                                )

                                
                                

                                #Make Time Lable
                                timeLable = TS.fTimeLable2  (  f_Scope        = TimeStamp_df.Scope[index] # (source=daySheet   ,interval=indexScope   ,inList=list (df.index.map(str)))
                                                            ,f_HourConstant = TimeStamp_df.HourConstant[index]
                                                            ,f_MinConstant  = TimeStamp_df.MinConstant[index]
                                                            ,f_SecConstant  = TimeStamp_df.SecConstant[index]
                                                            ,f_leftSide     = TimeStamp_df.leftSide[index]
                                                            ,source         = "IBKR"
                                                            ,interval       = indexScope
                                                            ,inList         = list (df.index.map(str))
                                                        )

                                #Make Price Label
                                priceLabel = TS.LabelPrice6 (    pMax = dfH + dfD
                                                                ,pMin = dfL - dfD
                                                            )

                                
                                # #Saving plot to a file -> Link: https://github.com/matplotlib/mplfinance/blob/master/examples/savefig.ipynb
                                # setName = '\\' + tickerName + '_' + str("%03d"%imgCount) + '_'  + indexScope + '_' + TimeStamp_df.Name[index] 
                                # save = dict(fname= (rxChartPath + setName + imageType), dpi= 300, pad_inches= 0)

                                #Make a chart
                                fig, axlist = mpf.plot(  data                   = df
                                                        ,title                  = (tickerName + ' ' + indexScope)
                                                        ,type                   = 'candle' 
                                                        # ,mav                    = (20,50)
                                                        ,volume                 = True
                                                        ,show_nontrading        = False
                                                        ,tight_layout           = True
                                                        ,figratio               = (1,1)
                                                        ,figscale               = 3
                                                        ,scale_padding          = 0.30
                                                        ,figsize               = (12,6)
                                                        # ,ylim                   = (((df.Low.min())*0.95) ,((df.High.max())*1.05)) # set min and max of Chart
                                                        ,ylim                   = ((dfL - dfD) ,(dfH + dfD)) # set min and max of Chart (dfH + dfD)
                                                        ,xrotation              = 0
                                                        ,yscale                 = "linear" # y-axis scale: "linear", "log", "symlog", or "logit"
                                                        ,volume_yscale          = "linear" # Volume y-axis scale: "linear", "log", "symlog", or "logit"
                                                        ,style                  = ChartStyle
                                                        ,marketcolor_overrides  = mco
                                                        ,mco_faceonly           = False
                                                        ,addplot                = EMA
                                                        ,hlines                 = Fobi_Var.fibolines
                                                        ,returnfig=True
                                                        # ,savefig                = save # link: https://github.com/matplotlib/mplfinance/blob/master/examples/savefig.ipynb
                                                        # ,marketcolor_overrides = mco  # link https://github.com/matplotlib/mplfinance/blob/master/examples/marketcolor_overrides.ipynb
                                )
                                
                                #Add Time Lable      Help link:   https://github.com/matplotlib/mplfinance/issues/573
                                axlist[-2].set_xticks(timeLable["ticks"])
                                axlist[-2].set_xticklabels(timeLable["tlabs"], ha='center')
                                axlist[-2].set_xticks(timeLable["mitks"],minor=True)
                                axlist[-2].set_xticklabels(timeLable["milab"], ha='center', minor=True, rotation=0) 

                                x2 = axlist[0].secondary_xaxis('top')       # Help Link :-  https://matplotlib.org/stable/gallery/ticks/multilevel_ticks.html#sphx-glr-gallery-ticks-multilevel-ticks-py
                                x2.set_xticks(timeLable["ticks"])           # Help Link :-  https://matplotlib.org/stable/api/axes_api.html
                                x2.set_xticklabels(timeLable["Utlabs"], ha='center')
                                x2.set_xticks(timeLable["mitks"],minor=True)
                                x2.set_xticklabels(timeLable["Umilab"], ha='center', minor=True, rotation=0)
                                
                                #Add Price Lable     Help Link:-   https://github.com/matplotlib/mplfinance/issues/604 
                                axlist[0].set_yticks( priceLabel['ticks'] )
                                axlist[0].set_yticklabels( priceLabel['tlabs'] )
                                                
                                y2 = axlist[0].secondary_yaxis('left')      # Help Link :-  https://github.com/matplotlib/mplfinance/issues/587
                                y2.set_yticks(priceLabel['ticks'])
                                y2.set_yticklabels( priceLabel['tlabs'] )
                                
                                
                                #Saving plot to a file -> Link: https://github.com/matplotlib/mplfinance/blob/master/examples/savefig.ipynb
                                setName = '\\' + tickerName + '_' + str("%03d"%imgCount) + '_'  + indexScope + '_' + TimeStamp_df.Name[index] 
                                save = dict(fname= (rxChartPath + setName + imageType), dpi= 300, pad_inches= 0)
                                
                                
                                #Saving Chart        Help link:   https://github.com/matplotlib/mplfinance/issues/573
                                fig.savefig( fname= (rxChartPath + setName + imageType), dpi= 300, pad_inches= 0, bbox_inches='tight' )
                                imgCount += 1

                                # mpf.show()
                                # matplotlib.pyplot.savefig()
                                # fig.save()
                                
                                # axlist[-2].save()
                                # fig.savefig( 'myplot.pdf', bbox_inches='tight' )           
###_____________________________________________________________________________________________________________________________________________________________________________________







AttributeList =["style_name","base_mpl_style","marketcolors","candle_Up","candle_Down","edge_Up","edge_Down","wick_Up","wick_Down","ohlc_Up","ohlc_Down","volume_Up","volume_Down","vcdopcod","alpha","mavcolors","y_on_right","gridcolor","gridstyle","facecolor","figcolor","rc","axes.edgecolor","axes.linewidth","axes.labelsize","axes.labelweight","axes.grid","axes.grid.axis","axes.grid.which","grid.alpha","grid.color","grid.linestyle","grid.linewidth","figure.facecolor","patch.linewidth","lines.linewidth","font.weight","font.size","figure.titlesize","figure.titleweight","base_mpf_style"]


AttributeList =[
                 "style_name"
                ,"base_mpl_style"
                ,"marketcolors"
                ,"candle_Up"
                ,"candle_Down"
                ,"edge_Up"
                ,"edge_Down"
                ,"wick_Up"
                ,"wick_Down"
                ,"ohlc_Up"
                ,"ohlc_Down"
                ,"volume_Up"
                ,"volume_Down"
                ,"vcdopcod"
                ,"alpha"
                ,"mavcolors"
                ,"y_on_right"
                ,"gridcolor"
                ,"gridstyle"
                ,"facecolor"
                ,"figcolor"
                ,"rc"
                ,"axes.edgecolor"
                ,"axes.linewidth"
                ,"axes.labelsize"
                ,"axes.labelweight"
                ,"axes.grid"
                ,"axes.grid.axis"
                ,"axes.grid.which"
                ,"grid.alpha"
                ,"grid.color"
                ,"grid.linestyle"
                ,"grid.linewidth"
                ,"figure.facecolor"
                ,"patch.linewidth"
                ,"lines.linewidth"
                ,"font.weight"
                ,"font.size"
                ,"figure.titlesize"
                ,"figure.titleweight"
                ,"base_mpf_style"]

def tAttributeDic():
    AttributeDic["style_name"]
    AttributeDic["base_mpl_style"]
    AttributeDic["marketcolors"]
    AttributeDic["candle_Up"]
    AttributeDic["candle_Down"]
    AttributeDic["edge_Up"]
    AttributeDic["edge_Down"]
    AttributeDic["wick_Up"]
    AttributeDic["wick_Down"]
    AttributeDic["ohlc_Up"]
    AttributeDic["ohlc_Down"]
    AttributeDic["volume_Up"]
    AttributeDic["volume_Down"]
    AttributeDic["vcdopcod"]
    AttributeDic["alpha"]
    AttributeDic["mavcolors"]
    AttributeDic["y_on_right"]
    AttributeDic["gridcolor"]
    AttributeDic["gridstyle"]
    AttributeDic["facecolor"]
    AttributeDic["figcolor"]
    AttributeDic["rc"]
    AttributeDic["axes.edgecolor"]
    AttributeDic["axes.linewidth"]
    AttributeDic["axes.labelsize"]
    AttributeDic["axes.labelweight"]
    AttributeDic["axes.grid"]
    AttributeDic["axes.grid.axis"]
    AttributeDic["axes.grid.which"]
    AttributeDic["grid.alpha"]
    AttributeDic["grid.color"]
    AttributeDic["grid.linestyle"]
    AttributeDic["grid.linewidth"]
    AttributeDic["figure.facecolor"]
    AttributeDic["patch.linewidth"]
    AttributeDic["lines.linewidth"]
    AttributeDic["font.weight"]
    AttributeDic["font.size"]
    AttributeDic["figure.titlesize"]
    AttributeDic["figure.titleweight"]
    AttributeDic["base_mpf_style"]









#--------------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------



# fun6()

if(False):

    dfS = fGetTimeStampDataFrame(    
                timeStampPath      = r"D:\Python Tools\ChartMaker\SourceDocuments\InPut_Excel\Watch_List_Zamzam .xlsx"
                ,timeStampSheet     = "Sheet2"                  
            )



    stayloo,timescopo = fGetStyle2(    
                dfS
                ,ShowOutPut = True                              
            )
    print(stayloo)

    dfD = fGetDayDataFrame()
    print(dfD)



    testoo = fun7(    
                dfD
                ,dfS  
                ,scope              =  "1m" # ["5s","1m","5m","30m","1h","1d"]  
                )

    print(testoo)

    print("***@@@***")
    print(timescopo)
    print("***@@@***")

    print("__________________________")
    print("Test Program Now is End !!")
    print("__________________________")


#--------------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------


# Help Link :-      https://stackoverflow.com/questions/2003870/how-can-i-select-all-of-the-sundays-for-a-year-using-python

# from datetime import datetime,date, timedelta

# def allsundays(year):
#    d = date(year, 1, 1)                    # January 1st
#    d += timedelta(days = 0 - d.weekday())  # First Sunday
#    while d.year == year:
#       yield d
#       d += timedelta(days = 7)

# for d in allsundays(2024):
#    print(d)




# dfD = fGetDayDataFrame()
# last_day = dfD.index[-1]
# last_day = datetime.strptime(last_day, '%Y-%m-%d').date()
# print(last_day,type(last_day))


# next_Monday = last_day + timedelta(days = 0 - last_day.weekday())
# print(next_Monday)

if(False):

    last_day= date(2024, 4, 8)
    print(last_day.year,last_day.month,last_day.day)

    temp_Monday = last_day + timedelta(days = 0 - last_day.weekday())
    temp_Sunday = last_day + timedelta(days = 6 - last_day.weekday())

    monday_list = []
    Sunday_list = []

    monday_list.append(str(temp_Monday))
    Sunday_list.append(str(temp_Sunday))

    for i in range( 8 ):
        temp_Monday -= timedelta(days = 7)
        temp_Sunday -= timedelta(days = 7)

        monday_list.append(str(temp_Monday))
        Sunday_list.append(str(temp_Sunday))

    print(monday_list)
    print(Sunday_list)
        









class FiboClass2():
    
    def __init__(self):
        self.fibolines=dict( hlines=(0, 0, 0, 0, 0, 0, 0),
                        colors=['#4ACE14','#E21919','#27DFDD', '#E6F226', '#27DFDD', '#E21919','#4ACE14'],
                        linestyle='solid',
                        linewidths=3,
                        alpha= 0.8 
                    )
        self.HighLow =   {
                        "Last5Year"	    :{	"High":	0.00	,	"Low":	0.00	},
                        "Last1Year"	    :{	"High":	0.00	,	"Low":	0.00	},
                        "CurrentYear"   :{	"High":	0.00	,	"Low":	0.00	},
                        "Last1Month"	:{	"High":	0.00	,	"Low":	0.00	},
                        "CurrentMonth"	:{	"High":	0.00	,	"Low":	0.00	},
                        "Last1Week" 	:{	"High":	0.00	,	"Low":	0.00	},
                        "CurrentWeek" 	:{	"High":	0.00	,	"Low":	0.00	},
                        "Yesterday"	    :{	"High":	0.00	,	"Low":	0.00	},
                        "PreMarketI"	:{	"High":	0.00	,	"Low":	0.00	},
                        "PreMarketII"	:{	"High":	0.00	,	"Low":	0.00	},
                        "MarketOpen"	:{	"High":	0.00	,	"Low":	0.00	},
                        "Midday"	    :{	"High":	0.00	,	"Low":	0.00	},
                        "PowerHour"	    :{	"High":	0.00	,	"Low":	0.00	},
                        "AfterMarketI"	:{	"High":	0.00	,	"Low":	0.00	},
                        "AfterMarketII"	:{	"High":	0.00	,	"Low":	0.00	},
                        "Final"	        :{	"High":	0.00	,	"Low":	0.00	},
                        "EMA200"	    :{	"High":	0.00	,	"Low":	0.00	},
                        "Random1"	    :{	"High":	0.00	,	"Low":	0.00	},
                        "Random2"	    :{	"High":	0.00	,	"Low":	0.00	},
                        "Random3"	    :{	"High":	0.00	,	"Low":	0.00	},
                        "Random4"	    :{	"High":	0.00	,	"Low":	0.00	},
                        "Random5"	    :{	"High":	0.00	,	"Low":	0.00	}
                    }

    def HLday ( self 
                ,dfD
                ,dfT 
                # ,filepath_xl
                ,showOutPut = True
                ,dayIndex = "MarketOpen"
                ,dayList = ["Yesterday","PreMarketI","PreMarketII","MarketOpen","Midday","PowerHour","AfterMarketI","AfterMarketII","Final"]
                ,dayBook =  {
                                "Yesterday"	    :{"TimeStart":"04:00:00" ,"TimeEnd":"20:00:00" ,"StartDelta":-2	,"EndDelta":-2},
                                "PreMarketI"	:{"TimeStart":"04:00:00" ,"TimeEnd":"06:00:00" ,"StartDelta":-1	,"EndDelta":-1},
                                "PreMarketII"	:{"TimeStart":"04:00:00" ,"TimeEnd":"08:00:00" ,"StartDelta":-1	,"EndDelta":-1},
                                "MarketOpen"	:{"TimeStart":"04:00:00" ,"TimeEnd":"09:00:00" ,"StartDelta":-1	,"EndDelta":-1},
                                "Midday"	    :{"TimeStart":"04:00:00" ,"TimeEnd":"10:00:00" ,"StartDelta":-1	,"EndDelta":-1},
                                "PowerHour"	    :{"TimeStart":"04:00:00" ,"TimeEnd":"14:00:00" ,"StartDelta":-1	,"EndDelta":-1},
                                "AfterMarketI"	:{"TimeStart":"04:00:00" ,"TimeEnd":"16:00:00" ,"StartDelta":-1	,"EndDelta":-1},
                                "AfterMarketII" :{"TimeStart":"04:00:00" ,"TimeEnd":"18:00:00" ,"StartDelta":-1	,"EndDelta":-1},
                                "Final"	        :{"TimeStart":"04:00:00" ,"TimeEnd":"20:00:00" ,"StartDelta":-1	,"EndDelta":-1}
                            }
                ):

        for dayIndex in dayList :
            EDay = str( dfD.index[dayBook[dayIndex]["EndDelta"]] )
            SDay = str( dfD.index[dayBook[dayIndex]["StartDelta"]] )
            ETime=EDay + " " + dayBook[dayIndex]["TimeEnd"]  
            STime=SDay + " " + dayBook[dayIndex]["TimeStart"]
            if(showOutPut):print(STime)
            if(showOutPut):print(ETime)
            df = dfT[STime:ETime] #XXXXXXXX
            High = self.HighLow[dayIndex]["High"] = df['High'].max()
            Low  = self.HighLow[dayIndex]["Low"]  = df['Low'].min()
            if(showOutPut):print(dayIndex,"High",High)
            if(showOutPut):print(dayIndex,"Low",Low)

    def HLweek ( self 
                ,dfD
                ,dfT 
                # ,filepath_xl
                ,showOutPut = True
                ):
        # last_day= date(2024, 4, 8)
        last_day = dfD.index[-1]
        last_day = datetime.strptime(last_day, '%Y-%m-%d').date()
        if(showOutPut):print(last_day.year,last_day.month,last_day.day)

        temp_Monday = last_day + timedelta(days = 0 - last_day.weekday())
        temp_Sunday = last_day + timedelta(days = 6 - last_day.weekday())

        monday_list = []
        Sunday_list = []

        monday_list.append(str(temp_Monday))
        Sunday_list.append(str(temp_Sunday))

        for i in range( 2 ):
            temp_Monday -= timedelta(days = 7)
            temp_Sunday -= timedelta(days = 7)

            monday_list.append(str(temp_Monday))
            Sunday_list.append(str(temp_Sunday))

        if(showOutPut):print(monday_list)
        if(showOutPut):print(Sunday_list)

        # dfT = ImportData.fun2 (filePath_fun = filepath_xl , bookName = 'IBKR 30m')    #   'YAHOO'      'IBKR 30m'  IBKR 1H

        # CurrentWeek
        STime=monday_list[0] + ' ' + "04:00:00"
        ETime=Sunday_list[0] + ' ' + "20:00:00"        
        df = dfT[STime:ETime] #XXXXXXXX
        High = self.HighLow["CurrentWeek"]["High"] = df['High'].max()
        Low  = self.HighLow["CurrentWeek"]["Low"]  = df['Low'].min()
        if(showOutPut):print("CurrentWeek","High",High)
        if(showOutPut):print("CurrentWeek","Low",Low)

        # Last1Week
        STime=monday_list[1] + ' ' + "04:00:00"
        ETime=Sunday_list[1] + ' ' + "20:00:00"        
        df = dfT[STime:ETime] #XXXXXXXX
        High = self.HighLow["Last1Week"]["High"] = df['High'].max()
        Low  =self.HighLow["Last1Week"]["Low"]  = df['Low'].min()
        if(showOutPut):print("Last1Week","High",High)
        if(showOutPut):print("Last1Week","Low",Low)

    def HLmonth ( self 
                ,dfD
                ,dfT 
                # ,filepath_xl
                ,showOutPut = True
                ):

        last_day = dfD.index[-1]
        last_day = datetime.strptime(last_day, '%Y-%m-%d').date()
        if(showOutPut):print(last_day.year,last_day.month,last_day.day)

        # CurrentMonth
        first_day = last_day.replace(day=1)    #Help Link :-   https://stackoverflow.com/questions/12468823/python-datetime-setting-fixed-hour-and-minute-after-using-strptime-to-get-day
        STime = str(first_day) + ' ' + "04:00:00"
        ETime = str(last_day)  + ' ' + "20:00:00" 
        if(showOutPut):print(STime)
        if(showOutPut):print(ETime)
        df = dfT[STime:ETime] #XXXXXXXX
        High = self.HighLow["CurrentMonth"]["High"] = df['High'].max()
        Low  = self.HighLow["CurrentMonth"]["Low"]  = df['Low'].min()
        if(showOutPut):print("CurrentMonth","High",High)
        if(showOutPut):print("CurrentMonth","Low",Low) 

        # Last1Month
        from dateutil.relativedelta import relativedelta
        first_LMD = first_day - relativedelta(months=1)     #Help Link :- https://www.geeksforgeeks.org/add-months-to-datetime-object-in-python/
        last_LMD  = first_day - timedelta(days = 1)
        STime = str(first_LMD) + ' ' + "04:00:00"
        ETime = str(last_LMD)  + ' ' + "20:00:00" 
        if(showOutPut):print(STime)
        if(showOutPut):print(ETime)
        df = dfT[STime:ETime] #XXXXXXXX
        High = self.HighLow["Last1Month"]["High"] = df['High'].max()
        Low  = self.HighLow["Last1Month"]["Low"]  = df['Low'].min()
        if(showOutPut):print("Last1Month","High",High)
        if(showOutPut):print("Last1Month","Low",Low)

    def HLyear ( self 
                ,dfD
                ,dfT 
                # ,filepath_xl
                ,showOutPut = True
                ):

        last_day = dfD.index[-1]
        last_day = datetime.strptime(last_day, '%Y-%m-%d').date()
        if(showOutPut):print(last_day.year,last_day.month,last_day.day)

        # CurrentYear
        first_day = last_day.replace(day=1,month=1)    #Help Link :-   https://stackoverflow.com/questions/12468823/python-datetime-setting-fixed-hour-and-minute-after-using-strptime-to-get-day
        STime = str(first_day) + ' ' + "04:00:00"
        ETime = str(last_day)  + ' ' + "20:00:00" 
        if(showOutPut):print(STime)
        if(showOutPut):print(ETime)
        df = dfD[STime:ETime] #XXXXXXXX
        High = self.HighLow["CurrentYear"]["High"] = df['High'].max()
        Low  = self.HighLow["CurrentYear"]["Low"]  = df['Low'].min()
        if(showOutPut):print("CurrentYear","High",High)
        if(showOutPut):print("CurrentYear","Low",Low) 

        # Last1Year
        from dateutil.relativedelta import relativedelta
        first_LMD = first_day - relativedelta(years=1)     #Help Link :- https://www.geeksforgeeks.org/add-months-to-datetime-object-in-python/
        last_LMD  = first_day - timedelta(days = 1)
        STime = str(first_LMD) + ' ' + "04:00:00"
        ETime = str(last_LMD)  + ' ' + "20:00:00" 
        if(showOutPut):print(STime)
        if(showOutPut):print(ETime)
        df = dfD[STime:ETime] #XXXXXXXX
        High = self.HighLow["Last1Year"]["High"] = df['High'].max()
        Low  = self.HighLow["Last1Year"]["Low"]  = df['Low'].min()
        if(showOutPut):print("Last1Year","High",High)
        if(showOutPut):print("Last1Year","Low",Low)

    def HL_EMA200 ( self 
                    ,dfD
                    ,dfT 
                    # ,filepath_xl
                    ,showOutPut = True
                ):

        self.HighLow["EMA200"]["High"] = dfD.EMA200[-1]

    def GetHighLow(  self 
                    ,dfD
                    ,dfT 
                    # ,filepath_xl
                    ,showOutPut = True
                    ):
        self.HLday      (dfD,dfT,showOutPut)
        self.HLweek     (dfD,dfT,showOutPut)
        self.HLmonth    (dfD,dfT,showOutPut)
        self.HLyear     (dfD,dfT,showOutPut)
        self.HL_EMA200  (dfD,dfT,showOutPut)

    def MakeHorizontalLine  (    self
                                ,style = "Sheet2"
                                ,fibonacciIndex = "MarketOpen"
                                ,showOutPut = True
                            ):
        # 1- Check Style
        if(style in ["Ehab","Ehab1","Sheet2"]):
            # 2- Ist need fibonacci lines?
            if(fibonacciIndex in self.HighLow):
                if(showOutPut):print("it need fibonacci lines")
                if(showOutPut):print("Import Data")
                High = self.HighLow[fibonacciIndex]["High"]
                Low  = self.HighLow[fibonacciIndex]["Low"]
            else:
                if(showOutPut):print("it dont need fibonacci lines")
                High = -100
                Low  = -100
            # 3- Culclate fibonacci lines   Help Link :-  https://eodhd.com/financial-academy/technical-analysis-examples/fibonacci-sequence-in-trading-with-python/
            Fib100 = {'Value':(Low + ((High - Low) * 1.000 ))  ,'linestyle':'solid'  ,'linewidth':3   ,'alpha':0.8   ,'color':'#4ACE14'}
            Fib764 = {'Value':(Low + ((High - Low) * 0.764 ))  ,'linestyle':'solid'  ,'linewidth':3   ,'alpha':0.8   ,'color':'#E21919'}
            Fib618 = {'Value':(Low + ((High - Low) * 0.618 ))  ,'linestyle':'solid'  ,'linewidth':3   ,'alpha':0.8   ,'color':'#27DFDD'}
            Fib500 = {'Value':(Low + ((High - Low) * 0.500 ))  ,'linestyle':'solid'  ,'linewidth':3   ,'alpha':0.8   ,'color':'#E6F226'}
            Fib382 = {'Value':(Low + ((High - Low) * 0.382 ))  ,'linestyle':'-.'     ,'linewidth':3   ,'alpha':0.8   ,'color':'#27DFDD'}
            Fib236 = {'Value':(Low + ((High - Low) * 0.236 ))  ,'linestyle':'-.'     ,'linewidth':3   ,'alpha':0.8   ,'color':'#E21919'}
            Fib000 = {'Value':(Low + ((High - Low) * 0.000 ))  ,'linestyle':'-.'     ,'linewidth':3   ,'alpha':0.8   ,'color':'#4ACE14'}
            
            # 3- Make  fibonacci lines
            self.fibolines=dict   (      hlines      =(Fib100['Value']       ,Fib764['Value']        ,Fib618['Value']        ,Fib500['Value']        ,Fib382['Value']        ,Fib236['Value']        ,Fib000['Value'])
                                        ,colors      =[Fib100['color']       ,Fib764['color']        ,Fib618['color']        ,Fib500['color']        ,Fib382['color']        ,Fib236['color']        ,Fib000['color']]
                                        ,linestyle   =[Fib100['linestyle']   ,Fib764['linestyle']    ,Fib618['linestyle']    ,Fib500['linestyle']    ,Fib382['linestyle']    ,Fib236['linestyle']    ,Fib000['linestyle']]
                                        ,linewidths  =[Fib100['linewidth']   ,Fib764['linewidth']    ,Fib618['linewidth']    ,Fib500['linewidth']    ,Fib382['linewidth']    ,Fib236['linewidth']    ,Fib000['linewidth']]
                                        ,alpha       =[Fib100['alpha']       ,Fib764['alpha']        ,Fib618['alpha']        ,Fib500['alpha']        ,Fib382['alpha']        ,Fib236['alpha']        ,Fib000['alpha']]
                                    )

        elif(style in ["Zamzam"]):
            if(fibonacciIndex in ["All"]):
                z_LineIndex = "Last1Year"
                z_LineList = ["Last1Year","CurrentYear","Last1Month","CurrentMonth","Last1Week","CurrentWeek","Yesterday","MarketOpen","EMA200"]
                z_LineBook =    {
                                "Last1Year"	    :{'Value':self.HighLow["Last1Year"]["High"]	    ,'linestyle':'solid'  ,'linewidth':1  ,'alpha':0.8	,'color':'#09F6BD'},
                                "CurrentYear"	:{'Value':self.HighLow["CurrentYear"]["High"]	,'linestyle':'solid'  ,'linewidth':1  ,'alpha':0.8	,'color':'#13E1EC'},
                                "Last1Month"	:{'Value':self.HighLow["Last1Month"]["High"]	,'linestyle':'solid'  ,'linewidth':1  ,'alpha':0.8	,'color':'#E01F7D'},
                                "CurrentMonth"	:{'Value':self.HighLow["CurrentMonth"]["High"]	,'linestyle':'solid'  ,'linewidth':1  ,'alpha':0.8	,'color':'#BC38C7'},
                                "Last1Week"	    :{'Value':self.HighLow["Last1Week"]["High"]	    ,'linestyle':'solid'  ,'linewidth':1  ,'alpha':0.8	,'color':'#C3893C'},
                                "CurrentWeek"	:{'Value':self.HighLow["CurrentWeek"]["High"]	,'linestyle':'solid'  ,'linewidth':1  ,'alpha':0.8	,'color':'#AAB847'},
                                "Yesterday"	    :{'Value':self.HighLow["Yesterday"]["High"]	    ,'linestyle':'solid'  ,'linewidth':1  ,'alpha':0.8	,'color':'#4D15EA'},
                                "MarketOpen"    :{'Value':self.HighLow["MarketOpen"]["High"]	,'linestyle':'solid'  ,'linewidth':1  ,'alpha':0.8	,'color':'#0FF045'},
                                "EMA200"	    :{'Value':self.HighLow["EMA200"]["High"]	    ,'linestyle':'solid'  ,'linewidth':1  ,'alpha':0.8	,'color':'#F20D2E'}
                                }
                Z_hlines        = []
                Z_colors        = []
                Z_linestyle     = []
                Z_linewidths    = []
                Z_alpha         = []
                for z_LineIndex in z_LineList :  Z_hlines.append     (z_LineBook[z_LineIndex]['Value'])
                for z_LineIndex in z_LineList :  Z_colors.append     (z_LineBook[z_LineIndex]['color'])
                for z_LineIndex in z_LineList :  Z_linestyle.append  (z_LineBook[z_LineIndex]['linestyle'])
                for z_LineIndex in z_LineList :  Z_linewidths.append (z_LineBook[z_LineIndex]['linewidth'])
                for z_LineIndex in z_LineList :  Z_alpha.append      (z_LineBook[z_LineIndex]['alpha'])
                self.fibolines=dict(     hlines      =  Z_hlines
                                        ,colors      =  Z_colors
                                        ,linestyle   =  Z_linestyle
                                        ,linewidths  =  Z_linewidths
                                        ,alpha       =  Z_alpha
                                    )







if(False):
    xlFileLink = r"D:\Python Tools\ChartMaker\SourceDocuments\OutPut_Excel\Session [20231128_111940]\GME\OK GME.xlsx"
    fobiVar1 = FiboClass()
    # print(fobiVar1.__dict__)

    dfD = fGetDayDataFrame(filePath = xlFileLink)
    print(dfD)

    dfT = ImportData.fun2 (filePath_fun = xlFileLink , bookName = 'IBKR 30m')    #   'YAHOO'      'IBKR 30m'  IBKR 1H
    print(dfT)

    fobiVar1.HLday   (dfD,dfT)
    fobiVar1.HLweek  (dfD,dfT)
    fobiVar1.HLmonth (dfD,dfT)
    fobiVar1.HLyear (dfD,dfT)
    fobiVar1.HL_EMA200  (dfD,dfT)
    

    print(fobiVar1.__dict__)






if ( False ):
                # ,dfD
                # ,dfT 
                # # ,filepath_xl
                # ,showOutPut = True

        showOutPut = True
        last_day= date(2024, 3, 28)

        # last_day = dfD.index[-1]
        # last_day = datetime.strptime(last_day, '%Y-%m-%d').date()
        if(showOutPut):print(last_day.year,last_day.month,last_day.day)

        # CurrentMonth
        first_day = last_day.replace(day=1)    #Help Link :-   https://stackoverflow.com/questions/12468823/python-datetime-setting-fixed-hour-and-minute-after-using-strptime-to-get-day
        STime = str(first_day) + ' ' + "04:00:00"
        ETime = str(last_day)  + ' ' + "20:00:00" 
        if(showOutPut):print(STime)
        if(showOutPut):print(ETime)
        # df = dfT[STime:ETime] #XXXXXXXX
        # High = self.HighLow["CurrentMonth"]["High"] = df['High'].max()
        # Low  = self.HighLow["CurrentMonth"]["Low"]  = df['Low'].min()
        # if(showOutPut):print("CurrentMonth","High",High)
        # if(showOutPut):print("CurrentMonth","Low",Low) 

        # Last1Month
        from dateutil.relativedelta import relativedelta
        first_LMD = first_day - relativedelta(months=1)     #Help Link :- https://www.geeksforgeeks.org/add-months-to-datetime-object-in-python/
        last_LMD  = first_day - timedelta(days = 1)
        STime = str(first_LMD) + ' ' + "04:00:00"
        ETime = str(last_LMD)  + ' ' + "20:00:00" 
        if(showOutPut):print(STime)
        if(showOutPut):print(ETime)
        # df = dfT[STime:ETime] #XXXXXXXX
        # High = self.HighLow["Last1Month"]["High"] = df['High'].max()
        # Low  = self.HighLow["Last1Month"]["Low"]  = df['Low'].min()
        # if(showOutPut):print("Last1Month","High",High)
        # if(showOutPut):print("Last1Month","Low",Low) 




if(False):
    listFor=[] 
    for i in range(8): listFor.append(i)
    print(listFor)
    goba = (1,4,8,6,7,3,2,1)   # 1,4,8,6,7,3,2,1
    print(goba , type(goba))


if(False):
    import time
    timeep =      1013610600
    timeout = time.strftime("%a, %d %b %Y %H:%M:%S +0000", time.localtime(timeep))    
    print(timeout)

    timeout2 = time.strftime("%Y/%m/%d ", time.localtime(timeep))
    print(timeout2)




#_________________________________________________________________________________________________________________________________________________________
#_________________________________________________________________________________________________________________________________________________________
#_________________________________________________________________________________________________________________________________________________________
#_________________________________________________________________________________________________________________________________________________________
#_________________________________________________________________________________________________________________________________________________________
#_________________________________________________________________________________________________________________________________________________________
#_________________________________________________________________________________________________________________________________________________________
#_________________________________________________________________________________________________________________________________________________________
#_________________________________________________________________________________________________________________________________________________________




#_________________________________________________________________________________________________________________________________________________________
def funSamaryImport (filePath_fun, bookName = 'Samary') :
    # Import  data
    #------------------------------------------------  
    book = openpyxl.load_workbook(filePath_fun)
    sheet = book[bookName]
    df= pd.DataFrame(sheet.values)
    # format  data
    #------------------------------------------------  
    #Convert row to column header for Pandas DataFrame : Link https://stackoverflow.com/questions/26147180/convert-row-to-column-header-for-pandas-dataframe
    df.columns = df.iloc[0]
    df = df.drop(0)
    df.reset_index(inplace=True)
    df = df.drop(columns=['index', None])
    return df


def funYahooImport (filePath_fun, bookName = "Yahoo 1m") :

    # Set Variable    
    intraDayList = ['Yahoo Hours','Yahoo 30m','Yahoo 15m','Yahoo 5m','Yahoo 2m','Yahoo 1m']
    swingDayList = ['Yahoo Dayes']
    
    # Import  data
    #------------------------------------------------  
    book = openpyxl.load_workbook(filePath_fun)
    sheet = book[bookName]
    df= pd.DataFrame(sheet.values)


    # format  data
    #------------------------------------------------  
    #Convert row to column header for Pandas DataFrame : Link https://stackoverflow.com/questions/26147180/convert-row-to-column-header-for-pandas-dataframe
    df.columns = df.iloc[0]
    df = df.drop(0)

    #change object to datetime64[ns]
    if (bookName in intraDayList):
        df.Datetime = pd.to_datetime(df.Datetime.astype('datetime64[ns]'))
        df = df.set_index('Datetime')
    elif (bookName in swingDayList):
        df.Date = pd.to_datetime(df.Date.astype('datetime64[ns]'))
        df = df.set_index('Date')

    #change object to float
    # Link https://stackoverflow.com/questions/36814100/pandas-to-numeric-for-multiple-columns
    cols = df.columns
    df[cols] = df[cols].apply(pd.to_numeric, errors='coerce')


    return df


def funIBKRimport (filePath_fun, bookName = 'IBKR 1m') :

    # Set Variable    
    intraDayList = ['IBKR 1s','IBKR 5s','IBKR 10s','IBKR 15s','IBKR 30s',
                    'IBKR 1m','IBKR 2m','IBKR 3m','IBKR 5m','IBKR 10m','IBKR 15m','IBKR 20m','IBKR 30m',
                    'IBKR 1H','IBKR 2H','IBKR 3H','IBKR 4H','IBKR 8H' ]
    swingDayList = ['IBKR 1Day','IBKR 1week','IBKR 1Month']
    
    # Import  data
    #------------------------------------------------  
    book = openpyxl.load_workbook(filePath_fun)
    sheet = book[bookName]
    df= pd.DataFrame(sheet.values)


    # format  data
    #------------------------------------------------  
    #Convert row to column header for Pandas DataFrame : Link https://stackoverflow.com/questions/26147180/convert-row-to-column-header-for-pandas-dataframe
    df.columns = df.iloc[0]
    df = df.drop(0)

    #change object to datetime64[ns]
    if (bookName in intraDayList):
        df.Datetime = df.Datetime.str.replace("US/Eastern", "")
        df.Datetime = pd.to_datetime(df.Datetime.astype('datetime64[ns]'))
        df = df.set_index('Datetime')
    elif (bookName in swingDayList):
        df.Datetime = pd.to_datetime(df.Datetime.astype('datetime64[ns]'))
        df = df.set_index('Datetime')

    #change object to float
    # Link https://stackoverflow.com/questions/36814100/pandas-to-numeric-for-multiple-columns
    cols = df.columns
    df[cols] = df[cols].apply(pd.to_numeric, errors='coerce')


    return df


#_________________________________________________________________________________________________________________________________________________________


def fGetYahooFundamentalDataFrame(    
            filePath_fun      = r"D:\Python Tools\ChartMaker\SourceDocuments\OutPut_Excel\Session [20231128_111940]\GME\OK GME.xlsx"
            ,bookName     = "Yahoo Fundamentals"                 
        ):

            #----------------------------------------------------------------------------------------------------------------------------------------------     
            # Time Stamp sheet:- Import  data and Set column header 
            #   1-Import  data
            book = openpyxl.load_workbook(filePath_fun)
            sheet = book[bookName]
            dfS= pd.DataFrame(sheet.values)
            #   2-Set column header :- Convert row to column header for Pandas DataFrame : Link https://stackoverflow.com/questions/26147180/convert-row-to-column-header-for-pandas-dataframe
            dfS.columns = dfS.iloc[0]
            dfS = dfS.drop(0)
            #----------------------------------------------------------------------------------------------------------------------------------------------     
            return dfS



if(False):
    fExcelPath = r"D:\Python Tools\ChartMaker\SourceDocuments\OutPut_Excel\Session [20231128_111940]\GME\OK GME.xlsx"
    dfS = None
    
    dfF = fGetYahooFundamentalDataFrame  (filePath_fun = fExcelPath ,bookName = "Yahoo Fundamentals" )    # "Yahoo Fundamentals"
    dfD = funYahooImport  (filePath_fun = fExcelPath ,bookName = "Yahoo Dayes" )    # "Yahoo Dayes"
    dfT = funIBKRimport   (filePath_fun = fExcelPath ,bookName = 'IBKR 30m')        # "IBKR 30m"
    dfM = funIBKRimport   (filePath_fun = fExcelPath ,bookName = "IBKR 1m")         # "IBKR 1m" 
    
    print(dfF)
    print("_______symbol_______")
    EDay = str( dfD.index[-1])
    Z_Date                  = EDay
    
    Z_Ticker                        =       (dfF.Data[dfF[None] == "symbol"].values[0])
    Z_Company_Name                  =       (dfF.Data[dfF[None] == "shortName"].values[0])
    Z_Activity_Sector               =       (dfF.Data[dfF[None] == "sector"].values[0])
    Z_Issuer_Country                =       (dfF.Data[dfF[None] == "country"].values[0])
    Z_Market_Cap_Million            =       (dfF.Data[dfF[None] == "marketCap"].values[0])
    Z_Float_Million                 =       (dfF.Data[dfF[None] == "floatShares"].values[0])
    Z_Short_Interest                =       (dfF.Data[dfF[None] == "shortRatio"].values[0])                    # Need to consult Rabi3
    Z_ATR                           =       "0" 
    
    Z_Recent_Reverse_Split          = int   (dfF.Data[dfF[None] == "lastSplitDate"     ].values[0])
    Z_Recent_Reverse_Split          = time.strftime("%Y/%m/%d ", time.localtime(Z_Recent_Reverse_Split))    # Need to consult Rabi3

    Z_Active_Shelf_Registration     = int (dfF.Data[dfF[None] == "firstTradeDateEpochUtc"].values[0])  
    Z_Active_Shelf_Registration     = time.strftime("%Y/%m/%d ", time.localtime(Z_Active_Shelf_Registration))    # Need to consult Rabi3                                                                         # Need to consult Rabi3
    
    Z_Catalyst                      =       "0"         # Need to consult Rabi3  
    
    #Z_Pre_Runner   # Need to consult Rabi3
    EDay = str( dfD.index[-2])
    SDay = str( dfD.index[-61])
    df = dfD.loc[SDay:EDay] #XXXXXXXX
    AvrVol = (dfF.Data[dfF[None] == "averageVolume"].values[0])   
    CHEKO = df.loc [    (df["Volume"]>=(2*AvrVol)) 
                    &   ((2*df["Open"])>=(df["High"]-df["Low"]))
                    &   (df["Open"] != df["High"])
                    ]
    # print("_______IF_STATEMENT_______")        
    # print(CHEKO)
    if(len(CHEKO.index)>0): Z_Pre_Runner  = "yes"    
    else:                   Z_Pre_Runner  = "No"  
    
    
    Z_Previous_Day_High =  dfD['High' ][-2] 
    Z_Previous_Day_Close = dfD['Close'][-2] 

    EDay = str( dfD.index[-1])                  # Need More Test
    STime=EDay[0:11] + "04:00:00"
    df = dfT.loc[STime] 
    Z_price_0400am =  df.High

    Z_Low_price_before_first_move = "0" # Need to consult Rabi3
    Z_Low_time_before_first_move = "0"  # Need to consult Rabi3
    Z_First_Dip = "0"                   # Need to consult Rabi3

    EDay = str( dfD.index[-1])
    ETime=EDay[0:11] + "09:29:00"  
    STime=EDay[0:11] + "04:00:00"
    df = dfM.loc[STime:ETime] #XXXXXXXX
    Z_PreMarket_High_Level = df['High'].max()
    Z_PreMarket_High_Time = df.loc[df['High']==df['High'].max()].index[0]
    Z_PreMarket_High_Time = str(Z_PreMarket_High_Time)[-8:]

    Z_Market_Open_Level = dfD['Open' ][-1] 

    Z_Gap = ((dfD['Open' ][-1] - dfD['Close'][-2])/dfD['Close'][-2])*100

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "09:30:00"
    ETime=EDay[0:11] + "15:59:00"      
    df = dfM.loc[STime:ETime] #XXXXXXXX
    Z_IntraDay_High_Level = df['High'].max()
    Z_IntraDay_High_Time = df.loc[df['High']==df['High'].max()].index[0]
    Z_IntraDay_High_Time = str(Z_IntraDay_High_Time)[-8:]

    STime = df.loc[df['High']==df['High'].max()].index[0]
    ETime=EDay[0:11] + "19:59:00"
    df = dfM.loc[STime:ETime] #XXXXXXXX
    Z_Low_price_after_IntraDay_High = df['Low'].min()
    Z_Low_time_after_IntraDay_High = df.loc[df['Low']==df['Low'].min()].index[0]
    Z_Low_time_after_IntraDay_High = str(Z_Low_time_after_IntraDay_High)[-8:]
    Z_Sell_off = ((Z_IntraDay_High_Level - Z_Low_price_after_IntraDay_High)/Z_IntraDay_High_Level)*100  # Need to consult Rabi3

    Z_Day_Close_Level = dfD['Close'][-1] 

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "16:00:00"
    ETime=EDay[0:11] + "19:59:00"      
    df = dfM.loc[STime:ETime] #XXXXXXXX
    Z_After_Hours_High_Level = df['High'].max()
    Z_After_Hours_High_Time = df.loc[df['High']==df['High'].max()].index[0]
    Z_After_Hours_High_Time = str(Z_After_Hours_High_Time)[-8:]

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "04:00:00"
    ETime=EDay[0:11] + "20:00:00"      
    df = dfT.loc[STime:ETime] #XXXXXXXX
    Z_Max_Gain__From_Previous_Day = ((df['High'].max() - dfD['Close'][-1])/dfD['Close'][-1])*100

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "09:30:00"
    ETime=EDay[0:11] + "20:00:00"      
    df = dfT.loc[STime:ETime] #XXXXXXXX
    Z_Max_Gain__From_Market_Open = ((df['High'].max() - dfD['Close'][-1])/dfD['Close'][-1])*100
    
    Z_Halts_to_the_up_side  ="No"   # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_Halts_to_the_down_side ="No"  # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "04:00:00"
    ETime=EDay[0:11] + "06:30:00"      
    df = dfT.loc[STime:ETime] #XXXXXXXX
    Z_Volume_0700am = df['Volume'].sum()
    if(Z_Float_Million):    Z_float_rotation_0700_am = Z_Volume_0700am / Z_Float_Million
    else:                   Z_float_rotation_0700_am = "0"

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "04:00:00"
    ETime=EDay[0:11] + "09:00:00"      
    df = dfT.loc[STime:ETime] #XXXXXXXX
    Z_Volume_0930am = df['Volume'].sum()
    if(Z_Float_Million):Z_float_rotation_0930_am = Z_Volume_0930am / Z_Float_Million
    else:               Z_float_rotation_0930_am = "0"

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "04:00:00"
    ETime=EDay[0:11] + "20:00:00"      
    df = dfT.loc[STime:ETime] #XXXXXXXX
    Z_Volume_Full_Day = df['Volume'].sum()
    if(Z_Float_Million):Z_float_rotation_full_day = Z_Volume_Full_Day / Z_Float_Million
    else:               Z_float_rotation_full_day = "0"

    Z_move_0700_am                  = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_PreMarket_move_Non_0700_am    = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_move_0930_am                  = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_IntraDay_move_After_1000_am   = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_After_Hours_move              = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_breakouts_5min                = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_breakouts_1min                = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase

    Z_General_Comment               = "-"     # Need to consult Rabi3
    Z_Chart_Screenshot              = "-"     # Need to consult Rabi3


    #__________________________________________________________________________________________________ 
    # Add one row in an existing Pandas DataFrame   Help Link :- https://www.geeksforgeeks.org/how-to-add-one-row-in-an-existing-pandas-dataframe/
    dfN =   {
                 "Date"                             :   Z_Date  
                ,"Ticker"                           :   Z_Ticker
                ,"Company Name"                     :   Z_Company_Name
                ,"Activity - Sector"                :   Z_Activity_Sector
                ,"Issuer Country"                   :   Z_Issuer_Country
                ,"Market Cap (Million)"             :	Z_Market_Cap_Million
                ,"Float (Million)"                  :	Z_Float_Million
                ,"Short Interest %"                 :	Z_Short_Interest
                ,"ATR"                              :	Z_ATR
                ,"Recent Reverse Split"             :	Z_Recent_Reverse_Split
                ,"Active Shelf Registration"        :	Z_Active_Shelf_Registration
                ,"Catalyst"                         :	Z_Catalyst
                ,"Pre Runner"                       :	Z_Pre_Runner
                ,"Previous Day High"                :	Z_Previous_Day_High
                ,"Previous Day Close"               :	Z_Previous_Day_Close
                ,"4:00am price"                     :	Z_price_0400am
                ,"Low price before first move"      :	Z_Low_price_before_first_move
                ,"Low time before first move"       :	Z_Low_time_before_first_move
                ,"First Dip %"                      :	Z_First_Dip
                ,"PreMarket High Level"             :	Z_PreMarket_High_Level
                ,"PreMarket High Time"              :	Z_PreMarket_High_Time
                ,"Market Open Level"                :	Z_Market_Open_Level
                ,"Gap %"                            :	Z_Gap
                ,"IntraDay High Level"              :	Z_IntraDay_High_Level
                ,"IntraDay High Time"               :	Z_IntraDay_High_Time
                ,"Low price after IntraDay High"    :	Z_Low_price_after_IntraDay_High
                ,"Low time after IntraDay High"     :	Z_Low_time_after_IntraDay_High
                ,"Sell off %"                       :	Z_Sell_off
                ,"Day Close Level"                  :	Z_Day_Close_Level
                ,"After Hours High Level"           :	Z_After_Hours_High_Level
                ,"After Hours High Time"            :	Z_After_Hours_High_Time
                ,"Max % Gain (From Previous Day)"   :	Z_Max_Gain__From_Previous_Day
                ,"Max % Gain (From Market Open)"    :	Z_Max_Gain__From_Market_Open
                ,"Halts to the up side"             :	Z_Halts_to_the_up_side
                ,"Halts to the down side"           :	Z_Halts_to_the_down_side
                ,"Volume (7am)"                     :	Z_Volume_0700am
                ,"7:00 am float rotation"           :	Z_float_rotation_0700_am
                ,"Volume (9:30am)"                  :	Z_Volume_0930am
                ,"9:30 am float rotation"           :	Z_float_rotation_0930_am 
                ,"Volume (Full Day)"                :	Z_Volume_Full_Day
                ,"full day float rotation"          :	Z_float_rotation_full_day 
                ,"7am move"                         :	Z_move_0700_am 
                ,"PreMarket move (Non 7am)"         :	Z_PreMarket_move_Non_0700_am
                ,"9:30am move"                      :	Z_move_0930_am 
                ,"IntraDay move (After 10am)"       :	Z_IntraDay_move_After_1000_am
                ,"After Hours move"                 :	Z_After_Hours_move
                ,"5min breakouts"                   :	Z_breakouts_5min 
                ,"1min breakouts"                   :	Z_breakouts_1min 
                ,"General Comment"                  :	Z_General_Comment
                ,"Chart Screenshot"                 :	Z_Chart_Screenshot

            }
    print(dfN)
    dfS = dfS.append(dfN, ignore_index = True)






def Zamzamsammry(
                    #  dfS
                    fExcelPath = r"D:\Python Tools\ChartMaker\SourceDocuments\OutPut_Excel\Session [20231128_111940]\GME\OK GME.xlsx"
                    ,showOutPut = False
                    
                ):
    
    
    dfF = fGetYahooFundamentalDataFrame  (filePath_fun = fExcelPath ,bookName = "Yahoo Fundamentals" )    # "Yahoo Fundamentals"
    dfD = funYahooImport  (filePath_fun = fExcelPath ,bookName = "Yahoo Dayes" )    # "Yahoo Dayes"
    dfT = funIBKRimport   (filePath_fun = fExcelPath ,bookName = 'IBKR 30m')        # "IBKR 30m"
    dfM = funIBKRimport   (filePath_fun = fExcelPath ,bookName = "IBKR 1m")         # "IBKR 1m" 
    
    if(showOutPut):print(dfF)
    print("_______symbol_______")
    EDay = str( dfD.index[-1])
    Z_Date                  = EDay
    
    Z_Ticker                        =       (dfF.Data[dfF[None] == "symbol"].values[0])
    Z_Company_Name                  =       (dfF.Data[dfF[None] == "shortName"].values[0])
    Z_Activity_Sector               =       (dfF.Data[dfF[None] == "sector"].values[0])
    Z_Issuer_Country                =       (dfF.Data[dfF[None] == "country"].values[0])
    Z_Market_Cap_Million            =       (dfF.Data[dfF[None] == "marketCap"].values[0])
    Z_Float_Million                 =       (dfF.Data[dfF[None] == "floatShares"].values[0])
    Z_Short_Interest                =       (dfF.Data[dfF[None] == "shortRatio"].values[0])                    # Need to consult Rabi3
    Z_ATR                           =       "0" 
    
    Z_Recent_Reverse_Split          = int   (dfF.Data[dfF[None] == "lastSplitDate"     ].values[0])
    Z_Recent_Reverse_Split          = time.strftime("%Y/%m/%d ", time.localtime(Z_Recent_Reverse_Split))    # Need to consult Rabi3

    Z_Active_Shelf_Registration     = int (dfF.Data[dfF[None] == "firstTradeDateEpochUtc"].values[0])  
    Z_Active_Shelf_Registration     = time.strftime("%Y/%m/%d ", time.localtime(Z_Active_Shelf_Registration))    # Need to consult Rabi3                                                                         # Need to consult Rabi3
    
    Z_Catalyst                      =       "0"         # Need to consult Rabi3  
    
    #Z_Pre_Runner   # Need to consult Rabi3
    EDay = str( dfD.index[-2])
    SDay = str( dfD.index[-61])
    df = dfD.loc[SDay:EDay] #XXXXXXXX
    AvrVol = (dfF.Data[dfF[None] == "averageVolume"].values[0])   
    CHEKO = df.loc [    (df["Volume"]>=(2*AvrVol)) 
                    &   ((2*df["Open"])>=(df["High"]-df["Low"]))
                    &   (df["Open"] != df["High"])
                    ]
    if(showOutPut):print("_______IF_STATEMENT_______")        
    if(showOutPut):print(CHEKO)
    if(len(CHEKO.index)>0): Z_Pre_Runner  = "yes"    
    else:                   Z_Pre_Runner  = "No"  
    
    
    Z_Previous_Day_High =  dfD['High' ][-2] 
    Z_Previous_Day_Close = dfD['Close'][-2] 

    EDay = str( dfD.index[-1])                  # Need More Test
    STime=EDay[0:11] + "04:00:00"
    df = dfT.loc[STime] 
    Z_price_0400am =  df.High

    Z_Low_price_before_first_move = "0" # Need to consult Rabi3
    Z_Low_time_before_first_move = "0"  # Need to consult Rabi3
    Z_First_Dip = "0"                   # Need to consult Rabi3

    EDay = str( dfD.index[-1])
    ETime=EDay[0:11] + "09:29:00"  
    STime=EDay[0:11] + "04:00:00"
    df = dfM.loc[STime:ETime] #XXXXXXXX
    Z_PreMarket_High_Level = df['High'].max()
    Z_PreMarket_High_Time = df.loc[df['High']==df['High'].max()].index[0]
    Z_PreMarket_High_Time = str(Z_PreMarket_High_Time)[-8:]

    Z_Market_Open_Level = dfD['Open' ][-1] 

    Z_Gap = ((dfD['Open' ][-1] - dfD['Close'][-2])/dfD['Close'][-2])*100

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "09:30:00"
    ETime=EDay[0:11] + "15:59:00"      
    df = dfM.loc[STime:ETime] #XXXXXXXX
    Z_IntraDay_High_Level = df['High'].max()
    Z_IntraDay_High_Time = df.loc[df['High']==df['High'].max()].index[0]
    Z_IntraDay_High_Time = str(Z_IntraDay_High_Time)[-8:]

    STime = df.loc[df['High']==df['High'].max()].index[0]
    ETime=EDay[0:11] + "19:59:00"
    df = dfM.loc[STime:ETime] #XXXXXXXX
    Z_Low_price_after_IntraDay_High = df['Low'].min()
    Z_Low_time_after_IntraDay_High = df.loc[df['Low']==df['Low'].min()].index[0]
    Z_Low_time_after_IntraDay_High = str(Z_Low_time_after_IntraDay_High)[-8:]
    Z_Sell_off = ((Z_IntraDay_High_Level - Z_Low_price_after_IntraDay_High)/Z_IntraDay_High_Level)*100  # Need to consult Rabi3

    Z_Day_Close_Level = dfD['Close'][-1] 

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "16:00:00"
    ETime=EDay[0:11] + "19:59:00"      
    df = dfM.loc[STime:ETime] #XXXXXXXX
    Z_After_Hours_High_Level = df['High'].max()
    Z_After_Hours_High_Time = df.loc[df['High']==df['High'].max()].index[0]
    Z_After_Hours_High_Time = str(Z_After_Hours_High_Time)[-8:]

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "04:00:00"
    ETime=EDay[0:11] + "20:00:00"      
    df = dfT.loc[STime:ETime] #XXXXXXXX
    Z_Max_Gain__From_Previous_Day = ((df['High'].max() - dfD['Close'][-1])/dfD['Close'][-1])*100

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "09:30:00"
    ETime=EDay[0:11] + "20:00:00"      
    df = dfT.loc[STime:ETime] #XXXXXXXX
    Z_Max_Gain__From_Market_Open = ((df['High'].max() - dfD['Close'][-1])/dfD['Close'][-1])*100
    
    Z_Halts_to_the_up_side  ="No"   # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_Halts_to_the_down_side ="No"  # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "04:00:00"
    ETime=EDay[0:11] + "06:30:00"      
    df = dfT.loc[STime:ETime] #XXXXXXXX
    Z_Volume_0700am = df['Volume'].sum()
    if(Z_Float_Million):    Z_float_rotation_0700_am = Z_Volume_0700am / Z_Float_Million
    else:                   Z_float_rotation_0700_am = "0"

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "04:00:00"
    ETime=EDay[0:11] + "09:00:00"      
    df = dfT.loc[STime:ETime] #XXXXXXXX
    Z_Volume_0930am = df['Volume'].sum()
    if(Z_Float_Million):Z_float_rotation_0930_am = Z_Volume_0930am / Z_Float_Million
    else:               Z_float_rotation_0930_am = "0"

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "04:00:00"
    ETime=EDay[0:11] + "20:00:00"      
    df = dfT.loc[STime:ETime] #XXXXXXXX
    Z_Volume_Full_Day = df['Volume'].sum()
    if(Z_Float_Million):Z_float_rotation_full_day = Z_Volume_Full_Day / Z_Float_Million
    else:               Z_float_rotation_full_day = "0"

    Z_move_0700_am                  = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_PreMarket_move_Non_0700_am    = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_move_0930_am                  = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_IntraDay_move_After_1000_am   = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_After_Hours_move              = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_breakouts_5min                = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_breakouts_1min                = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase

    Z_General_Comment               = "-"     # Need to consult Rabi3
    Z_Chart_Screenshot              = "-"     # Need to consult Rabi3


    #__________________________________________________________________________________________________ 
    # Add one row in an existing Pandas DataFrame   Help Link :- https://www.geeksforgeeks.org/how-to-add-one-row-in-an-existing-pandas-dataframe/
    dfN =   pd.DataFrame(
            {
                 "Date"                             :   Z_Date  
                ,"Ticker"                           :   Z_Ticker
                ,"Company Name"                     :   Z_Company_Name
                ,"Activity - Sector"                :   Z_Activity_Sector
                ,"Issuer Country"                   :   Z_Issuer_Country
                ,"Market Cap (Million)"             :	Z_Market_Cap_Million
                ,"Float (Million)"                  :	Z_Float_Million
                ,"Short Interest %"                 :	Z_Short_Interest
                ,"ATR"                              :	Z_ATR
                ,"Recent Reverse Split"             :	Z_Recent_Reverse_Split
                ,"Active Shelf Registration"        :	Z_Active_Shelf_Registration
                ,"Catalyst"                         :	Z_Catalyst
                ,"Pre Runner"                       :	Z_Pre_Runner
                ,"Previous Day High"                :	Z_Previous_Day_High
                ,"Previous Day Close"               :	Z_Previous_Day_Close
                ,"4:00am price"                     :	Z_price_0400am
                ,"Low price before first move"      :	Z_Low_price_before_first_move
                ,"Low time before first move"       :	Z_Low_time_before_first_move
                ,"First Dip %"                      :	Z_First_Dip
                ,"PreMarket High Level"             :	Z_PreMarket_High_Level
                ,"PreMarket High Time"              :	Z_PreMarket_High_Time
                ,"Market Open Level"                :	Z_Market_Open_Level
                ,"Gap %"                            :	Z_Gap
                ,"IntraDay High Level"              :	Z_IntraDay_High_Level
                ,"IntraDay High Time"               :	Z_IntraDay_High_Time
                ,"Low price after IntraDay High"    :	Z_Low_price_after_IntraDay_High
                ,"Low time after IntraDay High"     :	Z_Low_time_after_IntraDay_High
                ,"Sell off %"                       :	Z_Sell_off
                ,"Day Close Level"                  :	Z_Day_Close_Level
                ,"After Hours High Level"           :	Z_After_Hours_High_Level
                ,"After Hours High Time"            :	Z_After_Hours_High_Time
                ,"Max % Gain (From Previous Day)"   :	Z_Max_Gain__From_Previous_Day
                ,"Max % Gain (From Market Open)"    :	Z_Max_Gain__From_Market_Open
                ,"Halts to the up side"             :	Z_Halts_to_the_up_side
                ,"Halts to the down side"           :	Z_Halts_to_the_down_side
                ,"Volume (7am)"                     :	Z_Volume_0700am
                ,"7:00 am float rotation"           :	Z_float_rotation_0700_am
                ,"Volume (9:30am)"                  :	Z_Volume_0930am
                ,"9:30 am float rotation"           :	Z_float_rotation_0930_am 
                ,"Volume (Full Day)"                :	Z_Volume_Full_Day
                ,"full day float rotation"          :	Z_float_rotation_full_day 
                ,"7am move"                         :	Z_move_0700_am 
                ,"PreMarket move (Non 7am)"         :	Z_PreMarket_move_Non_0700_am
                ,"9:30am move"                      :	Z_move_0930_am 
                ,"IntraDay move (After 10am)"       :	Z_IntraDay_move_After_1000_am
                ,"After Hours move"                 :	Z_After_Hours_move
                ,"5min breakouts"                   :	Z_breakouts_5min 
                ,"1min breakouts"                   :	Z_breakouts_1min 
                ,"General Comment"                  :	Z_General_Comment
                ,"Chart Screenshot"                 :	Z_Chart_Screenshot

            }, index=[0])
    if(showOutPut):print(dfN)
    # dfS = dfS.append(dfN, ignore_index = True)
    return dfN



if (True):
    filePath_Distnation = r"D:\Python Tools\ChartMaker\SourceDocuments\OutPut_Excel"        # filePath_Distnation = r"D:\Python Tools\ChartMaker\SourceDocuments\OutPut_Excel\gv "

    filePath_fun  = r"D:\Python Tools\ChartMaker\SourceDocuments\InPut_Excel\Watch_List_Zamzam .xlsx"
    bookName      = "RL"                 


    # datetime object containing current date and time --- Help Link :- https://www.programiz.com/python-programming/datetime/current-datetime#google_vignette
    now = datetime.now()
    dt_string = now.strftime("[%Y%m%d_%H%M%S]") # YYmmdd_HHMMSS

    # Create Session 
    sessionSymbol =  "Zamzam_Session" + ' ' + str(dt_string)
    pathSession = os.path.join(filePath_Distnation, sessionSymbol)
    # print(pathFile)
    os.mkdir(pathSession)



    #----------------------------------------------------------------------------------------------------------------------------------------------     
    # Time Stamp sheet:- Import  data and Set column header 
    #   1-Import  data
    book = openpyxl.load_workbook(filePath_fun)
    sheet = book[bookName]
    watchList_xl= pd.DataFrame(sheet.values)
    #   2-Set column header :- Convert row to column header for Pandas DataFrame : Link https://stackoverflow.com/questions/26147180/convert-row-to-column-header-for-pandas-dataframe
    watchList_xl.columns = watchList_xl.iloc[0]
    watchList_xl = watchList_xl.drop(0)
    print(watchList_xl)

    # dfT = funIBKRimport (filePath_fun = watchList_xl.Excel_Link[2], bookName = 'IBKR 1m')
    # print(dfT)

    for TikerIndex in watchList_xl.index:
        print (TikerIndex,watchList_xl.Excel_Link[TikerIndex])
        
        # Make file
        pathFile = os.path.join(pathSession, watchList_xl.Symbol[TikerIndex])
        # print(pathFile)
        os.mkdir(pathFile)
    
        GetChart.fun9(   # Set Varibles
            #------------------------------------------------
                 tickerName = watchList_xl.Symbol[TikerIndex]                    
                ,filePathExcel = watchList_xl.Excel_Link[TikerIndex]
                ,filePathChart = pathFile    #r'D:\Python Tools\ChartMaker\SourceDocuments\OutPut_jpg'            #C:\Users\lenovo\Desktop\Python Project\Ehab\Results\Chart test.jpg
                
                ,filePathWL = r"D:\Python Tools\ChartMaker\SourceDocuments\InPut_Excel\Watch_List_Zamzam .xlsx"
                ,timeStampSheet = watchList_xl.Style[TikerIndex]
                
                ,flagYahoo = False
                ,flagIBKR = True
                ,imageType= '.png'
        )


    #______________________________________________________________
    #Make sammry Sheet
    dfTotallSammru = pd.DataFrame(columns=[
                                     "Date"                             
                                    ,"Ticker"                           
                                    ,"Company Name"                     
                                    ,"Activity - Sector"                
                                    ,"Issuer Country"                  
                                    ,"Market Cap (Million)"             
                                    ,"Float (Million)"                  
                                    ,"Short Interest %"                 
                                    ,"ATR"                              
                                    ,"Recent Reverse Split"             
                                    ,"Active Shelf Registration"        
                                    ,"Catalyst"                         
                                    ,"Pre Runner"                       
                                    ,"Previous Day High"                
                                    ,"Previous Day Close"               
                                    ,"4:00am price"                     
                                    ,"Low price before first move"      
                                    ,"Low time before first move"       
                                    ,"First Dip %"                      
                                    ,"PreMarket High Level"             
                                    ,"PreMarket High Time"              
                                    ,"Market Open Level"                
                                    ,"Gap %"                            
                                    ,"IntraDay High Level"              
                                    ,"IntraDay High Time"               
                                    ,"Low price after IntraDay High"    
                                    ,"Low time after IntraDay High"     
                                    ,"Sell off %"                       
                                    ,"Day Close Level"                  
                                    ,"After Hours High Level"           
                                    ,"After Hours High Time"            
                                    ,"Max % Gain (From Previous Day)"   
                                    ,"Max % Gain (From Market Open)"    
                                    ,"Halts to the up side"             
                                    ,"Halts to the down side"           
                                    ,"Volume (7am)"                     
                                    ,"7:00 am float rotation"           
                                    ,"Volume (9:30am)"                  
                                    ,"9:30 am float rotation"           
                                    ,"Volume (Full Day)"                
                                    ,"full day float rotation"          
                                    ,"7am move"                         
                                    ,"PreMarket move (Non 7am)"         
                                    ,"9:30am move"                      
                                    ,"IntraDay move (After 10am)"       
                                    ,"After Hours move"                 
                                    ,"5min breakouts"                   
                                    ,"1min breakouts"                   
                                    ,"General Comment"                  
                                    ,"Chart Screenshot"                 




                                        ])
    
    for TikerIndex in watchList_xl.index:
        dfs = Zamzamsammry  (    fExcelPath =  watchList_xl.Excel_Link[TikerIndex]    # watchList_xl.Style[TikerIndex]
                                ,showOutPut = False
                            )
        dfTotallSammru = dfTotallSammru.append(dfs, ignore_index = True)


    print(dfTotallSammru)
    filePathTicker = pathSession + "\Zamzomfinal" + ".xlsx"
    dfTotallSammru.to_excel( filePathTicker , "Sammry")
        



















