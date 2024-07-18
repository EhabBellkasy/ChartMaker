
# import the libraries

import datetime as dt
from datetime import date
import pandas as pd
import os
from datetime import datetime

import openpyxl
import time


import sys # Link:- https://stackoverflow.com/questions/4383571/importing-files-from-different-folder
sys.path.append('D:\Python Tools\ChartMaker\GetData')
import GetData
sys.path.append('D:\Python Tools\ChartMaker\GetChart')
import GetChart 
sys.path.append("D:\Python Tools\ChartMaker\IBKR\DataIBKR")
import DataIBKR 
sys.path.append("D:\Python Tools\ChartMaker\WebScraping")
import WebScraping


if(False):

   # set variables:

   filePathWL = r"D:\Python Tools\ChartMaker\SourceDocuments\InPut_Excel\Watch_List.xlsx"
   filePath_Distnation = r"D:\Python Tools\ChartMaker\SourceDocuments\OutPut_Excel"        # filePath_Distnation = r"D:\Python Tools\ChartMaker\SourceDocuments\OutPut_Excel\gv "






   # datetime object containing current date and time --- Help Link :- https://www.programiz.com/python-programming/datetime/current-datetime#google_vignette
   now = datetime.now()
   dt_string = now.strftime("[%Y%m%d_%H%M%S]") # YYmmdd_HHMMSS


   IBKR=DataIBKR.IBKR()

   # import Watch List from excel 
   print ("import watch List from excel ")
   watchList_xl = pd.read_excel(filePathWL, keep_default_na=False)            # Ticker "NA" Problem Solve Help Link :- https://stackoverflow.com/questions/33952142/prevent-pandas-from-interpreting-na-as-nan-in-a-string
   tickersWL = watchList_xl.Symbol.to_list()
   watchList_xl = watchList_xl.set_index('Symbol')


   # Create Session 
   sessionSymbol =  "Session" + ' ' + str(dt_string)
   pathSession = os.path.join(filePath_Distnation, sessionSymbol)
   # print(pathFile)
   os.mkdir(pathSession)

   for ticker_index in tickersWL :
      print (F' Ticker is : {ticker_index}')

      END_day = watchList_xl.Date[ticker_index]
      # print (F' End Day is : {END_day}')

      END_day_yahoo = END_day + dt.timedelta(days=1)
      # print (F' End Day for yahoo is : {END_day_yahoo}')

      END_day_IBKR = str(END_day)[0:10]
      END_day_IBKR = END_day_IBKR.replace('-','')
      # print (F' End Day for IBKR is : {END_day_IBKR}')
      
      IBKR.qSymbol = ticker_index
      IBKR.qEndDate = END_day_IBKR

      # Make file
      pathFile = os.path.join(pathSession, ticker_index)
      # print(pathFile)
      os.mkdir(pathFile)

      excelSymbol = "OK "
      pathExcel = os.path.join(pathFile, excelSymbol)
      # print(pathExcel+'#')

      GetData.fun(    ticker = ticker_index,
                     END = END_day_yahoo,
                     filePath = pathExcel , 
                     yahoo_Fundamentals = True,
                     yahoo_NEWS = True,
                     yahoo_Days = True,
                     daysNumber = 380,
                     yahoo_Hour = True,
                     daysHours = 30,
                     yahoo_30Min = True,
                     days30Min = 20,
                     yahoo_15Min = True,
                     days15Min = 10,
                     yahoo_5Min = True,
                     days5Min = 10,
                     yahoo_2Min = True,
                     days2Min = 10,
                     yahoo_1Min = True,
                     days1Min = 7
                  )
      
      IBKR.fMakePackage   (   nfMakeAdd=False, # True: Make / False: Add
                              nfFilePath = pathExcel,
                              nfSheetList =[  '1M','1W','1 day',
                                          '8 hours','4 hours','3 hours','2 hours','1 hour',
                                          '30 mins','20 mins','15 mins','10 mins','5 mins','3 mins','2 mins','1 min',
                                          '30 secs','15 secs','10 secs','5 secs'#,'1 secs'
                                       ],
                              nfConfigSet ={
                                       '1 secs' :{'SizeIndex': 0, 'DurationIndex':1, 'DurationLenth': 1, 'Rolling':60, 'SheetName':'IBKR 1s'     },
                                       '5 secs' :{'SizeIndex': 1, 'DurationIndex':1, 'DurationLenth': 2, 'Rolling':60, 'SheetName':'IBKR 5s'     },
                                       '10 secs':{'SizeIndex': 2, 'DurationIndex':1, 'DurationLenth': 2, 'Rolling':30, 'SheetName':'IBKR 10s'    },
                                       '15 secs':{'SizeIndex': 3, 'DurationIndex':1, 'DurationLenth': 2, 'Rolling':20, 'SheetName':'IBKR 15s'    },
                                       '30 secs':{'SizeIndex': 4, 'DurationIndex':1, 'DurationLenth': 2, 'Rolling':10, 'SheetName':'IBKR 30s'    },
                                       '1 min'  :{'SizeIndex': 5, 'DurationIndex':1, 'DurationLenth': 2, 'Rolling': 5, 'SheetName':"IBKR 1m"     },
                                       '2 mins' :{'SizeIndex': 6, 'DurationIndex':1, 'DurationLenth': 2, 'Rolling': 5, 'SheetName':'IBKR 2m'     },
                                       '3 mins' :{'SizeIndex': 7, 'DurationIndex':1, 'DurationLenth': 2, 'Rolling':10, 'SheetName':'IBKR 3m'     },
                                       '5 mins' :{'SizeIndex': 8, 'DurationIndex':1, 'DurationLenth': 3, 'Rolling': 6, 'SheetName':'IBKR 5m'     },
                                       '10 mins':{'SizeIndex': 9, 'DurationIndex':1, 'DurationLenth': 9, 'Rolling': 6, 'SheetName':'IBKR 10m'    },
                                       '15 mins':{'SizeIndex':10, 'DurationIndex':1, 'DurationLenth': 9, 'Rolling':12, 'SheetName':'IBKR 15m'    },
                                       '20 mins':{'SizeIndex':11, 'DurationIndex':1, 'DurationLenth':12, 'Rolling':12, 'SheetName':'IBKR 20m'    },
                                       '30 mins':{'SizeIndex':12, 'DurationIndex':1, 'DurationLenth':12, 'Rolling':11, 'SheetName':'IBKR 30m'    },
                                       '1 hour' :{'SizeIndex':13, 'DurationIndex':1, 'DurationLenth':34, 'Rolling':16, 'SheetName':'IBKR 1H'     },
                                       '2 hours':{'SizeIndex':14, 'DurationIndex':1, 'DurationLenth':34, 'Rolling': 8, 'SheetName':'IBKR 2H'     },
                                       '3 hours':{'SizeIndex':15, 'DurationIndex':1, 'DurationLenth':34, 'Rolling':30, 'SheetName':'IBKR 3H'     },
                                       '4 hours':{'SizeIndex':16, 'DurationIndex':1, 'DurationLenth':34, 'Rolling':20, 'SheetName':'IBKR 4H'     },
                                       '8 hours':{'SizeIndex':17, 'DurationIndex':1, 'DurationLenth':34, 'Rolling':42, 'SheetName':'IBKR 8H'     },
                                       '1 day'  :{'SizeIndex':18, 'DurationIndex':4, 'DurationLenth': 5, 'Rolling':21, 'SheetName':'IBKR 1Day'   },
                                       '1W'     :{'SizeIndex':19, 'DurationIndex':4, 'DurationLenth': 5, 'Rolling':12, 'SheetName':'IBKR 1week'  },
                                       '1M'     :{'SizeIndex':20, 'DurationIndex':4, 'DurationLenth': 5, 'Rolling':12, 'SheetName':'IBKR 1Month' }
                                       }
                           )


      #  print(pathExcel)
      #  print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
      #  print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
      #  print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
      #  print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
      GetData.funSamaryUpdate( ticker = ticker_index
                              ,fExcelPath = pathExcel
                              ,sheetName = 'Samary'
                              )

      IBKR.fFundamentalData(  fnFilePath = pathExcel,                         
                              fnFandaList = ['ReportSnapshot','ReportsFinSummary','ReportRatios','ReportsFinStatements','RESC']
                           )

      WebScraping.fun(webScrapingTiker=ticker_index, webScrapingDisPath= pathExcel)

      # GetChart.fun(   tickerName = ticker_index,     
      #                 filePathExcel = pathExcel + ticker_index + '.xlsx',
      #                 filePathChart = pathFile,            #C:\Users\lenovo\Desktop\Python Project\Ehab\Results\Chart test.jpg
      #                 daySheet = "Yahoo Dayes",
      #                 imageType= '.png'
      #             )
      
      # GetChart.fun2   (   # Set Varibles
      #                         #------------------------------------------------
      #                         tickerName      = ticker_index,     
      #                         filePathExcel   = pathExcel + ticker_index + '.xlsx',
      #                         filePathChart   = pathFile,            #C:\Users\lenovo\Desktop\Python Project\Ehab\Results\Chart test.jpg
      #                         flagYahoo       = True,
      #                         flagIBKR        = True,
      #                         imageType       = '.png'
      #                 )

      GetChart.fun8   (   # Set Varibles
                           #------------------------------------------------
                           tickerName      = ticker_index     
                        ,filePathExcel   = pathExcel + ticker_index + '.xlsx'
                        ,filePathChart   = pathFile            #C:\Users\lenovo\Desktop\Python Project\Ehab\Results\Chart test.jpg
                        ,flagYahoo       = False
                        ,flagIBKR        = True
                        ,imageType       = '.png'
                     )

   print (f' Watch list is : ')
   print (watchList_xl)
   print (F' Tickers are : {tickersWL}')
   print("------------------------")






#_________________________________________________________________________________________________________________________________________________________
def funSamaryImport (filePath_fun, bookName = 'Samary') :
    # Import  data
    #------------------------------------------------  
    book = openpyxl.load_workbook(filePath_fun)
    sheet = book[bookName]
    df= pd.DataFrame(sheet.values)
    # format  data
    #------------------------------------------------  
    #Convert row to column header for Pandas DataFrame : Link https://stackoverflow.com/questions/26147180/convert-row-to-column-header-for-pandas-dataframe
    df.columns = df.iloc[0]
    df = df.drop(0)
    df.reset_index(inplace=True)
    df = df.drop(columns=['index', None])
    return df


def funYahooImport (filePath_fun, bookName = "Yahoo 1m") :

    # Set Variable    
    intraDayList = ['Yahoo Hours','Yahoo 30m','Yahoo 15m','Yahoo 5m','Yahoo 2m','Yahoo 1m']
    swingDayList = ['Yahoo Dayes']
    
    # Import  data
    #------------------------------------------------  
    book = openpyxl.load_workbook(filePath_fun)
    sheet = book[bookName]
    df= pd.DataFrame(sheet.values)


    # format  data
    #------------------------------------------------  
    #Convert row to column header for Pandas DataFrame : Link https://stackoverflow.com/questions/26147180/convert-row-to-column-header-for-pandas-dataframe
    df.columns = df.iloc[0]
    df = df.drop(0)

    #change object to datetime64[ns]
    if (bookName in intraDayList):
        df.Datetime = pd.to_datetime(df.Datetime.astype('datetime64[ns]'))
        df = df.set_index('Datetime')
    elif (bookName in swingDayList):
        df.Date = pd.to_datetime(df.Date.astype('datetime64[ns]'))
        df = df.set_index('Date')

    #change object to float
    # Link https://stackoverflow.com/questions/36814100/pandas-to-numeric-for-multiple-columns
    cols = df.columns
    df[cols] = df[cols].apply(pd.to_numeric, errors='coerce')


    return df


def funIBKRimport (filePath_fun, bookName = 'IBKR 1m') :

    # Set Variable    
    intraDayList = ['IBKR 1s','IBKR 5s','IBKR 10s','IBKR 15s','IBKR 30s',
                    'IBKR 1m','IBKR 2m','IBKR 3m','IBKR 5m','IBKR 10m','IBKR 15m','IBKR 20m','IBKR 30m',
                    'IBKR 1H','IBKR 2H','IBKR 3H','IBKR 4H','IBKR 8H' ]
    swingDayList = ['IBKR 1Day','IBKR 1week','IBKR 1Month']
    
    # Import  data
    #------------------------------------------------  
    book = openpyxl.load_workbook(filePath_fun)
    sheet = book[bookName]
    df= pd.DataFrame(sheet.values)


    # format  data
    #------------------------------------------------  
    #Convert row to column header for Pandas DataFrame : Link https://stackoverflow.com/questions/26147180/convert-row-to-column-header-for-pandas-dataframe
    df.columns = df.iloc[0]
    df = df.drop(0)

    #change object to datetime64[ns]
    if (bookName in intraDayList):
        df.Datetime = df.Datetime.str.replace("US/Eastern", "")
        df.Datetime = pd.to_datetime(df.Datetime.astype('datetime64[ns]'))
        df = df.set_index('Datetime')
    elif (bookName in swingDayList):
        df.Datetime = pd.to_datetime(df.Datetime.astype('datetime64[ns]'))
        df = df.set_index('Datetime')

    #change object to float
    # Link https://stackoverflow.com/questions/36814100/pandas-to-numeric-for-multiple-columns
    cols = df.columns
    df[cols] = df[cols].apply(pd.to_numeric, errors='coerce')


    return df



def fGetYahooFundamentalDataFrame(    
            filePath_fun      = r"D:\Python Tools\ChartMaker\SourceDocuments\OutPut_Excel\Session [20231128_111940]\GME\OK GME.xlsx"
            ,bookName     = "Yahoo Fundamentals"                 
        ):

            #----------------------------------------------------------------------------------------------------------------------------------------------     
            # Time Stamp sheet:- Import  data and Set column header 
            #   1-Import  data
            book = openpyxl.load_workbook(filePath_fun)
            sheet = book[bookName]
            dfS= pd.DataFrame(sheet.values)
            #   2-Set column header :- Convert row to column header for Pandas DataFrame : Link https://stackoverflow.com/questions/26147180/convert-row-to-column-header-for-pandas-dataframe
            dfS.columns = dfS.iloc[0]
            dfS = dfS.drop(0)
            #----------------------------------------------------------------------------------------------------------------------------------------------     
            return dfS


#_________________________________________________________________________________________________________________________________________________________





def Zamzamsammry(
                    #  dfS
                    fExcelPath = r"D:\Python Tools\ChartMaker\SourceDocuments\OutPut_Excel\Session [20231128_111940]\GME\OK GME.xlsx"
                    ,showOutPut = False
                    
                ):
    
    
    dfF = fGetYahooFundamentalDataFrame  (filePath_fun = fExcelPath ,bookName = "Yahoo Fundamentals" )    # "Yahoo Fundamentals"
    dfD = funYahooImport  (filePath_fun = fExcelPath ,bookName = "Yahoo Dayes" )    # "Yahoo Dayes"
    dfT = funIBKRimport   (filePath_fun = fExcelPath ,bookName = 'IBKR 30m')        # "IBKR 30m"
    dfM = funIBKRimport   (filePath_fun = fExcelPath ,bookName = "IBKR 1m")         # "IBKR 1m" 
    
    if(showOutPut):print(dfF)
    print("_______symbol_______")
    EDay = str( dfD.index[-1])
    Z_Date                  = EDay
    
    Z_Ticker                        =       (dfF.Data[dfF[None] == "symbol"].values[0])
    Z_Company_Name                  =       (dfF.Data[dfF[None] == "shortName"].values[0])
    Z_Activity_Sector               =       (dfF.Data[dfF[None] == "sector"].values[0])
    Z_Issuer_Country                =       (dfF.Data[dfF[None] == "country"].values[0])
    Z_Market_Cap_Million            =       (dfF.Data[dfF[None] == "marketCap"].values[0])
    Z_Float_Million                 =       (dfF.Data[dfF[None] == "floatShares"].values[0])
    Z_Short_Interest                =       (dfF.Data[dfF[None] == "shortRatio"].values[0])                    # Need to consult Rabi3
    Z_ATR                           =       "0" 
    
    Z_Recent_Reverse_Split          = int   (dfF.Data[dfF[None] == "lastSplitDate"     ].values[0])
    Z_Recent_Reverse_Split          = time.strftime("%Y/%m/%d ", time.localtime(Z_Recent_Reverse_Split))    # Need to consult Rabi3

    Z_Active_Shelf_Registration     = int (dfF.Data[dfF[None] == "firstTradeDateEpochUtc"].values[0])  
    Z_Active_Shelf_Registration     = time.strftime("%Y/%m/%d ", time.localtime(Z_Active_Shelf_Registration))    # Need to consult Rabi3                                                                         # Need to consult Rabi3
    
    Z_Catalyst                      =       "0"         # Need to consult Rabi3  
    
    #Z_Pre_Runner   # Need to consult Rabi3
    EDay = str( dfD.index[-2])
    SDay = str( dfD.index[-61])
    df = dfD.loc[SDay:EDay] #XXXXXXXX
    AvrVol = (dfF.Data[dfF[None] == "averageVolume"].values[0])   
    CHEKO = df.loc [    (df["Volume"]>=(2*AvrVol)) 
                    &   ((2*df["Open"])>=(df["High"]-df["Low"]))
                    &   (df["Open"] != df["High"])
                    ]
    if(showOutPut):print("_______IF_STATEMENT_______")        
    if(showOutPut):print(CHEKO)
    if(len(CHEKO.index)>0): Z_Pre_Runner  = "yes"    
    else:                   Z_Pre_Runner  = "No"  
    
    
    Z_Previous_Day_High =  dfD['High' ][-2] 
    Z_Previous_Day_Close = dfD['Close'][-2] 

    EDay = str( dfD.index[-1])                  # Need More Test
    STime=EDay[0:11] + "04:00:00"
    df = dfT.loc[STime] 
    Z_price_0400am =  df.High

    Z_Low_price_before_first_move = "0" # Need to consult Rabi3
    Z_Low_time_before_first_move = "0"  # Need to consult Rabi3
    Z_First_Dip = "0"                   # Need to consult Rabi3

    EDay = str( dfD.index[-1])
    ETime=EDay[0:11] + "09:29:00"  
    STime=EDay[0:11] + "04:00:00"
    df = dfM.loc[STime:ETime] #XXXXXXXX
    Z_PreMarket_High_Level = df['High'].max()
    Z_PreMarket_High_Time = df.loc[df['High']==df['High'].max()].index[0]
    Z_PreMarket_High_Time = str(Z_PreMarket_High_Time)[-8:]

    Z_Market_Open_Level = dfD['Open' ][-1] 

    Z_Gap = ((dfD['Open' ][-1] - dfD['Close'][-2])/dfD['Close'][-2])*100

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "09:30:00"
    ETime=EDay[0:11] + "15:59:00"      
    df = dfM.loc[STime:ETime] #XXXXXXXX
    Z_IntraDay_High_Level = df['High'].max()
    Z_IntraDay_High_Time = df.loc[df['High']==df['High'].max()].index[0]
    Z_IntraDay_High_Time = str(Z_IntraDay_High_Time)[-8:]

    STime = df.loc[df['High']==df['High'].max()].index[0]
    ETime=EDay[0:11] + "19:59:00"
    df = dfM.loc[STime:ETime] #XXXXXXXX
    Z_Low_price_after_IntraDay_High = df['Low'].min()
    Z_Low_time_after_IntraDay_High = df.loc[df['Low']==df['Low'].min()].index[0]
    Z_Low_time_after_IntraDay_High = str(Z_Low_time_after_IntraDay_High)[-8:]
    Z_Sell_off = ((Z_IntraDay_High_Level - Z_Low_price_after_IntraDay_High)/Z_IntraDay_High_Level)*100  # Need to consult Rabi3

    Z_Day_Close_Level = dfD['Close'][-1] 

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "16:00:00"
    ETime=EDay[0:11] + "19:59:00"      
    df = dfM.loc[STime:ETime] #XXXXXXXX
    Z_After_Hours_High_Level = df['High'].max()
    Z_After_Hours_High_Time = df.loc[df['High']==df['High'].max()].index[0]
    Z_After_Hours_High_Time = str(Z_After_Hours_High_Time)[-8:]

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "04:00:00"
    ETime=EDay[0:11] + "20:00:00"      
    df = dfT.loc[STime:ETime] #XXXXXXXX
    Z_Max_Gain__From_Previous_Day = ((df['High'].max() - dfD['Close'][-1])/dfD['Close'][-1])*100

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "09:30:00"
    ETime=EDay[0:11] + "20:00:00"      
    df = dfT.loc[STime:ETime] #XXXXXXXX
    Z_Max_Gain__From_Market_Open = ((df['High'].max() - dfD['Close'][-1])/dfD['Close'][-1])*100
    
    Z_Halts_to_the_up_side  ="No"   # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_Halts_to_the_down_side ="No"  # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "04:00:00"
    ETime=EDay[0:11] + "06:30:00"      
    df = dfT.loc[STime:ETime] #XXXXXXXX
    Z_Volume_0700am = df['Volume'].sum()
    if(Z_Float_Million):    Z_float_rotation_0700_am = Z_Volume_0700am / Z_Float_Million
    else:                   Z_float_rotation_0700_am = "0"

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "04:00:00"
    ETime=EDay[0:11] + "09:00:00"      
    df = dfT.loc[STime:ETime] #XXXXXXXX
    Z_Volume_0930am = df['Volume'].sum()
    if(Z_Float_Million):Z_float_rotation_0930_am = Z_Volume_0930am / Z_Float_Million
    else:               Z_float_rotation_0930_am = "0"

    EDay = str( dfD.index[-1])
    STime=EDay[0:11] + "04:00:00"
    ETime=EDay[0:11] + "20:00:00"      
    df = dfT.loc[STime:ETime] #XXXXXXXX
    Z_Volume_Full_Day = df['Volume'].sum()
    if(Z_Float_Million):Z_float_rotation_full_day = Z_Volume_Full_Day / Z_Float_Million
    else:               Z_float_rotation_full_day = "0"

    Z_move_0700_am                  = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_PreMarket_move_Non_0700_am    = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_move_0930_am                  = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_IntraDay_move_After_1000_am   = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_After_Hours_move              = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_breakouts_5min                = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase
    Z_breakouts_1min                = "No"    # Need to consult Rabi3 ------- Solve in Pattern Recognition Phase

    Z_General_Comment               = "-"     # Need to consult Rabi3
    Z_Chart_Screenshot              = "-"     # Need to consult Rabi3


    #__________________________________________________________________________________________________ 
    # Add one row in an existing Pandas DataFrame   Help Link :- https://www.geeksforgeeks.org/how-to-add-one-row-in-an-existing-pandas-dataframe/
    dfN =   pd.DataFrame(
            {
                 "Date"                             :   Z_Date  
                ,"Ticker"                           :   Z_Ticker
                ,"Company Name"                     :   Z_Company_Name
                ,"Activity - Sector"                :   Z_Activity_Sector
                ,"Issuer Country"                   :   Z_Issuer_Country
                ,"Market Cap (Million)"             :	Z_Market_Cap_Million
                ,"Float (Million)"                  :	Z_Float_Million
                ,"Short Interest %"                 :	Z_Short_Interest
                ,"ATR"                              :	Z_ATR
                ,"Recent Reverse Split"             :	Z_Recent_Reverse_Split
                ,"Active Shelf Registration"        :	Z_Active_Shelf_Registration
                ,"Catalyst"                         :	Z_Catalyst
                ,"Pre Runner"                       :	Z_Pre_Runner
                ,"Previous Day High"                :	Z_Previous_Day_High
                ,"Previous Day Close"               :	Z_Previous_Day_Close
                ,"4:00am price"                     :	Z_price_0400am
                ,"Low price before first move"      :	Z_Low_price_before_first_move
                ,"Low time before first move"       :	Z_Low_time_before_first_move
                ,"First Dip %"                      :	Z_First_Dip
                ,"PreMarket High Level"             :	Z_PreMarket_High_Level
                ,"PreMarket High Time"              :	Z_PreMarket_High_Time
                ,"Market Open Level"                :	Z_Market_Open_Level
                ,"Gap %"                            :	Z_Gap
                ,"IntraDay High Level"              :	Z_IntraDay_High_Level
                ,"IntraDay High Time"               :	Z_IntraDay_High_Time
                ,"Low price after IntraDay High"    :	Z_Low_price_after_IntraDay_High
                ,"Low time after IntraDay High"     :	Z_Low_time_after_IntraDay_High
                ,"Sell off %"                       :	Z_Sell_off
                ,"Day Close Level"                  :	Z_Day_Close_Level
                ,"After Hours High Level"           :	Z_After_Hours_High_Level
                ,"After Hours High Time"            :	Z_After_Hours_High_Time
                ,"Max % Gain (From Previous Day)"   :	Z_Max_Gain__From_Previous_Day
                ,"Max % Gain (From Market Open)"    :	Z_Max_Gain__From_Market_Open
                ,"Halts to the up side"             :	Z_Halts_to_the_up_side
                ,"Halts to the down side"           :	Z_Halts_to_the_down_side
                ,"Volume (7am)"                     :	Z_Volume_0700am
                ,"7:00 am float rotation"           :	Z_float_rotation_0700_am
                ,"Volume (9:30am)"                  :	Z_Volume_0930am
                ,"9:30 am float rotation"           :	Z_float_rotation_0930_am 
                ,"Volume (Full Day)"                :	Z_Volume_Full_Day
                ,"full day float rotation"          :	Z_float_rotation_full_day 
                ,"7am move"                         :	Z_move_0700_am 
                ,"PreMarket move (Non 7am)"         :	Z_PreMarket_move_Non_0700_am
                ,"9:30am move"                      :	Z_move_0930_am 
                ,"IntraDay move (After 10am)"       :	Z_IntraDay_move_After_1000_am
                ,"After Hours move"                 :	Z_After_Hours_move
                ,"5min breakouts"                   :	Z_breakouts_5min 
                ,"1min breakouts"                   :	Z_breakouts_1min 
                ,"General Comment"                  :	Z_General_Comment
                ,"Chart Screenshot"                 :	Z_Chart_Screenshot

            }, index=[0])
    if(showOutPut):print(dfN)
    # dfS = dfS.append(dfN, ignore_index = True)
    return dfN



if (True):
    filePath_Distnation = r"D:\Python Tools\ChartMaker\SourceDocuments\OutPut_Excel"        # filePath_Distnation = r"D:\Python Tools\ChartMaker\SourceDocuments\OutPut_Excel\gv "

    filePath_fun  = r"D:\Python Tools\ChartMaker\SourceDocuments\InPut_Excel\Watch_List_Zamzam .xlsx"
    bookName      = "RL"                 


    # datetime object containing current date and time --- Help Link :- https://www.programiz.com/python-programming/datetime/current-datetime#google_vignette
    now = datetime.now()
    dt_string = now.strftime("[%Y%m%d_%H%M%S]") # YYmmdd_HHMMSS

    # Create Session 
    sessionSymbol =  "Zamzam_Session" + ' ' + str(dt_string)
    pathSession = os.path.join(filePath_Distnation, sessionSymbol)
    # print(pathFile)
    os.mkdir(pathSession)



    #----------------------------------------------------------------------------------------------------------------------------------------------     
    # Time Stamp sheet:- Import  data and Set column header 
    #   1-Import  data
    book = openpyxl.load_workbook(filePath_fun)
    sheet = book[bookName]
    watchList_xl= pd.DataFrame(sheet.values)
    #   2-Set column header :- Convert row to column header for Pandas DataFrame : Link https://stackoverflow.com/questions/26147180/convert-row-to-column-header-for-pandas-dataframe
    watchList_xl.columns = watchList_xl.iloc[0]
    watchList_xl = watchList_xl.drop(0)
    print(watchList_xl)

    # dfT = funIBKRimport (filePath_fun = watchList_xl.Excel_Link[2], bookName = 'IBKR 1m')
    # print(dfT)

    for TikerIndex in watchList_xl.index:
        print (TikerIndex,watchList_xl.Excel_Link[TikerIndex])
        
        # Make file
        pathFile = os.path.join(pathSession, watchList_xl.Symbol[TikerIndex])
        # print(pathFile)
        os.mkdir(pathFile)
    
        GetChart.fun9(   # Set Varibles
            #------------------------------------------------
                 tickerName = watchList_xl.Symbol[TikerIndex]                    
                ,filePathExcel = watchList_xl.Excel_Link[TikerIndex]
                ,filePathChart = pathFile    #r'D:\Python Tools\ChartMaker\SourceDocuments\OutPut_jpg'            #C:\Users\lenovo\Desktop\Python Project\Ehab\Results\Chart test.jpg
                
                ,filePathWL = r"D:\Python Tools\ChartMaker\SourceDocuments\InPut_Excel\Watch_List_Zamzam .xlsx"
                ,timeStampSheet = watchList_xl.Style[TikerIndex]
                
                ,flagYahoo = False
                ,flagIBKR = True
                ,imageType= '.png'
        )


    #______________________________________________________________
    #Make sammry Sheet
    dfTotallSammru = pd.DataFrame(columns=[
                                     "Date"                             
                                    ,"Ticker"                           
                                    ,"Company Name"                     
                                    ,"Activity - Sector"                
                                    ,"Issuer Country"                  
                                    ,"Market Cap (Million)"             
                                    ,"Float (Million)"                  
                                    ,"Short Interest %"                 
                                    ,"ATR"                              
                                    ,"Recent Reverse Split"             
                                    ,"Active Shelf Registration"        
                                    ,"Catalyst"                         
                                    ,"Pre Runner"                       
                                    ,"Previous Day High"                
                                    ,"Previous Day Close"               
                                    ,"4:00am price"                     
                                    ,"Low price before first move"      
                                    ,"Low time before first move"       
                                    ,"First Dip %"                      
                                    ,"PreMarket High Level"             
                                    ,"PreMarket High Time"              
                                    ,"Market Open Level"                
                                    ,"Gap %"                            
                                    ,"IntraDay High Level"              
                                    ,"IntraDay High Time"               
                                    ,"Low price after IntraDay High"    
                                    ,"Low time after IntraDay High"     
                                    ,"Sell off %"                       
                                    ,"Day Close Level"                  
                                    ,"After Hours High Level"           
                                    ,"After Hours High Time"            
                                    ,"Max % Gain (From Previous Day)"   
                                    ,"Max % Gain (From Market Open)"    
                                    ,"Halts to the up side"             
                                    ,"Halts to the down side"           
                                    ,"Volume (7am)"                     
                                    ,"7:00 am float rotation"           
                                    ,"Volume (9:30am)"                  
                                    ,"9:30 am float rotation"           
                                    ,"Volume (Full Day)"                
                                    ,"full day float rotation"          
                                    ,"7am move"                         
                                    ,"PreMarket move (Non 7am)"         
                                    ,"9:30am move"                      
                                    ,"IntraDay move (After 10am)"       
                                    ,"After Hours move"                 
                                    ,"5min breakouts"                   
                                    ,"1min breakouts"                   
                                    ,"General Comment"                  
                                    ,"Chart Screenshot"                 




                                        ])
    
    for TikerIndex in watchList_xl.index:
        dfs = Zamzamsammry  (    fExcelPath =  watchList_xl.Excel_Link[TikerIndex]    # watchList_xl.Style[TikerIndex]
                                ,showOutPut = False
                            )
        dfTotallSammru = dfTotallSammru.append(dfs, ignore_index = True)


    print(dfTotallSammru)
    filePathTicker = pathSession + "\Zamzomfinal" + ".xlsx"
    dfTotallSammru.to_excel( filePathTicker , "Sammry")
        
















































###_____________________________________________________________________________________________________________________________________________________________________________________
###_____________________________________________________________________________________________________________________________________________________________________________________
###_____________________________________________________________________________________________________________________________________________________________________________________
###_____________________________________________________________________________________________________________________________________________________________________________________                    
###_____________________________________________________________________________________________________________________________________________________________________________________





















































