import pandas as pd
import numpy as np
import math
import ImportData
import openpyxl










# 5sec 16min
'''
index1_5s_A = ["09:18:00","09:28:00","09:38:00","09:48:00","09:58:00","10:08:00","10:18:00","10:28:00","10:38:00","10:48:00","10:58:00","11:08:00","11:18:00","11:28:00","11:38:00","11:48:00"]
index2_5s_A = ["09:34:00","09:44:00","09:54:00","10:04:00","10:14:00","10:24:00","10:34:00","10:44:00","10:54:00","11:04:00","11:14:00","11:24:00","11:34:00","11:44:00","11:54:00","12:04:00"]
index3_5s_A = ["091800_to_093400","092800_to_094400","093800_to_095400"	,"094800_to_100400"	,"095800_to_101400"	,"100800_to_102400"	,"101800_to_103400"	,"102800_to_104400"	,"103800_to_105400"	,"104800_to_110400"	,"105800_to_111400"	,"110800_to_112400"	,"111800_to_113400"	,"112800_to_114400"	,"113800_to_115400"	,"114800_to_120400"	]
index4_5s_A = ["091800","092800","093800","094800","095800","100800","101800","102800","103800","104800","105800","110800","111800","112800","113800","114800"]
index5_5s_A = ["093400","094400","095400","100400","101400","102400","103400","104400","105400","110400","111400","112400","113400","114400","115400","120400"]
tDelta_5s_A = -1 #day
'''
index1_5s_A = ["09:28:00","09:18:00","09:28:00","09:38:00","09:48:00","09:58:00","10:08:00","10:18:00","10:28:00","10:38:00","10:48:00","10:58:00","11:08:00","11:18:00","11:28:00","11:38:00","11:48:00","11:58:00"]
index2_5s_A = ["10:01:00","09:34:00","09:44:00","09:54:00","10:04:00","10:14:00","10:24:00","10:34:00","10:44:00","10:54:00","11:04:00","11:14:00","11:24:00","11:34:00","11:44:00","11:54:00","12:04:00","12:14:00"]
index3_5s_A = ["093000_to_100000","091800_to_093400","092800_to_094400","093800_to_095400","094800_to_100400","095800_to_101400","100800_to_102400","101800_to_103400","102800_to_104400","103800_to_105400","104800_to_110400","105800_to_111400","110800_to_112400","111800_to_113400","112800_to_114400","113800_to_115400","114800_to_120400","115800_to_121400"]
index4_5s_A = ["093000","091800","092800","093800","094800","095800","100800","101800","102800","103800","104800","105800","110800","111800","112800","113800","114800","115800"]
index5_5s_A = ["100000","093400","094400","095400","100400","101400","102400","103400","104400","105400","110400","111400","112400","113400","114400","115400","120400","121400"]
tDelta_5s_A = -1 #day

# 5sec 30min
'''
index1_5s_B = ["08:00:00","08:30:00","09:00:00","09:30:00","10:00:00","10:30:00","11:00:00","11:30:00"]#"00:30","00:00",
index2_5s_B = ["08:30:00","09:00:00","09:30:00","10:00:00","10:30:00","11:00:00","11:30:00","12:00:00"] 
index3_5s_B = ["080000_to_083000","083000_to_090000","090000_to_093000","093000_to_100000","100000_to_103000","103000_to_110000","110000_to_113000","113000_to_120000"]
index4_5s_B = ["080000","083000","090000","093000","100000","103000","110000","113000"]
index5_5s_B = ["083000","090000","093000","100000","103000","110000","113000","120000"]
tDelta_5s_B = -1 #day
'''
index1_5s_B = ["07:59:00","08:29:00","08:59:00","09:29:00","09:59:00","10:29:00","10:59:00","11:29:00"]#"00:30","00:00",
index2_5s_B = ["08:31:00","09:01:00","09:31:00","10:01:00","10:31:00","11:01:00","11:31:00","12:01:00"] 
index3_5s_B = ["075900_to_083100","082900_to_090100","085900_to_093100","092900_to_100100","095900_to_103100","102900_to_110100","105900_to_113100","112900_to_120100"]
index4_5s_B = ["075900","082900","085900","092900","095900","102900","105900","112900"]
index5_5s_B = ["083100","090100","093100","100100","103100","110100","113100","120100"]
tDelta_5s_B = -1 #day

# 1min 2h32min
'''
index1_1m_A = ["09:19:00","11:20:00","13:20:00","15:20:00","17:28:00","04:00:00","05:55:00","07:55:00","08:55:00"]
index2_1m_A = ["11:51:00","13:52:00","15:52:00","17:52:00","20:00:00","06:33:00","08:28:00","10:28:00","11:28:00"]
index3_1m_A = ["091900_to_115100","112000_to_135200","132000_to_155200","152000_to_175200","172800_to_200000","040000_to_063300","055500_to_082800","075500_to_102800","085500_to_112800"]#,"_to_"
index4_1m_A = ["091900","112000","132000","152000","172800","040000","055500","075500","085500"]
index5_1m_A = ["115100","135200","155200","175200","200000","063300","082800","102800","112800"]
tDelta_1m_A = -1 #day
'''
index1_1m_A = ["09:19:00","04:00:00","05:55:00","07:55:00","09:55:00","11:55:00","13:55:00","15:55:00","17:28:00"]
index2_1m_A = ["11:51:00","06:33:00","08:28:00","10:28:00","12:28:00","14:28:00","16:28:00","18:28:00","20:00:00"]
index3_1m_A = ["091900_to_115100","040000_to_063300","055500_to_082800","075500_to_102800","095500_to_122800","115500_to_142800","135500_to_162800","155500_to_182800","172800_to_200000"]#,"_to_"
index4_1m_A = ["091900","040000","055500","075500","095500","115500","135500","155500","172800"]
index5_1m_A = ["115100","063300","082800","102800","122800","142800","162800","182800","200000"]
tDelta_1m_A = -1 #day

# 5min 8h5min 
index1_5m_A = ["04:00:00","08:45:00","12:05:00"]
index2_5m_A = ["11:55:00","16:40:00","20:00:00"]
index3_5m_A = ["040000_to_115500","084500_to_164000","120500_to_200000"]
index4_5m_A = ["040000","084500","120500"]
index5_5m_A = ["115500","164000","200000"]
tDelta_5m_A = -1 #day

# 30min 6days
index1_30m_A = ["04:00:00"]
index2_30m_A = ["20:00:00"]
index3_30m_A = ["040000_to_200000"]
index4_30m_A = ["040000"]
index5_30m_A = ["200000"]
tDelta_30m_A = -6 #days

# 1h 21days
index1_1h_A = ["04:00:00"]
index2_1h_A = ["20:00:00"]
index3_1h_A = ["040000_to_200000"]
index4_1h_A = ["040000"]
index5_1h_A = ["200000"]
tDelta_1h_A = -21 #days

# 1d 220days
index1_1d_A = ["04:00:00"]
index2_1d_A = ["20:00:00"]
index3_1d_A = ["040000_to_200000"]
index4_1d_A = ["040000"]
index5_1d_A = ["200000"]
tDelta_1d_A = -220 #days




def fun(scope       = "1m", # ["5s","1m","5m","30m","1h","1d"]
        indexType   = 1     # [1,2,3]
        ):
    
    if(scope == "5s"):
        if(indexType==1):
            index = index1_5s_A
            index.extend(index1_5s_B)
        elif(indexType==2):
            index = index2_5s_A
            index.extend(index2_5s_B)
        elif(indexType==3):
            index = index3_5s_A
            index.extend(index3_5s_B)
        else:
            print ("wrong index number only accept [1,2,3]")

    elif(scope == "1m"):
        if(indexType==1):
            index = index1_1m_A
        elif(indexType==2):
            index = index2_1m_A
        elif(indexType==3):
            index = index3_1m_A
        else:
            print ("wrong index number only accept [1,2,3]")

    elif(scope == "5m"):
        if(indexType==1):
            index = index1_5m_A
        elif(indexType==2):
            index = index2_5m_A
        elif(indexType==3):
            index = index3_5m_A
        else:
            print ("wrong index number only accept [1,2,3]")

    elif(scope == "30m"):
        if(indexType==1):
            index = index1_30m_A
        elif(indexType==2):
            index = index2_30m_A
        elif(indexType==3):
            index = index3_30m_A
        else:
            print ("wrong index number only accept [1,2,3]")

    elif(scope == "1h"):
        if(indexType==1):
            index = index1_1h_A
        elif(indexType==2):
            index = index2_1h_A
        elif(indexType==3):
            index = index3_1h_A
        else:
            print ("wrong index number only accept [1,2,3]")

    elif(scope == "1d"):
        if(indexType==1):
            index = index1_1d_A
        elif(indexType==2):
            index = index2_1d_A
        elif(indexType==3):
            index = index3_1d_A
        else:
            print ("wrong index number only accept [1,2,3]")

    else:
        print ("wrong Scope type only accept ['5s','1m','5m','30m','1h','1d']")

    return index

#--------------------------------------------------------------------------------------------------------------------------

def fun2(   filePath ,
            bookScope = "Yahoo Dayes" ,
            scope       = "1m" # ["5s","1m","5m","30m","1h","1d"]         
        ): 
    df = ImportData.fun (filePath_fun = filePath , bookName = bookScope)
    Eday = str( df.index[-1] )

    if      (scope=="5s"):
        taro = pd.DataFrame({
                                'index1': index1_5s_A + index1_5s_B,
                                'index2': index2_5s_A + index2_5s_B,
                                'index3': index3_5s_A + index3_5s_B
                                # 'index4': index2_5s_A + index4_5s_B,
                                # 'index5': index2_5s_A + index5_5s_B
                            })
        Sday = str( df.index[tDelta_5s_A] )

    elif    (scope=="1m"):
        taro = pd.DataFrame({
                                'index1': index1_1m_A,
                                'index2': index2_1m_A,
                                'index3': index3_1m_A
                                # 'index4': index4_1m_A,
                                # 'index5': index5_1m_A
                            })
        Sday = str( df.index[tDelta_1m_A] )

    elif    (scope=="5m"):
        taro = pd.DataFrame({
                                'index1': index1_5m_A,
                                'index2': index2_5m_A,
                                'index3': index3_5m_A
                                # 'index4': index4_5m_A,
                                # 'index5': index5_5m_A
                            })
        Sday = str( df.index[tDelta_5m_A] )

    elif    (scope=="30m"):
        taro = pd.DataFrame({
                                'index1': index1_30m_A,
                                'index2': index2_30m_A,
                                'index3': index3_30m_A
                                # 'index4': index4_30m_A,
                                # 'index5': index5_30m_A
                            })
        Sday = str( df.index[tDelta_30m_A] )

    elif    (scope=="1h"):
        taro = pd.DataFrame({
                                'index1': index1_1h_A,
                                'index2': index2_1h_A,
                                'index3': index3_1h_A
                                # 'index4': index4_1h_A,
                                # 'index5': index5_1h_A
                            })
        Sday = str( df.index[tDelta_1h_A] )

    elif    (scope=="1d"):
        taro = pd.DataFrame({
                                'index1': index1_1d_A,
                                'index2': index2_1d_A,
                                'index3': index3_1d_A
                                # 'index4': index4_1d_A,
                                # 'index5': index5_1d_A
                            })
        Sday = str( df.index[tDelta_1d_A] )

    else:
        print ("wrong Scope type only accept ['5s','1m','5m','30m','1h','1d']")
    

    taro["index4"] = Eday[:10].replace('-', '') +'_'+ (taro.index1.str.replace(':', '')) # link:- https://www.statology.org/pandas-remove-characters-from-string/
    taro["index5"] = Eday[:10].replace('-', '') +'_'+ (taro.index2.str.replace(':', ''))
    taro.index1 = Sday[:11] + taro.index1 
    taro.index2 = Eday[:11] + taro.index2 

    taro.index1 = pd.to_datetime(taro.index1.astype('datetime64[ns]'))
    taro.index2 = pd.to_datetime(taro.index2.astype('datetime64[ns]'))

    return taro

#--------------------------------------------------------------------------------------------------------------------------

def fun3(   filePath ,
            bookType    = "Yahoo", # ["Yahoo","IBKR"]
            bookScope   = "Yahoo Dayes" ,
            scope       = "1m", # ["5s","1m","5m","30m","1h","1d"]     
            tDelta      = -1    
        ): 
    df = ImportData.fun (filePath_fun = filePath , bookName = bookScope)
    Eday = str( df.index[-1] )
    Sday = str( df.index[tDelta] )

    if      (scope=="5s"):
        taro = pd.DataFrame({
                                'index1': index1_5s_A + index1_5s_B,
                                'index2': index2_5s_A + index2_5s_B,
                                'index3': index3_5s_A + index3_5s_B                                
                            })
        
    elif    (scope=="1m"):
        taro = pd.DataFrame({
                                'index1': index1_1m_A,
                                'index2': index2_1m_A,
                                'index3': index3_1m_A
                            })
        
    elif    (scope=="5m"):
        taro = pd.DataFrame({
                                'index1': index1_5m_A,
                                'index2': index2_5m_A,
                                'index3': index3_5m_A
                            })
        
    elif    (scope=="30m"):
        taro = pd.DataFrame({
                                'index1': index1_30m_A,
                                'index2': index2_30m_A,
                                'index3': index3_30m_A
                            })
        
    elif    (scope=="1h"):
        taro = pd.DataFrame({
                                'index1': index1_1h_A,
                                'index2': index2_1h_A,
                                'index3': index3_1h_A
                            })
        
    elif    (scope=="1d"):
        taro = pd.DataFrame({
                                'index1': index1_1d_A,
                                'index2': index2_1d_A,
                                'index3': index3_1d_A
                            })
        

    else:
        print ("wrong Scope type only accept ['5s','1m','5m','30m','1h','1d']")
    

    if      (bookType=="Yahoo" ):# ["Yahoo","IBKR"]
                    taro["index4"] = Eday[:10].replace('-', '') +'_'+ (taro.index1.str.replace(':', '')) # link:- https://www.statology.org/pandas-remove-characters-from-string/
                    taro["index5"] = Eday[:10].replace('-', '') +'_'+ (taro.index2.str.replace(':', ''))
                    taro.index1 = Sday[:11] + taro.index1 
                    taro.index2 = Eday[:11] + taro.index2 

                    taro.index1 = pd.to_datetime(taro.index1.astype('datetime64[ns]'))
                    taro.index2 = pd.to_datetime(taro.index2.astype('datetime64[ns]'))

    elif    (bookType=="IBKR"):
                    taro["index4"] = Eday[:8].replace('-', '') +'_'+ (taro.index1.str.replace(':', '')) # link:- https://www.statology.org/pandas-remove-characters-from-string/
                    taro["index5"] = Eday[:8].replace('-', '') +'_'+ (taro.index2.str.replace(':', ''))
                    taro.index1 = Sday[:8] + (taro.index1.str.replace(':', ''))  
                    taro.index2 = Eday[:8] + (taro.index2.str.replace(':', '')) 

                    taro.index1 = pd.to_datetime(taro.index1.astype('datetime64[ns]'))
                    taro.index2 = pd.to_datetime(taro.index2.astype('datetime64[ns]'))

    return taro

#--------------------------------------------------------------------------------------------------------------------------

def fun4(   filePath ,
            bookScope = "IBKR 1Day" ,
            scope       = "1m" # ["5s","1m","5m","30m","1h","1d"]         
        ): 
    df = ImportData.fun2 (filePath_fun = filePath , bookName = bookScope)
    # print ("##########################################################")
    # print (df)

    # df.Datetime = pd.to_datetime(df.Datetime.astype('datetime64[ns]'))
    # print ("##########################################################")
    # print (df)

    # df = df.set_index('Datetime')
    # print ("##########################################################")
    # print (df)

    Eday = str( df.index[-1] )
    print (Eday)

    if      (scope in ["5s",'5 secs']):
        taro = pd.DataFrame({
                                'index1': index1_5s_A + index1_5s_B,
                                'index2': index2_5s_A + index2_5s_B,
                                'index3': index3_5s_A + index3_5s_B
                                # 'index4': index2_5s_A + index4_5s_B,
                                # 'index5': index2_5s_A + index5_5s_B
                            })
        Sday = str( df.index[tDelta_5s_A] )

    elif    (scope in ["1m",'1 min']):      
        taro = pd.DataFrame({
                                'index1': index1_1m_A,
                                'index2': index2_1m_A,
                                'index3': index3_1m_A
                                # 'index4': index4_1m_A,
                                # 'index5': index5_1m_A
                            })
        Sday = str( df.index[tDelta_1m_A] )

    elif    (scope in ["5m",'5 mins']):
        taro = pd.DataFrame({
                                'index1': index1_5m_A,
                                'index2': index2_5m_A,
                                'index3': index3_5m_A
                                # 'index4': index4_5m_A,
                                # 'index5': index5_5m_A
                            })
        Sday = str( df.index[tDelta_5m_A] )

    elif    (scope in ["30m",'30 mins']):
        taro = pd.DataFrame({
                                'index1': index1_30m_A,
                                'index2': index2_30m_A,
                                'index3': index3_30m_A
                                # 'index4': index4_30m_A,
                                # 'index5': index5_30m_A
                            })
        Sday = str( df.index[tDelta_30m_A] )

    elif    (scope in ["1h",'1 hour']):     
        taro = pd.DataFrame({
                                'index1': index1_1h_A,
                                'index2': index2_1h_A,
                                'index3': index3_1h_A
                                # 'index4': index4_1h_A,
                                # 'index5': index5_1h_A
                            })
        Sday = str( df.index[tDelta_1h_A] )

    elif    (scope in ["1d",'1 day']):
        taro = pd.DataFrame({
                                'index1': index1_1d_A,
                                'index2': index2_1d_A,
                                'index3': index3_1d_A
                                # 'index4': index4_1d_A,
                                # 'index5': index5_1d_A
                            })
        Sday = str( df.index[tDelta_1d_A] )

    else:
        print ("wrong Scope type only accept ['5s','1m','5m','30m','1h','1d']")
    

    taro["index4"] = Eday[:10].replace('-', '') +'_'+ (taro.index1.str.replace(':', '')) # link:- https://www.statology.org/pandas-remove-characters-from-string/
    taro["index5"] = Eday[:10].replace('-', '') +'_'+ (taro.index2.str.replace(':', ''))
    taro.index1 = Sday[:11] + taro.index1 
    taro.index2 = Eday[:11] + taro.index2 

    taro.index1 = pd.to_datetime(taro.index1.astype('datetime64[ns]'))
    taro.index2 = pd.to_datetime(taro.index2.astype('datetime64[ns]'))

    return taro

#--------------------------------------------------------------------------------------------------------------------------


def fun5(    filePath           = r"D:\Python Tools\ChartMaker\SourceDocuments\OutPut_Excel\Session [20230921_221745]\NVDA\OK NVDA.xlsx"
            ,timeStampPath      = r"D:\Python Tools\ChartMaker\SourceDocuments\InPut_Excel\Time_Stamp.xlsx"
            ,timeStampSheet     = "Ver1"
            ,DaySheetName       = "IBKR 1Day" 
            ,scope              =  "1m" # ["5s","1m","5m","30m","1h","1d"]  
            ,scopeBook          = {
                                     "5s"       :"5s"
                                    ,'5 secs'   :"5s"
                                    ,"1m"       :"1m"
                                    ,'1 min'    :"1m"
                                    ,"5m"       :"5m"
                                    ,'5 mins'   :"5m"
                                    ,"30m"      :"30m"
                                    ,'30 mins'  :"30m"
                                    ,"1h"       :"1h"
                                    ,'1 hour'   :"1h"
                                    ,"1d"       :"1d" 
                                    ,'1 day'    :"1d" 
                                } 
      
        ):
            #----------------------------------------------------------------------------------------------------------------------------------------------     
            # Date sheet:- Import data and change index to string     Help Link :-    https://stackoverflow.com/questions/44741587/pandas-timestamp-series-to-string
            #   1-Import data
            dfD = ImportData.fun2 (filePath_fun = filePath , bookName = DaySheetName) # "IBKR 1Day"
            
            if(len(dfD.index)==0):
                if(DaySheetName == "IBKR 1Day"):
                    DaySheetName = "Yahoo Dayes"
                    dfD = ImportData.fun (filePath_fun = filePath , bookName = DaySheetName )   # "Yahoo Dayes"
                elif(DaySheetName == "Yahoo Dayes"):
                    DaySheetName = "IBKR 1Day"
                    dfD = ImportData.fun2 (filePath_fun = filePath , bookName = DaySheetName )   # "Yahoo Dayes"
                else:
                    print("Wrong Day Sheet Name")
                # dfD = ImportData.fun2 (filePath_fun = filePath , bookName = DaySheetName )   # "Yahoo Dayes"
            #   2-Change index Type to string
            dfD.index = dfD.index.astype(str)
            #----------------------------------------------------------------------------------------------------------------------------------------------     



            #----------------------------------------------------------------------------------------------------------------------------------------------     
            # Time Stamp sheet:- Import  data and Set column header 
            #   1-Import  data
            book = openpyxl.load_workbook(timeStampPath)
            sheet = book[timeStampSheet]
            dfS= pd.DataFrame(sheet.values)
            #   2-Set column header :- Convert row to column header for Pandas DataFrame : Link https://stackoverflow.com/questions/26147180/convert-row-to-column-header-for-pandas-dataframe
            dfS.columns = dfS.iloc[0]
            dfS = dfS.drop(0)
            #----------------------------------------------------------------------------------------------------------------------------------------------     


            
            #----------------------------------------------------------------------------------------------------------------------------------------------     
            # OutPut Dataframe "taro":-     #   01- Create dataframe :-
            #-------------------------------#   02- Rename column in dataframe :- 
            #-------------------------------#   03- Clean None Rows :-
            #-------------------------------#   04- Change object to float :-
            #-------------------------------#   05- Add date to index colum :-
            #-------------------------------#   06- Change type of index colum to Time stamp :-
            #-------------------------------#   07- Return OutPut Dataframe "taro" :-
            #   01- Create dataframe :- from "dfS" another dataframe    Help Link :- https://www.statology.org/pandas-create-dataframe-from-existing-dataframe/
            taro = dfS[[    
                                         ("Index1_"         + scopeBook[scope])
                                        ,("Index2_"         + scopeBook[scope])
                                        ,("Name_"           + scopeBook[scope])
                                        ,("StartDelta_"     + scopeBook[scope])
                                        ,("EndDelta_"       + scopeBook[scope])
                                        ,("Fibonacci_"      + scopeBook[scope])
                                        ,("Scope_"          + scopeBook[scope])
                                        ,("HourConstant_"   + scopeBook[scope])
                                        ,("MinConstant_"    + scopeBook[scope])
                                        ,("SecConstant_"    + scopeBook[scope])
                                        ,("leftSide_"       + scopeBook[scope])
                                        ,("Blank_"          + scopeBook[scope])
                                        ]].copy()                        
            #   02- Rename column in dataframe :-      Help Link :- https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.rename.html
            taro = taro.rename(columns={                                                                  
                                         "Index1_"         + scopeBook[scope]    :"Index1"
                                        ,"Index2_"         + scopeBook[scope]    :"Index2"
                                        ,"Name_"           + scopeBook[scope]    :"Name"
                                        ,"StartDelta_"     + scopeBook[scope]    :"StartDelta"
                                        ,"EndDelta_"       + scopeBook[scope]    :"EndDelta"
                                        ,"Fibonacci_"      + scopeBook[scope]    :"Fibonacci"
                                        ,"Scope_"          + scopeBook[scope]    :"Scope"
                                        ,"HourConstant_"   + scopeBook[scope]    :"HourConstant"
                                        ,"MinConstant_"    + scopeBook[scope]    :"MinConstant"
                                        ,"SecConstant_"    + scopeBook[scope]    :"SecConstant"
                                        ,"leftSide_"       + scopeBook[scope]    :"leftSide"
                                        ,"Blank_"          + scopeBook[scope]    :"Date"
                                        })
            #   03- Clean None Rows :- Help Link :- https://www.digitalocean.com/community/tutorials/pandas-dropna-drop-null-na-values-from-dataframe    ,   https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.dropna.html
            taro = taro.dropna(how='all')
            #   04- Change object to float :-   Help Link https://stackoverflow.com/questions/36814100/pandas-to-numeric-for-multiple-columns
            cols = taro.columns.drop(["Index1","Index2","Name","Fibonacci","Scope","Date"])   # set colums you dont want to change         
            taro[cols] = taro[cols].apply(pd.to_numeric, errors='coerce')
            #   05- Add date to index colum :-
            taro["Date"] = dfD.index[taro.EndDelta]
            taro.Index2 = taro.Date + " " + taro.Index2
            taro["Date"] = dfD.index[taro.StartDelta]
            taro.Index1 = taro.Date + " " + taro.Index1
            #   06- Change type of index colum to Time stamp :-
            
            # print(">>>>$$$$$$$$$$$$$$<<<<")
            # print(dfD)
            # print(taro)
            # print(">>>>$$$$$$$$$$$$$$<<<<")
            taro.Index1 = pd.to_datetime(taro.Index1.astype('datetime64[ns]'))
            taro.Index2 = pd.to_datetime(taro.Index2.astype('datetime64[ns]'))

            taro.reset_index(inplace=True)
            #   07- Return OutPut Dataframe "taro" :-
            return taro
            #----------------------------------------------------------------------------------------------------------------------------------------------     

#--------------------------------------------------------------------------------------------------------------------------

def fun6(    filePath           = r"C:\Users\lenovo\Desktop\TGL\BCG\OK BCG.xlsx"
            ,timeStampPath      = r"D:\Python Tools\ChartMaker\SourceDocuments\InPut_Excel\Time_Stamp.xlsx"
            ,timeStampSheet     = "Ver1"
            ,DaySheetName       = "IBKR 1Day" 
            ,scope              =  "5m" # ["5s","1m","5m","30m","1h","1d"]  
            ,scopeBook          = {
                                     "5s"       :"5s"
                                    ,'5 secs'   :"5s"
                                    ,"1m"       :"1m"
                                    ,'1 min'    :"1m"
                                    ,"5m"       :"5m"
                                    ,'5 mins'   :"5m"
                                    ,"30m"      :"30m"
                                    ,'30 mins'  :"30m"
                                    ,"1h"       :"1h"
                                    ,'1 hour'   :"1h"
                                    ,"1d"       :"1d" 
                                    ,'1 day'    :"1d" 
                                } 
      
        ):
            #----------------------------------------------------------------------------------------------------------------------------------------------     
            # Date sheet:- Import data and change index to string     Help Link :-    https://stackoverflow.com/questions/44741587/pandas-timestamp-series-to-string
            #   1-Import data
            dfD = ImportData.fun2 (filePath_fun = filePath , bookName = DaySheetName) # "IBKR 1Day"
            
            if(len(dfD.index)==0):
                if(DaySheetName == "IBKR 1Day"):
                    DaySheetName = "Yahoo Dayes"
                    dfD = ImportData.fun (filePath_fun = filePath , bookName = DaySheetName )   # "Yahoo Dayes"
                elif(DaySheetName == "Yahoo Dayes"):
                    DaySheetName = "IBKR 1Day"
                    dfD = ImportData.fun2 (filePath_fun = filePath , bookName = DaySheetName )   # "Yahoo Dayes"
                else:
                    print("Wrong Day Sheet Name")
                # dfD = ImportData.fun2 (filePath_fun = filePath , bookName = DaySheetName )   # "Yahoo Dayes"
            #   2-Change index Type to string
            dfD.index = dfD.index.astype(str)
            #----------------------------------------------------------------------------------------------------------------------------------------------     



            #----------------------------------------------------------------------------------------------------------------------------------------------     
            # Time Stamp sheet:- Import  data and Set column header 
            #   1-Import  data
            book = openpyxl.load_workbook(timeStampPath)
            sheet = book[timeStampSheet]
            dfS= pd.DataFrame(sheet.values)
            #   2-Set column header :- Convert row to column header for Pandas DataFrame : Link https://stackoverflow.com/questions/26147180/convert-row-to-column-header-for-pandas-dataframe
            dfS.columns = dfS.iloc[0]
            dfS = dfS.drop(0)
            #----------------------------------------------------------------------------------------------------------------------------------------------     


            
            #----------------------------------------------------------------------------------------------------------------------------------------------     
            # OutPut Dataframe "taro":-     #   01- Create dataframe :-
            #-------------------------------#   02- Rename column in dataframe :- 
            #-------------------------------#   03- Clean None Rows :-
            #-------------------------------#   04- Change object to float :-
            #-------------------------------#   05- Add date to index colum :-
            #-------------------------------#   06- Change type of index colum to Time stamp :-
            #-------------------------------#   07- Return OutPut Dataframe "taro" :-
            #   01- Create dataframe :- from "dfS" another dataframe    Help Link :- https://www.statology.org/pandas-create-dataframe-from-existing-dataframe/
            taro = dfS[[    
                                         ("Index1_"         + scopeBook[scope])
                                        ,("Index2_"         + scopeBook[scope])
                                        ,("Name_"           + scopeBook[scope])
                                        ,("StartDelta_"     + scopeBook[scope])
                                        ,("EndDelta_"       + scopeBook[scope])
                                        ,("Fibonacci_"      + scopeBook[scope])
                                        ,("Scope_"          + scopeBook[scope])
                                        ,("HourConstant_"   + scopeBook[scope])
                                        ,("MinConstant_"    + scopeBook[scope])
                                        ,("SecConstant_"    + scopeBook[scope])
                                        ,("leftSide_"       + scopeBook[scope])
                                        ,("Blank_"          + scopeBook[scope])
                                        ]].copy()                        
            #   02- Rename column in dataframe :-      Help Link :- https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.rename.html
            taro = taro.rename(columns={                                                                  
                                         "Index1_"         + scopeBook[scope]    :"Index1"
                                        ,"Index2_"         + scopeBook[scope]    :"Index2"
                                        ,"Name_"           + scopeBook[scope]    :"Name"
                                        ,"StartDelta_"     + scopeBook[scope]    :"StartDelta"
                                        ,"EndDelta_"       + scopeBook[scope]    :"EndDelta"
                                        ,"Fibonacci_"      + scopeBook[scope]    :"Fibonacci"
                                        ,"Scope_"          + scopeBook[scope]    :"Scope"
                                        ,"HourConstant_"   + scopeBook[scope]    :"HourConstant"
                                        ,"MinConstant_"    + scopeBook[scope]    :"MinConstant"
                                        ,"SecConstant_"    + scopeBook[scope]    :"SecConstant"
                                        ,"leftSide_"       + scopeBook[scope]    :"leftSide"
                                        ,"Blank_"          + scopeBook[scope]    :"Date"
                                        })
            #   03- Clean None Rows :- Help Link :- https://www.digitalocean.com/community/tutorials/pandas-dropna-drop-null-na-values-from-dataframe    ,   https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.dropna.html
            taro = taro.dropna(how='all')
            taro.reset_index(inplace=True, drop=True)
            #   04- Change object to float :-   Help Link https://stackoverflow.com/questions/36814100/pandas-to-numeric-for-multiple-columns
            cols = taro.columns.drop(["Index1","Index2","Name","Fibonacci","Scope","Date"])   # set colums you dont want to change         
            taro[cols] = taro[cols].apply(pd.to_numeric, errors='coerce')            
            #   05- Add date to index colum :-            
            taro["Date"] = dfD.index[taro.EndDelta]
            taro.Index2 = taro.Date + " " + taro.Index2                        
            for x in range (len(taro)): 
                # IF taro.StartDelta out of bounds for dfD.index set the firist Day is the start of dfD.index    
                if(len(dfD.index) >= abs(taro.loc[x].StartDelta)): 
                    # print(">>>>_______________________________<<<<")
                    # print(dfD.index[taro.loc[x].StartDelta])    
                    taro.Date[x] = dfD.index[taro.loc[x].StartDelta]      
                else:                       
                    taro.Date[x] =    dfD.index[0]                                    
            taro.Index1 = taro.Date + " " + taro.Index1
            #   06- Change type of index colum to Time stamp :-            
            print(">>>>$$$$$$$$$$$$$$<<<<")
            print(dfD)
            print(taro)
            print(">>>>$$$$$$$$$$$$$$<<<<")
            taro.Index1 = pd.to_datetime(taro.Index1.astype('datetime64[ns]'))
            taro.Index2 = pd.to_datetime(taro.Index2.astype('datetime64[ns]'))            
            #   07- Return OutPut Dataframe "taro" :-
            return taro
            #----------------------------------------------------------------------------------------------------------------------------------------------     


#--------------------------------------------------------------------------------------------------------------------------


def fun7(    
             dfD
            ,dfS            
            ,scope              =  "5m" # ["5s","1m","5m","30m","1h","1d"]  
            ,scopeBook          = {
                                     "5s"       :"5s"
                                    ,'5 secs'   :"5s"
                                    ,"1m"       :"1m"
                                    ,'1 min'    :"1m"
                                    ,"5m"       :"5m"
                                    ,'5 mins'   :"5m"
                                    ,"30m"      :"30m"
                                    ,'30 mins'  :"30m"
                                    ,"1h"       :"1h"
                                    ,'1 hour'   :"1h"
                                    ,"1d"       :"1d" 
                                    ,'1 day'    :"1d" 
                                } 
      
        ):
                        
            #----------------------------------------------------------------------------------------------------------------------------------------------     
            # OutPut Dataframe "taro":-     #   01- Create dataframe :-
            #-------------------------------#   02- Rename column in dataframe :- 
            #-------------------------------#   03- Clean None Rows :-
            #-------------------------------#   04- Change object to float :-
            #-------------------------------#   05- Add date to index colum :-
            #-------------------------------#   06- Change type of index colum to Time stamp :-
            #-------------------------------#   07- Return OutPut Dataframe "taro" :-
            #   01- Create dataframe :- from "dfS" another dataframe    Help Link :- https://www.statology.org/pandas-create-dataframe-from-existing-dataframe/
            taro = dfS[[    
                                         ("Index1_"         + scopeBook[scope])
                                        ,("Index2_"         + scopeBook[scope])
                                        ,("Name_"           + scopeBook[scope])
                                        ,("StartDelta_"     + scopeBook[scope])
                                        ,("EndDelta_"       + scopeBook[scope])
                                        ,("Fibonacci_"      + scopeBook[scope])
                                        ,("Scope_"          + scopeBook[scope])
                                        ,("HourConstant_"   + scopeBook[scope])
                                        ,("MinConstant_"    + scopeBook[scope])
                                        ,("SecConstant_"    + scopeBook[scope])
                                        ,("leftSide_"       + scopeBook[scope])
                                        ,("Blank_"          + scopeBook[scope])
                                        ]].copy()                        
            #   02- Rename column in dataframe :-      Help Link :- https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.rename.html
            taro = taro.rename(columns={                                                                  
                                         "Index1_"         + scopeBook[scope]    :"Index1"
                                        ,"Index2_"         + scopeBook[scope]    :"Index2"
                                        ,"Name_"           + scopeBook[scope]    :"Name"
                                        ,"StartDelta_"     + scopeBook[scope]    :"StartDelta"
                                        ,"EndDelta_"       + scopeBook[scope]    :"EndDelta"
                                        ,"Fibonacci_"      + scopeBook[scope]    :"Fibonacci"
                                        ,"Scope_"          + scopeBook[scope]    :"Scope"
                                        ,"HourConstant_"   + scopeBook[scope]    :"HourConstant"
                                        ,"MinConstant_"    + scopeBook[scope]    :"MinConstant"
                                        ,"SecConstant_"    + scopeBook[scope]    :"SecConstant"
                                        ,"leftSide_"       + scopeBook[scope]    :"leftSide"
                                        ,"Blank_"          + scopeBook[scope]    :"Date"
                                        })
            #   03- Clean None Rows :- Help Link :- https://www.digitalocean.com/community/tutorials/pandas-dropna-drop-null-na-values-from-dataframe    ,   https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.dropna.html
            taro = taro.dropna(how='all')
            taro.reset_index(inplace=True, drop=True)
            #   04- Change object to float :-   Help Link https://stackoverflow.com/questions/36814100/pandas-to-numeric-for-multiple-columns
            cols = taro.columns.drop(["Index1","Index2","Name","Fibonacci","Scope","Date"])   # set colums you dont want to change         
            taro[cols] = taro[cols].apply(pd.to_numeric, errors='coerce')            
            #   05- Add date to index colum :-            
            taro["Date"] = dfD.index[taro.EndDelta]
            taro.Index2 = taro.Date + " " + taro.Index2                        
            for x in range (len(taro)): 
                # IF taro.StartDelta out of bounds for dfD.index set the firist Day is the start of dfD.index    
                if(len(dfD.index) >= abs(taro.loc[x].StartDelta)): 
                    # print(">>>>_______________________________<<<<")
                    # print(dfD.index[taro.loc[x].StartDelta])    
                    taro.Date[x] = dfD.index[taro.loc[x].StartDelta]      
                else:                       
                    taro.Date[x] =    dfD.index[0]                                    
            taro.Index1 = taro.Date + " " + taro.Index1
            #   06- Change type of index colum to Time stamp :-            
            print(">>>>$$$$$$$$$$$$$$<<<<")
            print(dfD)
            print(taro)
            print(">>>>$$$$$$$$$$$$$$<<<<")
            taro.Index1 = pd.to_datetime(taro.Index1.astype('datetime64[ns]'))
            taro.Index2 = pd.to_datetime(taro.Index2.astype('datetime64[ns]'))            
            #   07- Return OutPut Dataframe "taro" :-
            return taro
            #----------------------------------------------------------------------------------------------------------------------------------------------     


#--------------------------------------------------------------------------------------------------------------------------















#--------------------------------------------------------------------------------------------------------------------------

def fTimeLable  (    f_Scope
                    ,f_HourConstant
                    ,f_MinConstant
                    ,f_SecConstant
                    ,f_leftSide
                    ,source="IBKR"
                    ,interval='1 min'
                    ,inList=[]
                ):
    print("###-TimeLable-###")
    
    # Set Varibles
    #------------------------------------------------

    tickFlag=True

    monthName=  [
                 "\nworng"
                ,"\nJan"
                ,"\nFeb"
                ,"\nMar"
                ,"\nApr"
                ,"\nMay"
                ,"\nJun"
                ,"\nJul"
                ,"\nAug"
                ,"\nSep"
                ,"\nOct"
                ,"\nNov"
                ,"\nDec"
                ,"" 
                ]

    srcIndex =  {
                    'IBKR'  :{'Y1': 0, 'Y2':4, 'M1': 5, 'M2':7, 'D1':8, 'D2':10,  'H1':11,  'H2':13, 'm1':14, 'm2':16, 's1':17, 's2':19 },
                    'Yahoo' :{'Y1': 0, 'Y2':4, 'M1': 5, 'M2':7, 'D1':8, 'D2':10, 'H1':11, 'H2':13, 'm1':13, 'm2':15, 's1':17, 's2':19 }
                }
        
    ntrvlHub =  {
                    '1 secs'	:{	'SheetName':'IBKR 1s'	    ,'IntervalSize':'1 min'	    ,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':1	    ,'secConstant':0.3      ,'leftSide':0	},		#	XX:XX:OO	Every 1  min
                    '5 secs'	:{	'SheetName':'IBKR 5s'   	,'IntervalSize':'1 min'	    ,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':1	    ,'secConstant':0.3	    ,'leftSide':-0.41  },		#	XX:XX:OO	Every 1  min
                    '10 secs'	:{	'SheetName':'IBKR 10s'  	,'IntervalSize':'1 min'	    ,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':1	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:XX:OO	Every 1  min
                    '15 secs'	:{	'SheetName':'IBKR 15s'  	,'IntervalSize':'1 min'	    ,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':1	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:XX:OO	Every 1  min
                    '30 secs'	:{	'SheetName':'IBKR 30s'  	,'IntervalSize':'5 min'	    ,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':5	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:O5:OO	Every 5  min
                    '1 min' 	:{	'SheetName':"IBKR 1m"   	,'IntervalSize':'5 min'	    ,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':5	    ,'secConstant':0.3	    ,'leftSide':-0.35   },		#	XX:O5:OO	Every 5  min
                    '2 mins' 	:{	'SheetName':'IBKR 2m'   	,'IntervalSize':'10 min'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':10	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:10:OO	Every 10 min
                    '3 mins' 	:{	'SheetName':'IBKR 3m'   	,'IntervalSize':'15 min'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':15	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:15:OO	Every 15 min
                    '5 mins' 	:{	'SheetName':'IBKR 5m'   	,'IntervalSize':'15 min'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':15	    ,'secConstant':0.3	    ,'leftSide':-0.36   },		#	XX:15:OO	Every 15 min
                    '10 mins'	:{	'SheetName':'IBKR 10m'  	,'IntervalSize':'30 mins'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':30	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:30:OO	Every 30 min
                    '15 mins'	:{	'SheetName':'IBKR 15m'  	,'IntervalSize':'1 hour'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:OO:OO	Every 1  hour
                    '20 mins'	:{	'SheetName':'IBKR 20m'  	,'IntervalSize':'1 hour'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:OO:OO	Every 1  hour
                    '30 mins'	:{	'SheetName':'IBKR 30m'	    ,'IntervalSize':'2 hours'	,'Scope':'intraday'	    ,'hourConstant':2	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':-0.38   },		#	O2:OO:OO	Every 2  hour
                    '1 hour' 	:{	'SheetName':'IBKR 1H'	    ,'IntervalSize':'2 hours'	,'Scope':'intraday'	    ,'hourConstant':8	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':-0.51   },		#	O2:OO:OO	Every 2  hour
                    '2 hours'	:{	'SheetName':'IBKR 2H'   	,'IntervalSize':'1 day'	    ,'Scope':'daily'	    ,'hourConstant':3	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	O4:OO:OO	Every 1  day
                    '3 hours'	:{	'SheetName':'IBKR 3H'	    ,'IntervalSize':'1 day'	    ,'Scope':'daily'	    ,'hourConstant':13	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	O4:OO:OO	Every 1  day
                    '4 hours'	:{	'SheetName':'IBKR 4H'   	,'IntervalSize':'1 day'	    ,'Scope':'daily'	    ,'hourConstant':11	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	O4:OO:OO	Every 1  day
                    '8 hours'	:{	'SheetName':'IBKR 8H'	    ,'IntervalSize':'1 day'	    ,'Scope':'daily'	    ,'hourConstant':0.3	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	O4:OO:OO	Every 1  day
                    '1 day' 	:{	'SheetName':'IBKR 1Day'	    ,'IntervalSize':'1 week'	,'Scope':'weekly'	    ,'hourConstant':1	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':-0.41   },		#	O4:OO:OO	Every 1  week
                    '1W'    	:{	'SheetName':'IBKR 1week'	,'IntervalSize':'1 month'	,'Scope':'monthly'	    ,'hourConstant':1	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	O4:OO:OO	Every 1  month
                    '1M'     	:{	'SheetName':'IBKR 1Month'   ,'IntervalSize':'1 year'	,'Scope':'yearly'	    ,'hourConstant':1	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	O4:OO:OO	Every 1  year
                                                                                                            
                    '1m'     	:{	'SheetName':'Yahoo 1m'	    ,'IntervalSize':'5 min'	    ,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':5	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:O5:OO	Every 5  min
                    '2m'     	:{	'SheetName':'Yahoo 2m'	    ,'IntervalSize':'10 min'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':10	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:10:OO	Every 10 min
                    '5m'     	:{	'SheetName':'Yahoo 5m'	    ,'IntervalSize':'15 min'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':15	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:15:OO	Every 15 min
                    '15m'   	:{	'SheetName':'Yahoo 15m'	    ,'IntervalSize':'1 hour'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':-0.21   },		#	XX:OO:OO	Every 1  hour
                    '30m'   	:{	'SheetName':'Yahoo 30m'	    ,'IntervalSize':'2 hours'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	O2:OO:OO	Every 2  hour
                    '1h'     	:{	'SheetName':'Yahoo Hours'	,'IntervalSize':'2 hours'	,'Scope':'daily'	    ,'hourConstant':1	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	O2:OO:OO	Every 2  hour
                    '1d'     	:{	'SheetName':'Yahoo Dayes'	,'IntervalSize':'1 week'	,'Scope':'monthly'	    ,'hourConstant':1	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   }		#	O4:OO:OO	Every 1  week
                }

    outLable =  {   
                    'ticks':[],
                    'tlabs':[],
                    'mitks':[],
                    'milab':[]
                }


    if(source in srcIndex) :
        print("source is ",source)
        if(interval in ntrvlHub) :
            print("interval is ",interval)
            print("Scope is ",f_Scope)
            if(len(inList) > 0):
                print("Lenth of InList is :",len(inList))
                print("First Time Lable is :",inList[0])

                for i in range (len(inList)) : 
                    # Set Varibles
                    #------------------------------------------------
                    if(interval in ['1 secs','5 secs','10 secs','15 secs','30 secs','1 min','2 mins','3 mins','5 mins','10 mins','15 mins','20 mins','30 mins','1 hour',
                                    '2 hours','3 hours','4 hours','8 hours','1m','2m','5m','15m','30m','1h']
                        ):
                            currentSecond   = (int(inList[i][(srcIndex[source]['s1']):(srcIndex[source]["s2"])]))   
                            currentMinute   = (int(inList[i][(srcIndex[source]["m1"]):(srcIndex[source]["m2"])]))
                            currentHour     = (int(inList[i][(srcIndex[source]["H1"]):(srcIndex[source]["H2"])]))
                            currentLable    =  str("%02d"%currentHour) + ':' + str("%02d"%currentMinute)      # lable HH:MM
                    
                    
                    currentDay      = (int(inList[i][(srcIndex[source]["D1"]):(srcIndex[source]["D2"])])) 
                    currentMonth    = (int(inList[i][(srcIndex[source]["M1"]):(srcIndex[source]["M2"])])) 
                    currentYear     = (int(inList[i][(srcIndex[source]["Y1"]):(srcIndex[source]["Y2"])]))  
                    #------------------------------------------------
                    if(i==0):
                            lastDay      = currentDay 
                            lastMonth    = currentMonth
                            lastYear     = currentYear
                    elif(i>0):
                            lastDay      = (int(inList[i-1][(srcIndex[source]["D1"]):(srcIndex[source]["D2"])])) 
                            lastMonth    = (int(inList[i-1][(srcIndex[source]["M1"]):(srcIndex[source]["M2"])])) 
                            lastYear     = (int(inList[i-1][(srcIndex[source]["Y1"]):(srcIndex[source]["Y2"])])) 
                    #------------------------------------------------
                    
                    dayLable        =  monthName[currentMonth] + '.' + str("%02d"%currentDay)        # lable MM.DD
                    yearLable       = str(currentYear) + '.' + dayLable                    # lable YYYY.MM.DD
                    #------------------------------------------------


                    if(f_Scope == 'intraday'):
                        #__________________________________________
                        #   Day lable:
                        if(     currentYear  >  lastYear        #   Year   Change:
                            or   currentMonth >  lastMonth       #   Month  Change:
                            or   currentDay   >  lastDay         #   Day    Change:
                        ):
                                    if(     currentYear  >  lastYear    #   Year   Change:
                                        ):
                                            condLable = yearLable
                                    else:
                                            condLable = dayLable
                                    outLable["ticks"].append(i + f_leftSide)
                                    outLable["tlabs"].append("\n \n " +currentLable)
                                    outLable["mitks"].append(i-0.1)   
                                    outLable["milab"].append("\n \n " +condLable)                                                 
                        #__________________________________________
                        #   clock lable:
                        elif(       (( currentSecond % f_SecConstant  ) == 0)     # Seconds   Condition
                                &   (( currentMinute % f_MinConstant  ) == 0)     # Minutes   Condition
                                &   (( currentHour   % f_HourConstant ) == 0)     # Hours     Condition
                            ):
                                    if(     tickFlag  == True     ):#   First  Candel:
                                            tickFlag  =  False
                                            condLable =  currentLable   #.replace('\n', '')
                                    else:
                                            tickFlag  =  True
                                            condLable =  '\n' + currentLable                                 
                                    outLable["ticks"].append(i + f_leftSide)
                                    outLable["tlabs"].append(condLable)
                        #__________________________________________
                        
                                
                
                    elif(f_Scope == 'daily'):
                        #__________________________________________
                        #   Day lable:
                        if(     i==0                            #   First  Candel:
                            or   currentYear  >  lastYear        #   Year   Change:
                            or   currentMonth >  lastMonth       #   Month  Change:
                            or   currentDay   >  lastDay         #   Day    Change:
                        ):
                                    
                                    if(     i==0                        #   First  Candel:
                                        or   currentYear  >  lastYear    #   Year   Change:
                                        ):
                                            condLable = yearLable
                                    else:
                                            condLable = dayLable
                                    outLable["ticks"].append(i + f_leftSide)
                                    outLable["tlabs"].append("\n \n " + currentLable)
                                    outLable["mitks"].append(i-0.1)   
                                    outLable["milab"].append("\n " + condLable)                                                 
                        #__________________________________________
                        #   clock lable:
                        elif(       (( currentSecond % f_SecConstant  ) == 0)     # Seconds   Condition
                                &   (( currentMinute % f_MinConstant  ) == 0)     # Minutes   Condition
                                &   (( currentHour   % f_HourConstant ) == 0)     # Hours     Condition
                            ):
                                    outLable["ticks"].append(i + f_leftSide)
                                    outLable["tlabs"].append(currentLable)
                        #__________________________________________


                    elif(f_Scope == 'weekly'):
                        #__________________________________________
                        #   Day lable:
                        if(         i==0                            #   First  Candel:
                            or       currentYear  >  lastYear        #   Year   Change:
                            or       currentMonth >  lastMonth       #   Month  Change: 
                        ):                                
                                    if(     i==0                        #   First  Candel:
                                        or   currentYear  >  lastYear    #   Year   Change:
                                        ):
                                            condLable = "\n \n " + yearLable.replace('\n', '')
                                    else:                               #   Month  Change: 
                                            condLable = "\n \n " + dayLable.replace('\n', '')
                                            if(i % 5 == 0):
                                                    # print("_____1_____")
                                                    if( i%2==0):                        #   First  Candel:                                                
                                                        tCondLable = dayLable.replace('\n', '')
                                                        # print("_____2_____")
                                                    else:
                                                        # print("_____3_____")
                                                        tCondLable = dayLable        
                                                    outLable["ticks"].append(i + f_leftSide)
                                                    outLable["tlabs"].append(tCondLable)

                                    outLable["mitks"].append(i+ f_leftSide -0.2)
                                    outLable["milab"].append(condLable)
                                                                                    
                        #__________________________________________
                        #   Week lable:
                        elif(       i % 5 == 0        #   Day    Change: currentDay
                            ):
                                    if(     i%2==0                        #   First  Candel:
                                        # or   currentYear  >  lastYear    #   Year   Change:
                                        ):
                                            condLable = dayLable.replace('\n', '')
                                    else:
                                            condLable = dayLable        
                                    outLable["ticks"].append(i + f_leftSide)
                                    outLable["tlabs"].append(condLable)
                        #__________________________________________


                    elif(f_Scope == 'monthly'):
                        #__________________________________________
                        #   Day lable:
                        if(         i==0                            #   First  Candel:
                            |       currentYear  >  lastYear        #   Year   Change:
                            |       currentMonth >  lastMonth       #   Month  Change: 
                        ):
                                    
                                    if(     i==0                        #   First  Candel:
                                        |   currentYear  >  lastYear    #   Year   Change:
                                        ):
                                            condLable = yearLable
                                    else:
                                            condLable = dayLable
                                    outLable["ticks"].append(i + f_leftSide)
                                    outLable["tlabs"].append(condLable)                                                
                        #__________________________________________


                    elif(f_Scope == 'yearly'):
                        #__________________________________________
                        #   Day lable:
                        if(         i==0                            #   First  Candel:
                            |       currentYear  >  lastYear        #   Year   Change:
                        ):
                                    outLable["ticks"].append(i + f_leftSide)
                                    outLable["tlabs"].append(yearLable)                                                
                        #__________________________________________

                    elif(f_Scope == 'interday'):
                        print("Scope is ",f_Scope)


                    '''
                    if(ntrvlHub[interval].Scope == 'intraday'):
                        print("Scope is ",ntrvlHub[interval].Scope)
                        if  (       (( currentSecond % ntrvlHub[interval].secConstant  ) == 0)     # Seconds   Condition
                                &   (( currentMinute % ntrvlHub[interval].minConstant  ) == 0)     # Minutes   Condition
                                &   (( currentHour   % ntrvlHub[interval].hourConstant ) == 0)     # Hours     Condition
                            ):
                                    currentLable =  str(currentHour) + ':' + str(currentMinute) # lable HH:MM
                                    outLable.ticks.append(i)
                                    outLable.tlabs.append(currentLable)
                                    print(i)
                                    print(currentLable) 

                        if  (       (i+1)in range (len(inList)) ):
                                    j=i+1 
                                    nextDay      = (int(inList[j][(srcIndex[source].D1):(srcIndex[source].D2)])) 
                                    nextMonth    = (int(inList[j][(srcIndex[source].M1):(srcIndex[source].M2)])) 
                                    nextYear     = (int(inList[j][(srcIndex[source].Y1):(srcIndex[source].Y2)]))  
                                    if(                 nextMonth>currentMonth      #   Month   changed
                                            |           nextDay>currentDay          #   day     changed     
                                        ): 
                                                dayLable =  monthName[nextMonth] + '.' + str(nextDay)
                                                if(     nextYear>currentYear):      #   year    changed 
                                                    dayLable = str(nextYear) + '.' + dayLable
                                                outLable.mitks.append(j)
                                                outLable.milab.append(dayLable)
    
                    '''
            else:
                print("No InList was Entered to Function. Lenth of InList is :",len(inList))



        else:
            print("interval is Wrong")
    else:
        print("source is Wrong")

    return outLable

#--------------------------------------------------------------------------------------------------------------------------
    
def fTimeLable2 (    f_Scope
                    ,f_HourConstant
                    ,f_MinConstant
                    ,f_SecConstant
                    ,f_leftSide
                    ,source="IBKR"
                    ,interval='1 min'
                    ,inList=[]
                ):
    print("###-TimeLable-###")
    
    # Set Varibles
    #------------------------------------------------

    tickFlag=True

    monthName=  [
                 "\nworng"
                ,"\nJan"
                ,"\nFeb"
                ,"\nMar"
                ,"\nApr"
                ,"\nMay"
                ,"\nJun"
                ,"\nJul"
                ,"\nAug"
                ,"\nSep"
                ,"\nOct"
                ,"\nNov"
                ,"\nDec"
                ,"" 
                ]

    srcIndex =  {
                    'IBKR'  :{'Y1': 0, 'Y2':4, 'M1': 5, 'M2':7, 'D1':8, 'D2':10,  'H1':11,  'H2':13, 'm1':14, 'm2':16, 's1':17, 's2':19 },
                    'Yahoo' :{'Y1': 0, 'Y2':4, 'M1': 5, 'M2':7, 'D1':8, 'D2':10, 'H1':11, 'H2':13, 'm1':13, 'm2':15, 's1':17, 's2':19 }
                }
        
    ntrvlHub =  {
                    '1 secs'	:{	'SheetName':'IBKR 1s'	    ,'IntervalSize':'1 min'	    ,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':1	    ,'secConstant':0.3      ,'leftSide':0	},		#	XX:XX:OO	Every 1  min
                    '5 secs'	:{	'SheetName':'IBKR 5s'   	,'IntervalSize':'1 min'	    ,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':1	    ,'secConstant':0.3	    ,'leftSide':-0.41  },		#	XX:XX:OO	Every 1  min
                    '10 secs'	:{	'SheetName':'IBKR 10s'  	,'IntervalSize':'1 min'	    ,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':1	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:XX:OO	Every 1  min
                    '15 secs'	:{	'SheetName':'IBKR 15s'  	,'IntervalSize':'1 min'	    ,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':1	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:XX:OO	Every 1  min
                    '30 secs'	:{	'SheetName':'IBKR 30s'  	,'IntervalSize':'5 min'	    ,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':5	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:O5:OO	Every 5  min
                    '1 min' 	:{	'SheetName':"IBKR 1m"   	,'IntervalSize':'5 min'	    ,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':5	    ,'secConstant':0.3	    ,'leftSide':-0.35   },		#	XX:O5:OO	Every 5  min
                    '2 mins' 	:{	'SheetName':'IBKR 2m'   	,'IntervalSize':'10 min'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':10	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:10:OO	Every 10 min
                    '3 mins' 	:{	'SheetName':'IBKR 3m'   	,'IntervalSize':'15 min'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':15	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:15:OO	Every 15 min
                    '5 mins' 	:{	'SheetName':'IBKR 5m'   	,'IntervalSize':'15 min'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':15	    ,'secConstant':0.3	    ,'leftSide':-0.36   },		#	XX:15:OO	Every 15 min
                    '10 mins'	:{	'SheetName':'IBKR 10m'  	,'IntervalSize':'30 mins'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':30	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:30:OO	Every 30 min
                    '15 mins'	:{	'SheetName':'IBKR 15m'  	,'IntervalSize':'1 hour'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:OO:OO	Every 1  hour
                    '20 mins'	:{	'SheetName':'IBKR 20m'  	,'IntervalSize':'1 hour'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:OO:OO	Every 1  hour
                    '30 mins'	:{	'SheetName':'IBKR 30m'	    ,'IntervalSize':'2 hours'	,'Scope':'intraday'	    ,'hourConstant':2	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':-0.38   },		#	O2:OO:OO	Every 2  hour
                    '1 hour' 	:{	'SheetName':'IBKR 1H'	    ,'IntervalSize':'2 hours'	,'Scope':'intraday'	    ,'hourConstant':8	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':-0.51   },		#	O2:OO:OO	Every 2  hour
                    '2 hours'	:{	'SheetName':'IBKR 2H'   	,'IntervalSize':'1 day'	    ,'Scope':'daily'	    ,'hourConstant':3	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	O4:OO:OO	Every 1  day
                    '3 hours'	:{	'SheetName':'IBKR 3H'	    ,'IntervalSize':'1 day'	    ,'Scope':'daily'	    ,'hourConstant':13	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	O4:OO:OO	Every 1  day
                    '4 hours'	:{	'SheetName':'IBKR 4H'   	,'IntervalSize':'1 day'	    ,'Scope':'daily'	    ,'hourConstant':11	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	O4:OO:OO	Every 1  day
                    '8 hours'	:{	'SheetName':'IBKR 8H'	    ,'IntervalSize':'1 day'	    ,'Scope':'daily'	    ,'hourConstant':0.3	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	O4:OO:OO	Every 1  day
                    '1 day' 	:{	'SheetName':'IBKR 1Day'	    ,'IntervalSize':'1 week'	,'Scope':'weekly'	    ,'hourConstant':1	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':-0.41   },		#	O4:OO:OO	Every 1  week
                    '1W'    	:{	'SheetName':'IBKR 1week'	,'IntervalSize':'1 month'	,'Scope':'monthly'	    ,'hourConstant':1	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	O4:OO:OO	Every 1  month
                    '1M'     	:{	'SheetName':'IBKR 1Month'   ,'IntervalSize':'1 year'	,'Scope':'yearly'	    ,'hourConstant':1	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	O4:OO:OO	Every 1  year
                                                                                                            
                    '1m'     	:{	'SheetName':'Yahoo 1m'	    ,'IntervalSize':'5 min'	    ,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':5	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:O5:OO	Every 5  min
                    '2m'     	:{	'SheetName':'Yahoo 2m'	    ,'IntervalSize':'10 min'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':10	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:10:OO	Every 10 min
                    '5m'     	:{	'SheetName':'Yahoo 5m'	    ,'IntervalSize':'15 min'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':15	    ,'secConstant':0.3	    ,'leftSide':0   },		#	XX:15:OO	Every 15 min
                    '15m'   	:{	'SheetName':'Yahoo 15m'	    ,'IntervalSize':'1 hour'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':-0.21   },		#	XX:OO:OO	Every 1  hour
                    '30m'   	:{	'SheetName':'Yahoo 30m'	    ,'IntervalSize':'2 hours'	,'Scope':'intraday'	    ,'hourConstant':1	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	O2:OO:OO	Every 2  hour
                    '1h'     	:{	'SheetName':'Yahoo Hours'	,'IntervalSize':'2 hours'	,'Scope':'daily'	    ,'hourConstant':1	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   },		#	O2:OO:OO	Every 2  hour
                    '1d'     	:{	'SheetName':'Yahoo Dayes'	,'IntervalSize':'1 week'	,'Scope':'monthly'	    ,'hourConstant':1	    ,'minConstant':0.3	    ,'secConstant':0.3	    ,'leftSide':0   }		#	O4:OO:OO	Every 1  week
                }

    outLable =  {   
                     'ticks' :[]
                    ,'tlabs' :[]
                    ,'Utlabs':[]
                    ,'mitks' :[]
                    ,'milab' :[]
                    ,'Umilab':[]
                }


    if(source in srcIndex) :
        print("source is ",source)
        if(interval in ntrvlHub) :
            print("interval is ",interval)
            print("Scope is ",f_Scope)
            if(len(inList) > 0):
                print("Lenth of InList is :",len(inList))
                print("First Time Lable is :",inList[0])

                for i in range (len(inList)) : 
                    # Set Varibles
                    #------------------------------------------------
                    if(interval in ['1 secs','5 secs','10 secs','15 secs','30 secs','1 min','2 mins','3 mins','5 mins','10 mins','15 mins','20 mins','30 mins','1 hour',
                                    '2 hours','3 hours','4 hours','8 hours','1m','2m','5m','15m','30m','1h']
                        ):
                            currentSecond   = (int(inList[i][(srcIndex[source]['s1']):(srcIndex[source]["s2"])]))   
                            currentMinute   = (int(inList[i][(srcIndex[source]["m1"]):(srcIndex[source]["m2"])]))
                            currentHour     = (int(inList[i][(srcIndex[source]["H1"]):(srcIndex[source]["H2"])]))
                            currentLable    =  str("%02d"%currentHour) + ':' + str("%02d"%currentMinute)      # lable HH:MM
                    
                    
                    currentDay      = (int(inList[i][(srcIndex[source]["D1"]):(srcIndex[source]["D2"])])) 
                    currentMonth    = (int(inList[i][(srcIndex[source]["M1"]):(srcIndex[source]["M2"])])) 
                    currentYear     = (int(inList[i][(srcIndex[source]["Y1"]):(srcIndex[source]["Y2"])]))  
                    #------------------------------------------------
                    if(i==0):
                            lastDay      = currentDay 
                            lastMonth    = currentMonth
                            lastYear     = currentYear
                    elif(i>0):
                            lastDay      = (int(inList[i-1][(srcIndex[source]["D1"]):(srcIndex[source]["D2"])])) 
                            lastMonth    = (int(inList[i-1][(srcIndex[source]["M1"]):(srcIndex[source]["M2"])])) 
                            lastYear     = (int(inList[i-1][(srcIndex[source]["Y1"]):(srcIndex[source]["Y2"])])) 
                    #------------------------------------------------
                    
                    dayLable        =  monthName[currentMonth] + '.' + str("%02d"%currentDay)        # lable MM.DD
                    yearLable       = str(currentYear) + '.' + dayLable                    # lable YYYY.MM.DD
                    #------------------------------------------------


                    if(f_Scope == 'intraday'):
                        #__________________________________________
                        #   Day lable:
                        if(     currentYear  >  lastYear        #   Year   Change:
                            or   currentMonth >  lastMonth       #   Month  Change:
                            or   currentDay   >  lastDay         #   Day    Change:
                        ):
                                    if(     currentYear  >  lastYear    #   Year   Change:
                                        ):
                                            condLable = yearLable
                                    else:
                                            condLable = dayLable
                                    outLable["ticks"].append(i + f_leftSide)
                                    outLable["tlabs"].append("\n \n " +currentLable)
                                    outLable["Utlabs"].append(currentLable + " \n \n" )
                                    outLable["mitks"].append(i-0.1)   
                                    outLable["milab"].append("\n \n " +condLable)
                                    outLable["Umilab"].append(condLable + " \n \n \n" )                                                 
                        #__________________________________________
                        #   clock lable:
                        elif(       (( currentSecond % f_SecConstant  ) == 0)     # Seconds   Condition
                                &   (( currentMinute % f_MinConstant  ) == 0)     # Minutes   Condition
                                &   (( currentHour   % f_HourConstant ) == 0)     # Hours     Condition
                            ):
                                    if(     tickFlag  == True     ):#   First  Candel:
                                            tickFlag  =  False
                                            condLable =  currentLable   #.replace('\n', '')
                                            condLable1 =  currentLable   #.replace('\n', '')
                                    else:
                                            tickFlag  =  True
                                            condLable =  '\n' + currentLable 
                                            condLable1 = currentLable + '\n'                                 
                                    outLable["ticks"].append(i + f_leftSide)
                                    outLable["tlabs"].append(condLable)
                                    outLable["Utlabs"].append(condLable1)
                        #__________________________________________
                        
                                
                
                    elif(f_Scope == 'daily'):
                        #__________________________________________
                        #   Day lable:
                        if(     i==0                            #   First  Candel:
                            or   currentYear  >  lastYear        #   Year   Change:
                            or   currentMonth >  lastMonth       #   Month  Change:
                            or   currentDay   >  lastDay         #   Day    Change:
                        ):
                                    
                                    if(     i==0                        #   First  Candel:
                                        or   currentYear  >  lastYear    #   Year   Change:
                                        ):
                                            condLable = yearLable
                                    else:
                                            condLable = dayLable
                                    outLable["ticks"].append(i + f_leftSide)
                                    outLable["tlabs"].append("\n \n " + currentLable)
                                    outLable["Utlabs"].append(currentLable + " \n \n" )
                                    outLable["mitks"].append(i-0.1)   
                                    outLable["milab"].append("\n " + condLable)
                                    outLable["Umilab"].append(condLable + " \n")                                                 
                        #__________________________________________
                        #   clock lable:
                        elif(       (( currentSecond % f_SecConstant  ) == 0)     # Seconds   Condition
                                &   (( currentMinute % f_MinConstant  ) == 0)     # Minutes   Condition
                                &   (( currentHour   % f_HourConstant ) == 0)     # Hours     Condition
                            ):
                                    outLable["ticks"].append(i + f_leftSide)
                                    outLable["tlabs"].append(currentLable)
                        #__________________________________________


                    elif(f_Scope == 'weekly'):
                        #__________________________________________
                        #   Day lable:
                        if(         i==0                            #   First  Candel:
                            or       currentYear  >  lastYear        #   Year   Change:
                            or       currentMonth >  lastMonth       #   Month  Change: 
                        ):                                
                                    if(     i==0                        #   First  Candel:
                                        or   currentYear  >  lastYear    #   Year   Change:
                                        ):
                                            condLable = "\n \n " + yearLable.replace('\n', '')
                                            condLable1 = yearLable.replace('\n', '') + "\n \n " 
                                    else:                               #   Month  Change: 
                                            condLable = "\n \n " + dayLable.replace('\n', '')
                                            condLable1 = dayLable.replace('\n', '') + "\n \n "
                                            if(i % 5 == 0):
                                                    # print("_____1_____")
                                                    if( i%2==0):                        #   First  Candel:                                                
                                                        tCondLable = dayLable.replace('\n', '')
                                                        tCondLable1 = dayLable.replace('\n', '')
                                                        # print("_____2_____")
                                                    else:
                                                        # print("_____3_____")
                                                        tCondLable = dayLable 
                                                        tCondLable1 = dayLable.replace('\n', '') + ' \n'       
                                                    outLable["ticks"].append(i + f_leftSide)
                                                    outLable["tlabs"].append(tCondLable)
                                                    outLable["Utlabs"].append(tCondLable1)

                                    outLable["mitks"].append(i+ f_leftSide -0.2)
                                    outLable["milab"].append(condLable)
                                    outLable["Umilab"].append(condLable1)
                                                                                    
                        #__________________________________________
                        #   Week lable:
                        elif(       i % 5 == 0        #   Day    Change: currentDay
                            ):
                                    if(     i%2==0                        #   First  Candel:
                                        # or   currentYear  >  lastYear    #   Year   Change:
                                        ):
                                            condLable = dayLable.replace('\n', '')
                                            CondLable1 = dayLable.replace('\n', '')
                                    else:
                                            condLable = dayLable
                                            CondLable1 = dayLable.replace('\n', '') + ' \n'        
                                    outLable["ticks"].append(i + f_leftSide)
                                    outLable["tlabs"].append(condLable)
                                    outLable["Utlabs"].append(CondLable1)
                        #__________________________________________


                    elif(f_Scope == 'monthly'):
                        #__________________________________________
                        #   Day lable:
                        if(         i==0                            #   First  Candel:
                            |       currentYear  >  lastYear        #   Year   Change:
                            |       currentMonth >  lastMonth       #   Month  Change: 
                        ):
                                    
                                    if(     i==0                        #   First  Candel:
                                        |   currentYear  >  lastYear    #   Year   Change:
                                        ):
                                            condLable = yearLable
                                    else:
                                            condLable = dayLable
                                    outLable["ticks"].append(i + f_leftSide)
                                    outLable["tlabs"].append(condLable)
                                    outLable["Utlabs"].append(condLable)                                                
                        #__________________________________________


                    elif(f_Scope == 'yearly'):
                        #__________________________________________
                        #   Day lable:
                        if(         i==0                            #   First  Candel:
                            |       currentYear  >  lastYear        #   Year   Change:
                        ):
                                    outLable["ticks"].append(i + f_leftSide)
                                    outLable["tlabs"].append(yearLable)
                                    outLable["Utlabs"].append(yearLable)                                                
                        #__________________________________________

                    elif(f_Scope == 'interday'):
                        print("Scope is ",f_Scope)


                    '''
                    if(ntrvlHub[interval].Scope == 'intraday'):
                        print("Scope is ",ntrvlHub[interval].Scope)
                        if  (       (( currentSecond % ntrvlHub[interval].secConstant  ) == 0)     # Seconds   Condition
                                &   (( currentMinute % ntrvlHub[interval].minConstant  ) == 0)     # Minutes   Condition
                                &   (( currentHour   % ntrvlHub[interval].hourConstant ) == 0)     # Hours     Condition
                            ):
                                    currentLable =  str(currentHour) + ':' + str(currentMinute) # lable HH:MM
                                    outLable.ticks.append(i)
                                    outLable.tlabs.append(currentLable)
                                    print(i)
                                    print(currentLable) 

                        if  (       (i+1)in range (len(inList)) ):
                                    j=i+1 
                                    nextDay      = (int(inList[j][(srcIndex[source].D1):(srcIndex[source].D2)])) 
                                    nextMonth    = (int(inList[j][(srcIndex[source].M1):(srcIndex[source].M2)])) 
                                    nextYear     = (int(inList[j][(srcIndex[source].Y1):(srcIndex[source].Y2)]))  
                                    if(                 nextMonth>currentMonth      #   Month   changed
                                            |           nextDay>currentDay          #   day     changed     
                                        ): 
                                                dayLable =  monthName[nextMonth] + '.' + str(nextDay)
                                                if(     nextYear>currentYear):      #   year    changed 
                                                    dayLable = str(nextYear) + '.' + dayLable
                                                outLable.mitks.append(j)
                                                outLable.milab.append(dayLable)
    
                    '''
            else:
                print("No InList was Entered to Function. Lenth of InList is :",len(inList))



        else:
            print("interval is Wrong")
    else:
        print("source is Wrong")

    return outLable

#--------------------------------------------------------------------------------------------------------------------------
































#--------------------------------------------------------------------------------------------------------------------------

def LabelPrice  (    pMax = 6
                    ,pMin = 2
                ):
    print("Calculate Price Label")
    outLable =  {   
                    'ticks':[],
                    'tlabs':[],
                    'mitks':[],
                    'milab':[]
                }
    pDlt = pMax - pMin
    pInc = 0.05
    pDec = "2"    # "Decimal point"
    pStr = ((pMin*10**1)//1)/(10**1)       # math.floor(pMin) 
    pEnd = ((pMax*10**1)//1)/(10**1)       #  math.ceil(pMax)
            
    if (pMax < 1):
        pInc = 00.0001      #   max < 1
        pDec = "4"
        pStr = round(pMin, 2)
        pEnd = round(pMax, 2)
        round(pMax, 1)
    elif((pDlt < 0.40) and (pMax > 1)):
        pInc = 00.01        #   Dlt = 0.40
    elif(pDlt <= 2 and pDlt > 0.40):
        pInc = 00.05        #   Dlt = 0.40
    elif(pDlt <= 4 and pDlt > 2):
        pInc = 00.10        #   Dlt = 0.40
    elif(pDlt <= 10 and pDlt > 4):
        pInc = 00.25        #   Dlt = 0.40
    elif(pDlt <= 20 and pDlt > 10):
        pInc = 00.50        #   Dlt = 0.40
    elif(pDlt <= 40 and pDlt > 20):
        pInc = 01.00        #   Dlt = 0.40
    elif(pDlt <= 100 and pDlt > 40):
        pInc = 02.50        #   Dlt = 0.40
    elif(pDlt <= 200 and pDlt > 100):
        pInc = 05.00        #   Dlt = 0.40
    elif(pDlt <= 400 and pDlt > 200):
        pInc = 10.00        #   Dlt = 0.40

    
    print(pMax)
    print(pMin)
    print(pDlt)
    print(pInc)
    print(pDec)
    # print()
    # print()
    
    
    outLable["ticks"] = np.arange(pStr,pEnd,pInc)

    # outLable["ticks"] = outLable["ticks"][::-1]
    
    pDec = "%." + pDec + "f"             #   "%.2f" % a
    outLable["tlabs"] =  [(pDec % x) for x in outLable["ticks"]]
    
    print(outLable["ticks"])
    print(outLable["tlabs"])



    print(outLable["ticks"])
    print(outLable["tlabs"])

    return outLable

#--------------------------------------------------------------------------------------------------------------------------
   
def LabelPrice5 (    pMax = 6
                    ,pMin = 2
                ):
    print("Calculate Price Label")
    #__________________________________________________________________________________________________________________
    # Set Varibles:-
    pDlt = round(pMax - pMin,4)
    ntrvIndex = 'Inc00_05'
    outLable =  {   
                    'ticks':[],
                    'tlabs':[],
                    'mitks':[],
                    'milab':[]
                }
    ntrvlHub =  {
                     'Inc00_0001'	:{	'p_Incerement':00.0001	    ,'p_Decimals':'4'	    ,'n_Decimals':4	    ,'n_Integers':0	    	}		#	pInc = 00.0001 
                    ,'Inc00_01'	    :{	'p_Incerement':00.01	    ,'p_Decimals':'2'	    ,'n_Decimals':2	    ,'n_Integers':0	    	}		#	pInc = 00.01
                    ,'Inc00_05'	    :{	'p_Incerement':00.05	    ,'p_Decimals':'2'	    ,'n_Decimals':2	    ,'n_Integers':0	    	}		#	pInc = 00.05
                    ,'Inc00_10'	    :{	'p_Incerement':00.10	    ,'p_Decimals':'2'	    ,'n_Decimals':2	    ,'n_Integers':0	    	}		#	pInc = 00.10
                    ,'Inc00_25'	    :{	'p_Incerement':00.25	    ,'p_Decimals':'2'	    ,'n_Decimals':2	    ,'n_Integers':0	    	}		#	pInc = 00.25
                    ,'Inc00_50'	    :{	'p_Incerement':00.50	    ,'p_Decimals':'2'	    ,'n_Decimals':2	    ,'n_Integers':0	    	}		#	pInc = 00.50
                    ,'Inc01_00'	    :{	'p_Incerement':01.00	    ,'p_Decimals':'2'	    ,'n_Decimals':2	    ,'n_Integers':0	    	}		#	pInc = 01.00
                    ,'Inc02_50'	    :{	'p_Incerement':02.50	    ,'p_Decimals':'2'	    ,'n_Decimals':2	    ,'n_Integers':1	    	}		#	pInc = 02.50
                    ,'Inc05_00'	    :{	'p_Incerement':05.00	    ,'p_Decimals':'2'	    ,'n_Decimals':2	    ,'n_Integers':1	    	}		#	pInc = 05.00
                    ,'Inc10_00'	    :{	'p_Incerement':10.00	    ,'p_Decimals':'2'	    ,'n_Decimals':2	    ,'n_Integers':2	    	}		#	pInc = 10.00
                    ,'Inc25_00'	    :{	'p_Incerement':25.00	    ,'p_Decimals':'2'	    ,'n_Decimals':2	    ,'n_Integers':2	    	}		#	pInc = 25.00
                    ,'Inc50_00'	    :{	'p_Incerement':50.00	    ,'p_Decimals':'2'	    ,'n_Decimals':2	    ,'n_Integers':2	    	}		#	pInc = 50.00 
                   

                }                
    if (pDlt > 0):
        #__________________________________________________________________________________________________________________
        # Choose Interval:- 
        if  (pMax < 1                   ) : ntrvIndex = 'Inc00_0001'  # pInc = 00.0001 
        elif(pDlt < 0.40 and pMax > 1   ) : ntrvIndex = 'Inc00_01'	  # pInc = 00.01
        elif(pDlt <= 2   and pDlt > 0.40) : ntrvIndex = 'Inc00_05'    # pInc = 00.05 
        elif(pDlt <= 4   and pDlt > 2   ) : ntrvIndex = 'Inc00_10'    # pInc = 00.10 
        elif(pDlt <= 10  and pDlt > 4   ) : ntrvIndex = 'Inc00_25'    # pInc = 00.25 
        elif(pDlt <= 20  and pDlt > 10  ) : ntrvIndex = 'Inc00_50'    # pInc = 00.50
        elif(pDlt <= 40  and pDlt > 20  ) : ntrvIndex = 'Inc01_00'    # pInc = 01.00         
        elif(pDlt <= 100 and pDlt > 40  ) : ntrvIndex = 'Inc02_50'    # pInc = 02.50 
        elif(pDlt <= 200 and pDlt > 100 ) : ntrvIndex = 'Inc05_00'    # pInc = 05.00 
        elif(pDlt <= 400 and pDlt > 200 ) : ntrvIndex = 'Inc10_00'    # pInc = 10.00
        elif(                pDlt > 400 ) : ntrvIndex = 'Inc25_00'    # pInc = 25.00 
        #__________________________________________________________________________________________________________________    
        # Extract parameters        
        pMax = float(pMax)
        pMin = float(pMin)

        pInc = ntrvlHub[ntrvIndex]['p_Incerement']
        pDec = ntrvlHub[ntrvIndex]['p_Decimals']    # "Decimal point"

        n_decimals = ntrvlHub[ntrvIndex]['n_Decimals']
        n_Integers = ntrvlHub[ntrvIndex]['n_Integers']
        #__________________________________________________________________________________________________________________    
        # Calculate Start and End range of Price Label
        pDN = round((pMin%(10**n_Integers)),n_decimals) 
        pUP = round((pMax%(10**n_Integers)),n_decimals) 
            
        pStr = round( pMin + (pInc - (pDN%pInc)) ,n_decimals)
        pEnd = round( pMax + (pInc - (pUP%pInc)) ,n_decimals)
        #__________________________________________________________________________________________________________________    
        # Calculate Price Label and Prepare Output
        outLable["ticks"] = np.arange(pStr,pEnd,pInc)
        
        pDec = "%." + pDec + "f"             #   "%.2f" % a
        outLable["tlabs"] =  [(pDec % x) for x in outLable["ticks"]]    
        #__________________________________________________________________________________________________________________
        # Print out Data
        print(outLable["ticks"])
        print(outLable["tlabs"])

        print("____________")
        print("Min:",pMin,"|Max:",pMax,"|Dlt:",pDlt,"|Inc:",pInc,"|Dec:",pDec)
        print("pStr=",pStr,"pDN=",pDN,"     |pEnd=",pEnd,"|pUP=",pUP)
        print(ntrvIndex)
        print("++++++++++++")
        # print()
    #__________________________________________________________________________________________________________________
    else:
        print ("Wrong InPut")
        print("Min:",pMin,"|Max:",pMax,"|Dlt:",pDlt)
    #__________________________________________________________________________________________________________________
    return outLable
    #__________________________________________________________________________________________________________________

#--------------------------------------------------------------------------------------------------------------------------


def LabelPrice6 (    pMax = 6
                    ,pMin = 2
                ):
    print("Calculate Price Label")
    #__________________________________________________________________________________________________________________
    # Set Varibles:-
    pDlt = round(pMax - pMin,4)
    ntrvIndex = 'Inc00_05'
    outLable =  {   
                    'ticks':[],
                    'tlabs':[],
                    'mitks':[],
                    'milab':[]
                }
    ntrvlHub =  {
                     "Inc00_0001"	:{	"p_Incerement":	00.0001	    ,"p_Decimals": "4"	  ,"n_Decimals": 4    ,"n_Integers": 0	  ,"MAX": 0000.0036	  ,"MIN":	0000.0000	}		#	pInc = 	00.0001
                    ,"Inc00_0005"	:{	"p_Incerement":	00.0005	    ,"p_Decimals": "4"	  ,"n_Decimals": 4	  ,"n_Integers": 0	  ,"MAX": 0000.0190	  ,"MIN":	0000.0036	}		#	pInc = 	00.0005
                    ,"Inc00_0010"	:{	"p_Incerement":	00.0010	    ,"p_Decimals": "4"	  ,"n_Decimals": 4    ,"n_Integers": 0	  ,"MAX": 0000.0370	  ,"MIN":	0000.0190	}		#	pInc = 	00.0010
                    ,"Inc00_0025"	:{	"p_Incerement":	00.0025	    ,"p_Decimals": "4"	  ,"n_Decimals": 4	  ,"n_Integers": 0	  ,"MAX": 0000.0925	  ,"MIN":	0000.0370	}		#	pInc = 	00.0025
                    ,"Inc00_0050"	:{	"p_Incerement":	00.0050	    ,"p_Decimals": "4"	  ,"n_Decimals": 4	  ,"n_Integers": 0	  ,"MAX": 0000.1800	  ,"MIN":	0000.0925	}		#	pInc = 	00.0050
                    ,"Inc00_0100"	:{	"p_Incerement":	00.0100	    ,"p_Decimals": "4"	  ,"n_Decimals": 4	  ,"n_Integers": 0	  ,"MAX": 0000.3700	  ,"MIN":	0000.1800	}		#	pInc = 	00.0100
                    ,"Inc00_0250"	:{	"p_Incerement":	00.0250	    ,"p_Decimals": "4"	  ,"n_Decimals": 4	  ,"n_Integers": 0	  ,"MAX": 0000.9250	  ,"MIN":	0000.3700	}		#	pInc = 	00.0250
                    ,"Inc00_0500"	:{	"p_Incerement":	00.0500	    ,"p_Decimals": "4"	  ,"n_Decimals": 4	  ,"n_Integers": 0	  ,"MAX": 0001.8500	  ,"MIN":	0000.9250	}		#	pInc = 	00.0500
                    ,"Inc00_1000"	:{	"p_Incerement":	00.1000	    ,"p_Decimals": "4"	  ,"n_Decimals": 4	  ,"n_Integers": 0	  ,"MAX": 0003.7000	  ,"MIN":	0001.8500	}		#	pInc = 	00.1000
    
                    ,"Inc00_01"	    :{	"p_Incerement":	00.0100	    ,"p_Decimals": "2"	  ,"n_Decimals": 2	  ,"n_Integers": 0	  ,"MAX": 0000.4000	  ,"MIN":	0000.0000	}		#	pInc = 	00.0100
                    ,"Inc00_05"	    :{	"p_Incerement":	00.0500	    ,"p_Decimals": "2"	  ,"n_Decimals": 2	  ,"n_Integers": 0	  ,"MAX": 0002.0000	  ,"MIN":	0000.4000	}		#	pInc = 	00.0500
                    ,"Inc00_10"	    :{	"p_Incerement":	00.1000	    ,"p_Decimals": "2"	  ,"n_Decimals": 2	  ,"n_Integers": 0	  ,"MAX": 0004.0000	  ,"MIN":	0002.0000	}		#	pInc = 	00.1000
                    ,"Inc00_25"	    :{	"p_Incerement":	00.2500	    ,"p_Decimals": "2"	  ,"n_Decimals": 2	  ,"n_Integers": 0	  ,"MAX": 0010.0000	  ,"MIN":	0004.0000	}		#	pInc = 	00.2500
                    ,"Inc00_50"	    :{	"p_Incerement":	00.5000	    ,"p_Decimals": "2"	  ,"n_Decimals": 2	  ,"n_Integers": 0	  ,"MAX": 0020.0000	  ,"MIN":	0010.0000	}		#	pInc = 	00.5000
                    ,"Inc01_00"	    :{	"p_Incerement":	01.0000	    ,"p_Decimals": "2"	  ,"n_Decimals": 2	  ,"n_Integers": 0	  ,"MAX": 0040.0000	  ,"MIN":	0020.0000	}		#	pInc = 	01.0000
                    ,"Inc02_50"	    :{	"p_Incerement":	02.5000	    ,"p_Decimals": "2"	  ,"n_Decimals": 2	  ,"n_Integers": 1	  ,"MAX": 0100.0000	  ,"MIN":	0040.0000	}		#	pInc = 	02.5000
                    ,"Inc05_00"	    :{	"p_Incerement":	05.0000	    ,"p_Decimals": "2"	  ,"n_Decimals": 2	  ,"n_Integers": 1	  ,"MAX": 0200.0000	  ,"MIN":	0100.0000	}		#	pInc = 	05.0000
                    ,"Inc10_00"	    :{	"p_Incerement":	10.0000	    ,"p_Decimals": "2"	  ,"n_Decimals": 2	  ,"n_Integers": 2	  ,"MAX": 0400.0000	  ,"MIN":	0200.0000	}		#	pInc = 	10.0000
                    ,"Inc25_00"	    :{	"p_Incerement":	25.0000	    ,"p_Decimals": "2"	  ,"n_Decimals": 2	  ,"n_Integers": 2	  ,"MAX": 0925.0000	  ,"MIN":	0400.0000	}		#	pInc = 	25.0000
                    ,"Inc50_00"	    :{	"p_Incerement":	50.0000	    ,"p_Decimals": "2"	  ,"n_Decimals": 2	  ,"n_Integers": 2	  ,"MAX": 1800.0000	  ,"MIN":	0925.0000	}		#	pInc = 	50.0000                   
                }                
    if (pDlt > 0):
        #__________________________________________________________________________________________________________________
        # Choose Interval:- 
        if  (pMax < 1 ) : 
            if  (pDlt <= ntrvlHub["Inc00_0001"]["MAX"]  and  pDlt > ntrvlHub["Inc00_0001"]["MIN"]  ):ntrvIndex = 'Inc00_0001'   # pInc = 00.0001 
            elif(pDlt <= ntrvlHub["Inc00_0005"]["MAX"]  and  pDlt > ntrvlHub["Inc00_0005"]["MIN"]  ):ntrvIndex = 'Inc00_0005'   # pInc = 00.0005
            elif(pDlt <= ntrvlHub["Inc00_0010"]["MAX"]  and  pDlt > ntrvlHub["Inc00_0010"]["MIN"]  ):ntrvIndex = 'Inc00_0010'   # pInc = 00.0010
            elif(pDlt <= ntrvlHub["Inc00_0025"]["MAX"]  and  pDlt > ntrvlHub["Inc00_0025"]["MIN"]  ):ntrvIndex = 'Inc00_0025'   # pInc = 00.0025
            elif(pDlt <= ntrvlHub["Inc00_0050"]["MAX"]  and  pDlt > ntrvlHub["Inc00_0050"]["MIN"]  ):ntrvIndex = 'Inc00_0050'   # pInc = 00.0050
            elif(pDlt <= ntrvlHub["Inc00_0100"]["MAX"]  and  pDlt > ntrvlHub["Inc00_0100"]["MIN"]  ):ntrvIndex = 'Inc00_0100'   # pInc = 00.0100
            elif(pDlt <= ntrvlHub["Inc00_0250"]["MAX"]  and  pDlt > ntrvlHub["Inc00_0250"]["MIN"]  ):ntrvIndex = 'Inc00_0250'   # pInc = 00.0250
            elif(pDlt <= ntrvlHub["Inc00_0500"]["MAX"]  and  pDlt > ntrvlHub["Inc00_0500"]["MIN"]  ):ntrvIndex = 'Inc00_0500'   # pInc = 00.0500
            elif(pDlt <= ntrvlHub["Inc00_1000"]["MAX"]  and  pDlt > ntrvlHub["Inc00_1000"]["MIN"]  ):ntrvIndex = 'Inc00_1000'   # pInc = 0001000                        
        elif(pMax >= 1 ) :     
            if  (pDlt <= ntrvlHub["Inc00_01"]["MAX"]    and  pDlt > ntrvlHub["Inc00_01"]["MIN"]    ):ntrvIndex = 'Inc00_01'     # pInc = 00.01
            elif(pDlt <= ntrvlHub["Inc00_05"]["MAX"]    and  pDlt > ntrvlHub["Inc00_05"]["MIN"]    ):ntrvIndex = 'Inc00_05'     # pInc = 00.05
            elif(pDlt <= ntrvlHub["Inc00_10"]["MAX"]    and  pDlt > ntrvlHub["Inc00_10"]["MIN"]    ):ntrvIndex = 'Inc00_10'     # pInc = 00.10
            elif(pDlt <= ntrvlHub["Inc00_25"]["MAX"]    and  pDlt > ntrvlHub["Inc00_25"]["MIN"]    ):ntrvIndex = 'Inc00_25'     # pInc = 00.25
            elif(pDlt <= ntrvlHub["Inc00_50"]["MAX"]    and  pDlt > ntrvlHub["Inc00_50"]["MIN"]    ):ntrvIndex = 'Inc00_50'     # pInc = 00.50
            elif(pDlt <= ntrvlHub["Inc01_00"]["MAX"]    and  pDlt > ntrvlHub["Inc01_00"]["MIN"]    ):ntrvIndex = 'Inc01_00'     # pInc = 01.00
            elif(pDlt <= ntrvlHub["Inc02_50"]["MAX"]    and  pDlt > ntrvlHub["Inc02_50"]["MIN"]    ):ntrvIndex = 'Inc02_50'     # pInc = 02.50
            elif(pDlt <= ntrvlHub["Inc05_00"]["MAX"]    and  pDlt > ntrvlHub["Inc05_00"]["MIN"]    ):ntrvIndex = 'Inc05_00'     # pInc = 05.00
            elif(pDlt <= ntrvlHub["Inc10_00"]["MAX"]    and  pDlt > ntrvlHub["Inc10_00"]["MIN"]    ):ntrvIndex = 'Inc10_00'     # pInc = 10.00
            elif(pDlt <= ntrvlHub["Inc25_00"]["MAX"]    and  pDlt > ntrvlHub["Inc25_00"]["MIN"]    ):ntrvIndex = 'Inc25_00'     # pInc = 25.00
            elif(pDlt <= ntrvlHub["Inc50_00"]["MAX"]    and  pDlt > ntrvlHub["Inc50_00"]["MIN"]    ):ntrvIndex = 'Inc50_00'     # pInc = 50.00
        #__________________________________________________________________________________________________________________    
        # Extract parameters        
        pMax = float(pMax)
        pMin = float(pMin)

        pInc = ntrvlHub[ntrvIndex]['p_Incerement']
        pDec = ntrvlHub[ntrvIndex]['p_Decimals']    # "Decimal point"

        n_decimals = ntrvlHub[ntrvIndex]['n_Decimals']
        n_Integers = ntrvlHub[ntrvIndex]['n_Integers']
        #__________________________________________________________________________________________________________________    
        # Calculate Start and End range of Price Label      
        pDN = round((pMin%(10**n_Integers)),n_decimals)     # Help Link :-  https://stackoverflow.com/questions/63430691/last-two-digits-of-decimal-number-with-a-mask
        pUP = round((pMax%(10**n_Integers)),n_decimals)     # Help Link :-  https://stackoverflow.com/questions/58065055/floor-and-ceil-with-number-of-decimals 
            
        pStr = round( pMin + (pInc - (pDN%pInc)) ,n_decimals)
        pEnd = round( pMax + (pInc - (pUP%pInc)) ,n_decimals)
        #__________________________________________________________________________________________________________________    
        # Calculate Price Label and Prepare Output
        outLable["ticks"] = np.arange(pStr,pEnd,pInc)
        
        pDec = "%." + pDec + "f" #   "%.2f" % a             # Help Link :-  https://stackoverflow.com/questions/455612/limiting-floats-to-two-decimal-points
        outLable["tlabs"] =  [(pDec % x) for x in outLable["ticks"]]    
        #__________________________________________________________________________________________________________________
        # Print out Data
        print(outLable["ticks"])
        print(outLable["tlabs"])

        print("____________")
        print("Min:",pMin,"|Max:",pMax,"|Dlt:",pDlt,"|Inc:",pInc,"|Dec:",pDec)
        print("pStr=",pStr,"pDN=",pDN,"     |pEnd=",pEnd,"|pUP=",pUP)
        print(ntrvIndex)
        print("++++++++++++")
        # print()
    #__________________________________________________________________________________________________________________
    else:
        print ("Wrong InPut")
        print("Min:",pMin,"|Max:",pMax,"|Dlt:",pDlt)
    #__________________________________________________________________________________________________________________
    return outLable
    #__________________________________________________________________________________________________________________




#--------------------------------------------------------------------------------------------------------------------------










































#--------------------------------------------------------------------------------------------------------------------------

def fGetDayDataFrame(    
                        filePath           = r"C:\Users\lenovo\Desktop\TGL\BCG\OK BCG.xlsx"            
                        ,DaySheetName       = "IBKR 1Day"                   
                    ):
            #----------------------------------------------------------------------------------------------------------------------------------------------     
            # Date sheet:- Import data and change index to string     Help Link :-    https://stackoverflow.com/questions/44741587/pandas-timestamp-series-to-string
            #   1-Import data
            dfD = ImportData.fun2 (filePath_fun = filePath , bookName = DaySheetName) # "IBKR 1Day"
            
            if(len(dfD.index)==0):
                if(DaySheetName == "IBKR 1Day"):
                    DaySheetName = "Yahoo Dayes"
                    dfD = ImportData.fun (filePath_fun = filePath , bookName = DaySheetName )   # "Yahoo Dayes"
                elif(DaySheetName == "Yahoo Dayes"):
                    DaySheetName = "IBKR 1Day"
                    dfD = ImportData.fun2 (filePath_fun = filePath , bookName = DaySheetName )   # "Yahoo Dayes"
                else:
                    print("Wrong Day Sheet Name")
                # dfD = ImportData.fun2 (filePath_fun = filePath , bookName = DaySheetName )   # "Yahoo Dayes"
            #   2-Change index Type to string
            dfD.index = dfD.index.astype(str)
            #----------------------------------------------------------------------------------------------------------------------------------------------     
            return dfD


#--------------------------------------------------------------------------------------------------------------------------



#--------------------------------------------------------------------------------------------------------------------------

def fGetTimeStampDataFrame(    
            timeStampPath      = r"D:\Python Tools\ChartMaker\SourceDocuments\InPut_Excel\Watch_List_Zamzam .xlsx"
            ,timeStampSheet     = "Sheet2"                  
        ):

            #----------------------------------------------------------------------------------------------------------------------------------------------     
            # Time Stamp sheet:- Import  data and Set column header 
            #   1-Import  data
            book = openpyxl.load_workbook(timeStampPath)
            sheet = book[timeStampSheet]
            dfS= pd.DataFrame(sheet.values)
            #   2-Set column header :- Convert row to column header for Pandas DataFrame : Link https://stackoverflow.com/questions/26147180/convert-row-to-column-header-for-pandas-dataframe
            dfS.columns = dfS.iloc[0]
            dfS = dfS.drop(0)
            #----------------------------------------------------------------------------------------------------------------------------------------------     
            return dfS


#--------------------------------------------------------------------------------------------------------------------------


#--------------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------

def fGetStyle(                 
             dfS
            ,AttributeIndex=0
            ,AttributeList =[
                 "style_name"
                ,"base_mpl_style"
                ,"marketcolors"
                ,"candle_Up"
                ,"candle_Down"
                ,"edge_Up"
                ,"edge_Down"
                ,"wick_Up"
                ,"wick_Down"
                ,"ohlc_Up"
                ,"ohlc_Down"
                ,"volume_Up"
                ,"volume_Down"
                ,"vcdopcod"
                ,"alpha"
                ,"mavcolors"
                ,"y_on_right"
                ,"gridcolor"
                ,"gridstyle"
                ,"facecolor"
                ,"figcolor"
                ,"rc"
                ,"axes.edgecolor"
                ,"axes.linewidth"
                ,"axes.labelsize"
                ,"axes.labelweight"
                ,"axes.grid"
                ,"axes.grid.axis"
                ,"axes.grid.which"
                ,"grid.alpha"
                ,"grid.color"
                ,"grid.linestyle"
                ,"grid.linewidth"
                ,"figure.facecolor"
                ,"patch.linewidth"
                ,"lines.linewidth"
                ,"font.weight"
                ,"font.size"
                ,"figure.titlesize"
                ,"figure.titleweight"
                ,"base_mpf_style"]
            ,AttributeDic = {
                                 "style_name"           :	"Ehab_Staylo"
                                ,"base_mpl_style"       :	"dark_background"
                                ,"marketcolors"         :	"marketcolors"
                                ,"candle_Up"            :	"#14CE1C"
                                ,"candle_Down"          :	"#CE1414"
                                ,"edge_Up"              :	"#14CE1C"
                                ,"edge_Down"            :	"#CE1414"
                                ,"wick_Up"              :	"#ffffff"
                                ,"wick_Down"            :	"#ffffff"
                                ,"ohlc_Up"              :	"#ffffff"
                                ,"ohlc_Down"            :	"#ffffff"
                                ,"volume_Up"            :	"#14CE1C"
                                ,"volume_Down"          :	"#CE1414"
                                ,"vcdopcod"             :	 False	
                                ,"alpha"                :	 2.00	
                                ,"mavcolors"            :	 ['#ec009c','#78ff8f','#fcf120']	
                                ,"y_on_right"           :	 True	
                                ,"gridcolor"            :	 None	
                                ,"gridstyle"            :	 None	
                                ,"facecolor"            :	"Black"
                                ,"figcolor"             :	"gray"
                                ,"rc"                   :	"rc"
                                ,"axes.edgecolor"       :	"white"
                                ,"axes.linewidth"       :	 1.50	
                                ,"axes.labelsize"       :	"large"
                                ,"axes.labelweight"     :	"semibold"
                                ,"axes.grid"            :	 True	
                                ,"axes.grid.axis"       :	"both"
                                ,"axes.grid.which"      :	"major"
                                ,"grid.alpha"           :	 0.90	
                                ,"grid.color"           :	"#EBEE24"
                                ,"grid.linestyle"       :	":"	
                                ,"grid.linewidth"       :	 1.00	
                                ,"figure.facecolor"     :	"#0a0a0a"
                                ,"patch.linewidth"      :	 1.00	
                                ,"lines.linewidth"      :	 1.00	
                                ,"font.weight"          :	"medium"
                                ,"font.size"            :	 8.00	
                                ,"figure.titlesize"     :	"x-large"
                                ,"figure.titleweight"   :	"semibold"
                                ,"base_mpf_style"       :	"mike"
                                }


        ):
                        
            

            #--------------------------------------------------------------------------------------------------------------------------
            #   01- Create dataframe :- from "dfS" another dataframe    Help Link :- https://www.statology.org/pandas-create-dataframe-from-existing-dataframe/
            StyRow = dfS[[    
                             ("Attribute")
                            ,("Value")
                        ]].copy()                        
            
            #   02- Clean None Rows :- Help Link :- https://www.digitalocean.com/community/tutorials/pandas-dropna-drop-null-na-values-from-dataframe    ,   https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.dropna.html
            StyRow = StyRow.dropna(how='all')
            StyRow.reset_index(inplace=True, drop=True)


            #--------------------------------------------------------------------------------------------------------------------------
            #   03- Read Data from StyRow
            for AttributeIndex in range(len(AttributeList)):

                stylIndex = AttributeList[AttributeIndex]
                stylRead = (StyRow.Value[StyRow['Attribute'] == AttributeList[AttributeIndex]].values[0])
                AttributeDic[stylIndex] = stylRead

                # print(stylIndex , "=" , stylRead  ,"--->", type(stylRead) )
                # print(">>>>|||||||||||||||<<<<")


            #--------------------------------------------------------------------------------------------------------------------------
            #   03- Cerat Staylo Dict
                    #Set Style:
                    # link : C:\Users\lenovo\anaconda3\Lib\site-packages\mplfinance\_styledata\mike.py
            Staylo =     dict(  style_name    = AttributeDic["style_name"],
                                base_mpl_style= AttributeDic["base_mpl_style"], 
                                marketcolors  = {'candle'   : {'up':AttributeDic["candle_Up"]   ,'down':AttributeDic["candle_Down"] },
                                                  'edge'    : {'up':AttributeDic["edge_Up"]     ,'down':AttributeDic["edge_Down"]   },
                                                  'wick'    : {'up':AttributeDic["wick_Up"]     ,'down':AttributeDic["wick_Down"]   },
                                                  'ohlc'    : {'up':AttributeDic["ohlc_Up"]     ,'down':AttributeDic["ohlc_Down"]   },
                                                  'volume'  : {'up':AttributeDic["volume_Up"]   ,'down':AttributeDic["volume_Down"] },
                                                  'vcdopcod':       AttributeDic["vcdopcod"], # Volume Color Depends On Price Change On Day
                                                  'alpha'   :       AttributeDic["alpha"],
                                                },
                                mavcolors     = AttributeDic["mavcolors"].split(","),  #    Help Link:- https://www.w3schools.com/python/ref_string_split.asp
                                y_on_right    = AttributeDic["y_on_right"],
                                gridcolor     = None if AttributeDic["gridcolor"] == 'None' else AttributeDic["gridcolor"], # None AttributeDic["gridcolor"],     Help Link:-   https://stackoverflow.com/questions/26481774/pythonic-way-to-convert-the-string-none-to-a-proper-none
                                gridstyle     = None if AttributeDic["gridstyle"] == 'None' else AttributeDic["gridstyle"], # None AttributeDic["gridstyle"],     Help Link:-   https://stackoverflow.com/questions/26481774/pythonic-way-to-convert-the-string-none-to-a-proper-none    
                                facecolor     = AttributeDic["facecolor"],
                                figcolor      = AttributeDic["figcolor"],
                                rc            = [ ('axes.edgecolor'     , AttributeDic["axes.edgecolor"]        ),
                                                  ('axes.linewidth'     , AttributeDic["axes.linewidth"]        ),
                                                  ('axes.labelsize'     , AttributeDic["axes.labelsize"]        ),
                                                  ('axes.labelweight'   , AttributeDic["axes.labelweight"]      ),
                                                  ('axes.grid'          , AttributeDic["axes.grid"]             ),
                                                  ('axes.grid.axis'     , AttributeDic["axes.grid.axis"]        ),
                                                  ('axes.grid.which'    , AttributeDic["axes.grid.which"]       ),
                                                  ('grid.alpha'         , AttributeDic["grid.alpha"]            ),
                                                  ('grid.color'         , AttributeDic["grid.color"]            ),
                                                  ('grid.linestyle'     , AttributeDic["grid.linestyle"]        ),
                                                  ('grid.linewidth'     , AttributeDic["grid.linewidth"]        ),
                                                  ('figure.facecolor'   , AttributeDic["figure.facecolor"]      ),
                                                  ('patch.linewidth'    , AttributeDic["patch.linewidth"]       ),
                                                  ('lines.linewidth'    , AttributeDic["lines.linewidth"]       ),
                                                  ('font.weight'        , AttributeDic["font.weight"]           ),
                                                  ('font.size'          , AttributeDic["font.size"]             ),
                                                  ('figure.titlesize'   , AttributeDic["figure.titlesize"]      ),
                                                  ('figure.titleweight' , AttributeDic["figure.titleweight"]    ),
                                                ],
                                base_mpf_style= AttributeDic["base_mpf_style"]
                              )

            
            #--------------------------------------------------------------------------------------------------------------------------            
            #   07- Return OutPut  "Staylo" :-

            if(False):
                print(">>>>**************<<<<")
                print(StyRow.loc[StyRow['Attribute'] == "base_mpl_style"])
                print(">>>>**************<<<<")

                print(">>>>$$$$$$$$$$$$$$<<<<")
                print(dfS)
                print(StyRow)
                print(">>>>$$$$$$$$$$$$$$<<<<")

                print(">>>>*******1*******<<<<")
                stylYest = (StyRow.Value[StyRow['Attribute'] == "gridcolor"].values[0])
                print(stylYest, type(stylYest), type(None))
                print(">>>>*******2*******<<<<")
                print(AttributeDic)
                print(">>>>*******3*******<<<<")
                print(Staylo)
                print(">>>>**************<<<<")

            return Staylo
            #----------------------------------------------------------------------------------------------------------------------------------------------     



def fGetStyle2(                 
             dfS
            ,ShowOutPut = False
            ,AttributeIndex=0
            ,AttributeList =[
                 "style_name"
                ,"base_mpl_style"
                ,"marketcolors"
                ,"candle_Up"
                ,"candle_Down"
                ,"edge_Up"
                ,"edge_Down"
                ,"wick_Up"
                ,"wick_Down"
                ,"ohlc_Up"
                ,"ohlc_Down"
                ,"volume_Up"
                ,"volume_Down"
                ,"vcdopcod"
                ,"alpha"
                ,"mavcolors"
                ,"y_on_right"
                ,"gridcolor"
                ,"gridstyle"
                ,"facecolor"
                ,"figcolor"
                ,"rc"
                ,"axes.edgecolor"
                ,"axes.linewidth"
                ,"axes.labelsize"
                ,"axes.labelweight"
                ,"axes.grid"
                ,"axes.grid.axis"
                ,"axes.grid.which"
                ,"grid.alpha"
                ,"grid.color"
                ,"grid.linestyle"
                ,"grid.linewidth"
                ,"figure.facecolor"
                ,"patch.linewidth"
                ,"lines.linewidth"
                ,"font.weight"
                ,"font.size"
                ,"figure.titlesize"
                ,"figure.titleweight"
                ,"base_mpf_style"]
            ,AttributeDic = {
                                 "style_name"           :	"Ehab_Staylo"
                                ,"base_mpl_style"       :	"dark_background"
                                ,"marketcolors"         :	"marketcolors"
                                ,"candle_Up"            :	"#14CE1C"
                                ,"candle_Down"          :	"#CE1414"
                                ,"edge_Up"              :	"#14CE1C"
                                ,"edge_Down"            :	"#CE1414"
                                ,"wick_Up"              :	"#ffffff"
                                ,"wick_Down"            :	"#ffffff"
                                ,"ohlc_Up"              :	"#ffffff"
                                ,"ohlc_Down"            :	"#ffffff"
                                ,"volume_Up"            :	"#14CE1C"
                                ,"volume_Down"          :	"#CE1414"
                                ,"vcdopcod"             :	 False	
                                ,"alpha"                :	 2.00	
                                ,"mavcolors"            :	 ['#ec009c','#78ff8f','#fcf120']	
                                ,"y_on_right"           :	 True	
                                ,"gridcolor"            :	 None	
                                ,"gridstyle"            :	 None	
                                ,"facecolor"            :	"Black"
                                ,"figcolor"             :	"gray"
                                ,"rc"                   :	"rc"
                                ,"axes.edgecolor"       :	"white"
                                ,"axes.linewidth"       :	 1.50	
                                ,"axes.labelsize"       :	"large"
                                ,"axes.labelweight"     :	"semibold"
                                ,"axes.grid"            :	 True	
                                ,"axes.grid.axis"       :	"both"
                                ,"axes.grid.which"      :	"major"
                                ,"grid.alpha"           :	 0.90	
                                ,"grid.color"           :	"#EBEE24"
                                ,"grid.linestyle"       :	":"	
                                ,"grid.linewidth"       :	 1.00	
                                ,"figure.facecolor"     :	"#0a0a0a"
                                ,"patch.linewidth"      :	 1.00	
                                ,"lines.linewidth"      :	 1.00	
                                ,"font.weight"          :	"medium"
                                ,"font.size"            :	 8.00	
                                ,"figure.titlesize"     :	"x-large"
                                ,"figure.titleweight"   :	"semibold"
                                ,"base_mpf_style"       :	"mike"
                                }


        ):
                        
            

            #--------------------------------------------------------------------------------------------------------------------------
            #   01- Create dataframe :- from "dfS" another dataframe    Help Link :- https://www.statology.org/pandas-create-dataframe-from-existing-dataframe/
            StyRow = dfS[[    
                             ("Attribute")
                            ,("Value")
                        ]].copy()                        
            
            #   02- Clean None Rows :- Help Link :- https://www.digitalocean.com/community/tutorials/pandas-dropna-drop-null-na-values-from-dataframe    ,   https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.dropna.html
            StyRow = StyRow.dropna(how='all')
            StyRow.reset_index(inplace=True, drop=True)


            #--------------------------------------------------------------------------------------------------------------------------
            #   03- Read Data from StyRow
            timeScope = (StyRow.Value[StyRow['Attribute'] == 'timeScope'].values[0]).split(",")
            if(ShowOutPut): print(timeScope)

            for AttributeIndex in range(len(AttributeList)):
                stylIndex = AttributeList[AttributeIndex]
                stylRead = (StyRow.Value[StyRow['Attribute'] == AttributeList[AttributeIndex]].values[0])
                AttributeDic[stylIndex] = stylRead
                if(ShowOutPut):
                    print(stylIndex , "=" , stylRead  ,"--->", type(stylRead) )
                    print(">>>>|||||||||||||||<<<<")


            #--------------------------------------------------------------------------------------------------------------------------
            #   03- Cerat Staylo Dict
                    #Set Style:
                    # link : C:\Users\lenovo\anaconda3\Lib\site-packages\mplfinance\_styledata\mike.py
            Staylo =     dict(  style_name    = AttributeDic["style_name"],
                                base_mpl_style= AttributeDic["base_mpl_style"], 
                                marketcolors  = {'candle'   : {'up':AttributeDic["candle_Up"]   ,'down':AttributeDic["candle_Down"] },
                                                  'edge'    : {'up':AttributeDic["edge_Up"]     ,'down':AttributeDic["edge_Down"]   },
                                                  'wick'    : {'up':AttributeDic["wick_Up"]     ,'down':AttributeDic["wick_Down"]   },
                                                  'ohlc'    : {'up':AttributeDic["ohlc_Up"]     ,'down':AttributeDic["ohlc_Down"]   },
                                                  'volume'  : {'up':AttributeDic["volume_Up"]   ,'down':AttributeDic["volume_Down"] },
                                                  'vcdopcod':       AttributeDic["vcdopcod"], # Volume Color Depends On Price Change On Day
                                                  'alpha'   :       AttributeDic["alpha"],
                                                },
                                mavcolors     = AttributeDic["mavcolors"].split(","),  #    Help Link:- https://www.w3schools.com/python/ref_string_split.asp
                                y_on_right    = AttributeDic["y_on_right"],
                                gridcolor     = None if AttributeDic["gridcolor"] == 'None' else AttributeDic["gridcolor"], # None AttributeDic["gridcolor"],     Help Link:-   https://stackoverflow.com/questions/26481774/pythonic-way-to-convert-the-string-none-to-a-proper-none
                                gridstyle     = None if AttributeDic["gridstyle"] == 'None' else AttributeDic["gridstyle"], # None AttributeDic["gridstyle"],     Help Link:-   https://stackoverflow.com/questions/26481774/pythonic-way-to-convert-the-string-none-to-a-proper-none    
                                facecolor     = AttributeDic["facecolor"],
                                figcolor      = AttributeDic["figcolor"],
                                rc            = [ ('axes.edgecolor'     , AttributeDic["axes.edgecolor"]        ),
                                                  ('axes.linewidth'     , AttributeDic["axes.linewidth"]        ),
                                                  ('axes.labelsize'     , AttributeDic["axes.labelsize"]        ),
                                                  ('axes.labelweight'   , AttributeDic["axes.labelweight"]      ),
                                                  ('axes.grid'          , AttributeDic["axes.grid"]             ),
                                                  ('axes.grid.axis'     , AttributeDic["axes.grid.axis"]        ),
                                                  ('axes.grid.which'    , AttributeDic["axes.grid.which"]       ),
                                                  ('grid.alpha'         , AttributeDic["grid.alpha"]            ),
                                                  ('grid.color'         , AttributeDic["grid.color"]            ),
                                                  ('grid.linestyle'     , AttributeDic["grid.linestyle"]        ),
                                                  ('grid.linewidth'     , AttributeDic["grid.linewidth"]        ),
                                                  ('figure.facecolor'   , AttributeDic["figure.facecolor"]      ),
                                                  ('patch.linewidth'    , AttributeDic["patch.linewidth"]       ),
                                                  ('lines.linewidth'    , AttributeDic["lines.linewidth"]       ),
                                                  ('font.weight'        , AttributeDic["font.weight"]           ),
                                                  ('font.size'          , AttributeDic["font.size"]             ),
                                                  ('figure.titlesize'   , AttributeDic["figure.titlesize"]      ),
                                                  ('figure.titleweight' , AttributeDic["figure.titleweight"]    ),
                                                ],
                                base_mpf_style= AttributeDic["base_mpf_style"]
                              )

            
            #--------------------------------------------------------------------------------------------------------------------------            
            #   07- Return OutPut  "Staylo" :-

            if(ShowOutPut):
                print(">>>>**************<<<<")
                print(StyRow.loc[StyRow['Attribute'] == "base_mpl_style"])
                print(">>>>**************<<<<")

                print(">>>>$$$$$$$$$$$$$$<<<<")
                print(dfS)
                print(StyRow)
                print(">>>>$$$$$$$$$$$$$$<<<<")

                print(">>>>*******1*******<<<<")
                stylYest = (StyRow.Value[StyRow['Attribute'] == "gridcolor"].values[0])
                print(stylYest, type(stylYest), type(None))
                print(">>>>*******2*******<<<<")
                print(AttributeDic)
                print(">>>>*******3*******<<<<")
                print(Staylo)
                print(">>>>**************<<<<")

            return Staylo,timeScope
            #----------------------------------------------------------------------------------------------------------------------------------------------     


#--------------------------------------------------------------------------------------------------------------------------












